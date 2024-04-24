# !/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
import time

# ========== 數值記錄 ==================================================
# ->高度控制器
R_TARGET_ALT = []                       # 記錄目標高
R_ACTUAL_ALT_CTRL = []                  # 記錄實際高
R_TARGET_THRUST = []                    # 記錄目標推力
R_ALT_ERROR = []                        # 記錄高度誤差
R_ALT_GP = []                           # 記錄P增益
R_ALT_GI = []                           # 記錄I增益
R_ALT_GD = []                           # 記錄D增益

# ->位置控制器
# 位置
R_POS_BLOCK_X = []                      # 記錄X座標
R_POS_BLOCK_Y = []                      # 記錄Y座標
R_POS_BLOCK_AREA = []                   # 記錄面積變化
R_POS_RECORD_SCALE = 1000               # 面積縮放比
R_POS_X_ERROR = []                      # 記錄誤差值
R_POS_Y_ERROR = []                      # 記錄誤差值
R_POS_X_GP = []                         # 記錄P增益
R_POS_X_GI = []                         # 記錄I增益
R_POS_X_GD = []                         # 記錄D增益
R_POS_Y_GP = []                         # 記錄P增益
R_POS_Y_GI = []                         # 記錄I增益
R_POS_Y_GD = []                         # 記錄D增益

# 速度
R_POS_TARGET_ROLL = []                  # 記錄目標滾轉角
R_POS_TARGET_PITCH = []                 # 記錄目標俯仰角
R_POS_ACTUAL_ROLL = []                  # 記錄實際滾轉角
R_POS_ACTUAL_PITCH = []                 # 記錄實際俯仰角
R_POS_VEL_X_CMD = []                    # 記錄X目標速度
R_POS_VEL_Y_CMD = []                    # 記錄Y目標速度
R_POS_VEL_X_ACT = []                    # 記錄X實際速度
R_POS_VEL_Y_ACT = []                    # 記錄Y實際速度
R_POS_VEL_X_ERROR = []                  # 記錄誤差值
R_POS_VEL_Y_ERROR = []                  # 記錄誤差值
R_POS_VEL_X_GP = []                     # 記錄P增益
R_POS_VEL_X_GI = []                     # 記錄I增益
R_POS_VEL_X_GD = []                     # 記錄D增益
R_POS_VEL_Y_GP = []                     # 記錄P增益
R_POS_VEL_Y_GI = []                     # 記錄I增益
R_POS_VEL_Y_GD = []                     # 記錄D增益

# ->高度感測器
R_ACTUAL_ALT = []                       # 記錄加權值
R_ACTUAL_ALT_SONAR = []                 # 記錄超音波高度計值
R_ACTUAL_ALT_BAROM = []                 # 記錄氣壓高度計值

# ->超音波高度計
R_ACTUAL_ALT_SONAR_RAW = []             # 記錄原始值
R_ACTUAL_ALT_SONAR_OFFSET = []          # 記錄補正值
R_ACTUAL_ALT_SONAR_AVE = []             # 記錄平均值
R_ACTUAL_ALT_SONAR_KF = []              # 記錄KF值

# ->氣壓高度計
R_ACTUAL_ALT_IMU_RAW = []               # 記錄原始值
R_ACTUAL_ALT_IMU_OFFSET = []            # 記錄補正值
R_ACTUAL_ALT_IMU_AVE = []               # 記錄平均值
R_ACTUAL_ALT_IMU_KF = []                # 記錄KF值


def data_sheet():
    # 建立活頁簿
    db = Workbook()

    # 建立分頁
    ds1 = db.create_sheet("ALT Controller")
    ds2 = db.create_sheet("POS Controller")
    ds2b = db.create_sheet("POS-VEL Controller")
    ds3 = db.create_sheet("HDG Controller")
    ds4 = db.create_sheet("ALT Sensor")
    ds5 = db.create_sheet("Parameter")

    # 高度控制器(先列在行)
    ds1.cell(1, 1, '高度控制器')
    ds1.cell(2, 1, '實際高度')
    for x in range(len(R_ACTUAL_ALT_CTRL)):
        ds1.cell(x + 3, 1, R_ACTUAL_ALT_CTRL[x])

    ds1.cell(2, 2, '目標高度')
    for x in range(len(R_TARGET_ALT)):
        ds1.cell(x + 3, 2, R_TARGET_ALT[x])

    ds1.cell(2, 3, '爬升率')
    for x in range(len(R_TARGET_THRUST)):
        ds1.cell(x + 3, 3, R_TARGET_THRUST[x])

    # 高度控制器PID響應
    ds1.cell(1, 5, '高度控制器PID響應')
    ds1.cell(2, 5, 'Error')
    for x in range(len(R_ALT_ERROR)):
        ds1.cell(x + 3, 5, R_ALT_ERROR[x])

    ds1.cell(2, 6, 'Gain P')
    for x in range(len(R_ALT_GP)):
        ds1.cell(x + 3, 6, R_ALT_GP[x])

    ds1.cell(2, 7, 'Gain I')
    for x in range(len(R_ALT_GI)):
        ds1.cell(x + 3, 7, R_ALT_GI[x])

    ds1.cell(2, 8, 'Gain D')
    for x in range(len(R_ALT_GD)):
        ds1.cell(x + 3, 8, R_ALT_GD[x])

    # 位置控制器-第二頁
    # 位置控制器X響應圖
    ds2.cell(1, 1, '位置控制器X響應圖')
    ds2.cell(2, 1, 'X位置')
    for x in range(len(R_POS_BLOCK_X)):
        ds2.cell(x + 3, 1, R_POS_BLOCK_X[x])

    ds2.cell(2, 2, 'X速度輸出')
    for x in range(len(R_POS_VEL_X_CMD)):
        ds2.cell(x + 3, 2, R_POS_VEL_X_CMD[x])

    ds2.cell(2, 3, '面積大小')
    for x in range(len(R_POS_BLOCK_AREA)):
        ds2.cell(x + 3, 3, R_POS_BLOCK_AREA[x])

    # 位置控制器Y響應圖
    ds2.cell(1, 5, '位置控制器Y響應圖')
    ds2.cell(2, 5, 'Y位置')
    for x in range(len(R_POS_BLOCK_Y)):
        ds2.cell(x + 3, 5, R_POS_BLOCK_Y[x])

    ds2.cell(2, 6, 'Y速度輸出')
    for x in range(len(R_POS_VEL_Y_CMD)):
        ds2.cell(x + 3, 6, R_POS_VEL_Y_CMD[x])

    ds2.cell(2, 7, '面積大小')
    for x in range(len(R_POS_BLOCK_AREA)):
        ds2.cell(x + 3, 7, R_POS_BLOCK_AREA[x])

    # 軌跡記錄
    ds2.cell(1, 9, '位置控制器Y響應圖')
    ds2.cell(2, 9, 'X位置')
    for x in range(len(R_POS_BLOCK_X)):
        ds2.cell(x + 3, 9, R_POS_BLOCK_X[x])

    ds2.cell(2, 10, 'Y位置')
    for x in range(len(R_POS_BLOCK_Y)):
        ds2.cell(x + 3, 10, R_POS_BLOCK_Y[x])

    # 位置控制器X-PID響應圖
    ds2.cell(1, 12, '位置控制器X-PID響應圖')
    ds2.cell(2, 12, 'Error')
    for x in range(len(R_POS_X_ERROR)):
        ds2.cell(x + 3, 12, R_POS_X_ERROR[x])

    ds2.cell(2, 13, 'Gain P')
    for x in range(len(R_POS_X_GP)):
        ds2.cell(x + 3, 13, R_POS_X_GP[x])

    ds2.cell(2, 14, 'Gain I')
    for x in range(len(R_POS_X_GI)):
        ds2.cell(x + 3, 14, R_POS_X_GI[x])

    ds2.cell(2, 15, 'Gain D')
    for x in range(len(R_POS_X_GD)):
        ds2.cell(x + 3, 15, R_POS_X_GD[x])

    # 位置控制器Y-PID響應圖
    ds2.cell(1, 17, '位置控制器Y-PID響應圖')
    ds2.cell(2, 17, 'Error')
    for x in range(len(R_POS_Y_ERROR)):
        ds2.cell(x + 3, 17, R_POS_Y_ERROR[x])

    ds2.cell(2, 18, 'Gain P')
    for x in range(len(R_POS_Y_GP)):
        ds2.cell(x + 3, 18, R_POS_Y_GP[x])

    ds2.cell(2, 19, 'Gain I')
    for x in range(len(R_POS_Y_GI)):
        ds2.cell(x + 3, 19, R_POS_Y_GI[x])

    ds2.cell(2, 20, 'Gain D')
    for x in range(len(R_POS_Y_GD)):
        ds2.cell(x + 3, 20, R_POS_Y_GD[x])

    # 速度控制器-第三頁
    # 速度控制器X響應圖
    ds2b.cell(1, 1, '速度控制器X響應圖')
    ds2b.cell(2, 1, 'X目標速度')
    for x in range(len(R_POS_VEL_X_CMD)):
        ds2b.cell(x + 3, 1, R_POS_VEL_X_CMD[x])

    ds2b.cell(2, 2, 'X實際速度')
    for x in range(len(R_POS_VEL_X_ACT)):
        ds2b.cell(x + 3, 2, R_POS_VEL_X_ACT[x])

    ds2b.cell(2, 3, '目標Roll')
    for x in range(len(R_POS_TARGET_ROLL)):
        ds2b.cell(x + 3, 3, R_POS_TARGET_ROLL[x])

    ds2b.cell(2, 4, '實際Roll')
    for x in range(len(R_POS_ACTUAL_ROLL)):
        ds2b.cell(x + 3, 4, R_POS_ACTUAL_ROLL[x])

    # 速度控制器Y響應圖
    ds2b.cell(1, 6, '速度控制器Y響應圖')
    ds2b.cell(2, 6, 'Y目標速度')
    for x in range(len(R_POS_VEL_Y_CMD)):
        ds2b.cell(x + 3, 6, R_POS_VEL_Y_CMD[x])

    ds2b.cell(2, 7, 'Y實際速度')
    for x in range(len(R_POS_VEL_Y_ACT)):
        ds2b.cell(x + 3, 7, R_POS_VEL_Y_ACT[x])

    ds2b.cell(2, 8, '目標Pitch')
    for x in range(len(R_POS_TARGET_PITCH)):
        ds2b.cell(x + 3, 8, R_POS_TARGET_PITCH[x])

    ds2b.cell(2, 9, '實際Pitch')
    for x in range(len(R_POS_ACTUAL_PITCH)):
        ds2b.cell(x + 3, 9, R_POS_ACTUAL_PITCH[x])

    # 速度控制器X-PID響應圖
    ds2b.cell(1, 11, '速度控制器X-PID響應圖')
    ds2b.cell(2, 11, 'Error')
    for x in range(len(R_POS_VEL_X_ERROR)):
        ds2b.cell(x + 3, 11, R_POS_VEL_X_ERROR[x])

    ds2b.cell(2, 12, 'Gain P')
    for x in range(len(R_POS_VEL_X_GP)):
        ds2b.cell(x + 3, 12, R_POS_VEL_X_GP[x])

    ds2b.cell(2, 13, 'Gain I')
    for x in range(len(R_POS_VEL_X_GI)):
        ds2b.cell(x + 3, 13, R_POS_VEL_X_GI[x])

    ds2b.cell(2, 14, 'Gain D')
    for x in range(len(R_POS_VEL_X_GD)):
        ds2b.cell(x + 3, 14, R_POS_VEL_X_GD[x])

    # 速度控制器Y-PID響應圖
    ds2b.cell(1, 16, '速度控制器Y-PID響應圖')
    ds2b.cell(2, 16, 'Error')
    for x in range(len(R_POS_VEL_Y_ERROR)):
        ds2b.cell(x + 3, 16, R_POS_VEL_Y_ERROR[x])

    ds2b.cell(2, 17, 'Gain P')
    for x in range(len(R_POS_VEL_Y_GP)):
        ds2b.cell(x + 3, 17, R_POS_VEL_Y_GP[x])

    ds2b.cell(2, 18, 'Gain I')
    for x in range(len(R_POS_VEL_Y_GI)):
        ds2b.cell(x + 3, 18, R_POS_VEL_Y_GI[x])

    ds2b.cell(2, 19, 'Gain D')
    for x in range(len(R_POS_VEL_Y_GD)):
        ds2b.cell(x + 3, 19, R_POS_VEL_Y_GD[x])

    # 第五頁-感測器
    # 高度感測器數值處理圖
    ds4.cell(1, 1, '高度感測器數值處理圖')
    ds4.cell(2, 1, '實際高')
    for x in range(len(R_ACTUAL_ALT)):
        ds4.cell(x + 3, 1, R_ACTUAL_ALT[x])

    ds4.cell(2, 2, '超音波')
    for x in range(len(R_ACTUAL_ALT_SONAR)):
        ds4.cell(x + 3, 2, R_ACTUAL_ALT_SONAR[x])

    ds4.cell(2, 3, 'IMU')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds4.cell(x + 3, 3, R_ACTUAL_ALT_IMU_KF[x])

    # 超音波高度計數值處理圖
    ds4.cell(1, 5, '超音波高度計數值處理圖')
    ds4.cell(2, 5, '原始值')
    for x in range(len(R_ACTUAL_ALT_SONAR_RAW)):
        ds4.cell(x + 3, 5, R_ACTUAL_ALT_SONAR_RAW[x])

    ds4.cell(2, 6, '補正後')
    for x in range(len(R_ACTUAL_ALT_SONAR_OFFSET)):
        ds4.cell(x + 3, 6, R_ACTUAL_ALT_SONAR_OFFSET[x])

    ds4.cell(2, 7, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_AVE)):
        ds4.cell(x + 3, 7, R_ACTUAL_ALT_SONAR_AVE[x])

    ds4.cell(2, 8, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_KF)):
        ds4.cell(x + 3, 8, R_ACTUAL_ALT_SONAR_KF[x])

    # IMU高度計數值處理圖
    ds4.cell(1, 10, 'IMU高度計數值處理圖')
    ds4.cell(2, 10, '原始值')
    for x in range(len(R_ACTUAL_ALT_IMU_RAW)):
        ds4.cell(x + 3, 10, R_ACTUAL_ALT_IMU_RAW[x])

    ds4.cell(2, 11, '補正後')
    for x in range(len(R_ACTUAL_ALT_IMU_OFFSET)):
        ds4.cell(x + 3, 11, R_ACTUAL_ALT_IMU_OFFSET[x])

    ds4.cell(2, 12, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_AVE)):
        ds4.cell(x + 3, 12, R_ACTUAL_ALT_IMU_AVE[x])

    ds4.cell(2, 13, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds4.cell(x + 3, 13, R_ACTUAL_ALT_IMU_KF[x])

    # 依照時間存檔
    t_save = time.localtime()
    db.save('Flight Data %d_%d_%d_%d_%d_%d.xlsx' %
            (t_save[0], t_save[1], t_save[2], t_save[3], t_save[4], t_save[5]))

data_sheet()






