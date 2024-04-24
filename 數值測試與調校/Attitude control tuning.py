# 姿態控制器角度調校

# 載入模組
from __future__ import print_function
from dronekit import connect, VehicleMode
from openpyxl import Workbook
# import dronekit_sitl
import matplotlib.pyplot as plt
import RPi.GPIO as GPIO
import numpy as np
import threading
import time
import math

# ====================================================================
# ========== 設定全域變數 ==============================================
# ====================================================================
# 加註*字為可調整之參數

# ========== UAV通訊參數 ==============================================
# 樹莓派Serial連線參數
CONNECT_PORT = '/dev/serial0'   # 傳輸口位址
CONNECT_BAUD = 921600           # 傳輸鮑率

# MacBook 3DR連線參數
# CONNECT_PORT = '/dev/tty.SLAB_USBtoUART'
# CONNECT_BAUD = 57600

# 軟體模擬連線參數
# sitl = dronekit_sitl.start_default()
# CONNECT_PORT = sitl.connection_string()
# CONNECT_BAUD = 921600

print('<啟動程序> 載入全域變數設定')

# ========== 旗標與變數 ================================================
# ->旗標
STATE_THREAD = False            # 多線程旗標
STATE_UAV = True                # 無人機連線狀態
STATE_ALT = False               # 高度控制器旗標
STATE_POS = False               # 位置控制器旗標
STATE_VEL = False               # 位置-速度控制器旗標
STATE_HDG = False               # 航向控制器旗標
STATE_SPD = False               # 飛行速度控制器旗標
STATE_SENSOR_SPD = False        # 飛行速度感測器旗標
STATE_SENSOR_ALT = False        # 高度感測器旗標
STATE_SENSOR_SONAR = False      # 超音波高度計旗標
STATE_SENSOR_CAM = False        # 影像感測器旗標
STATE_CMD = False               # 命令傳輸程式旗標
STATE_CAM = False               # 影像擷取程式旗標
STATE_BOX = False               # 貨物艙門狀態旗標
MODE = ''                       # 模式旗標(POS:定位模式，LINE:循跡模式)
MISSION_COLOR = ''              # 任務顏色(blue or green)
POS_COLOR = ''                  # 定位的色塊顏色

# ========== 控制器參數設定 =============================================
# ->迴圈頻率(預設20Hz)
FREQ_ALT = 20                   # 高度控制器頻率(不快於高度感測器頻率)
FREQ_CMD = 20                   # 命令傳輸頻率(不快於控制器頻率)
FREQ_SENSOR_ALT = 20            # 高度感測器頻率(不快於超音波感測器頻率)
FREQ_SENSOR_SONAR = 20          # 超音波感測器頻率(5~50Hz，越高雜訊越多)
FREQ_FR = 5

# ->目標值(命令值)
TARGET_ALT = 0                  # 目標高度，主程式輸入於高度控制器(m)
TARGET_ROLL = 0                 # 目標滾轉角(degree)
TARGET_PITCH = 0                # 目標俯仰角(degree)
TARGET_YAW = 0                  # 目標偏航角(degree)，不使用
TARGET_THRUST = 0               # 目標推力，表示升降速度(0.5為中點)
TARGET_YAW_RATE = 0             # 偏航角速度(degree/s)，需搭配控制器控制角度
TARGET_VEL = [0, 0]             # 目標移動速度(px/s，相機座標)
TARGET_SPD = 0                  # 目標飛行速度

# ->實際值(量測值)
ACTUAL_POS = [0, 0]             # 色塊實際位置(px)
ACTUAL_VEL = [0, 0]             # 色塊實際速度(px/s)
ACTUAL_SPD = 0                  # 實際飛行速度
ACTUAL_AREA = 0                 # 色塊面積(px^2)
ACTUAL_HDG = 0                  # 與線夾角(degree)
ACTUAL_ALT = 0                  # 實際高度(m)
ACTUAL_ALT_SONAR = 0            # 超音波高度計(m)
ACTUAL_ALT_IMU = 0              # IMU氣壓高度計(m)

# ->高度控制器參數
ALT_KP = 0.25                   # P增益值*
ALT_KI = 0                      # I增益值*
ALT_KD = 0.2                    # D增益值*
ALT_SPEED = 0.3                 # 昇降速率(m/s，與飛控參數一致)*

# ->飛行程序參數
TAKEOFF_SETTLE = 95             # 安定高度百分比(%，高度確認門檻)*
LANDING_CUTOFF = 0.05           # 油門關閉高度(m，停止油門輸出)*
FLIGHT_SPEED = 0.3              # 飛行速度(m/s)
TURN_SETTLE = 95                # 安定轉角百分比(%，停止轉向輸出)*
TURN_YAW_RATE = 20              # 轉角速度(degree/s)*

# ========== 高度感測器參數 =============================================
# ->高度感測器整合參數
IMU_SAMPLE_NUM_AVE = 10         # 均值濾波取樣次數*
WEIGHTED_SONAR = 5              # 超音波高度計權重*
WEIGHTED_IMU = 5                # IMU高度計權重*

# ->超音波高度計參數
SONAR_AIR_TEMP = 28             # 氣溫(攝氏)*
SONAR_TRIG_PIN = 2              # TRIG輸出腳位
SONAR_ECHO_PIN = 3              # ECHO輸入腳位
SONAR_TIMEOUT = 0.015           # 無回應超時限制(秒，過低時遠距離偵測將會變成超時)*
SONAR_OFFSET = 0                # 高度校正偏移量(m，向上為正)
SONAR_SAMPLE_NUM_CAL = 10       # 高度校正取樣次數*
SONAR_SAMPLE_NUM_AVE = 10       # 均值濾波取樣次數*
SONAR_KF_INPUT_MAX = 3          # 卡爾曼濾波器輸入值限制(高度，m)*
SONAR_KF_Q = 5                  # 卡爾曼濾波器滑順度(越小反應快，越不平；越大反應慢，但較平滑)*
SONAR_ERR_COUNT = 0             # 量測超時次數記錄
SONAR_FILTER_ALT = 1            # 突波濾波作用起始高度(m)
SONAR_MAX_CHANGE = 0.5          # 突波容許值(m)

# ->IMU高度計參數
IMU_OFFSET = 0                  # 高度校正偏移量(m，向上為正)
IMU_SAMPLE_NUM_CAL = 10         # 高度校正取樣次數*
IMU_KF_INPUT_MAX = 3            # 卡爾曼濾波器輸入值限制(高度，m)*
IMU_KF_Q = 5                    # 卡爾曼濾波器滑順度(越小反應快，越不平；越大反應慢，但較平滑)*

# ========== 數值記錄 ==================================================
# ->高度控制器
R_TARGET_ALT = []                       # 記錄目標高
R_ACTUAL_ALT_CTRL = []                  # 記錄實際高
R_TARGET_THRUST = []                    # 記錄目標推力
R_ALT_ERROR = []                        # 記錄高度誤差
R_ALT_GP = []                           # 記錄P增益
R_ALT_GI = []                           # 記錄I增益
R_ALT_GD = []                           # 記錄D增益

# ->高度感測器
R_ACTUAL_ALT = []                       # 記錄加權值
R_ACTUAL_ALT_SONAR = []                 # 記錄超音波高度計值
R_ACTUAL_ALT_BAROM = []                 # 記錄氣壓高度計值

# ->超音波高度計
R_ACTUAL_ALT_SONAR_RAW = []             # 記錄原始值
R_ACTUAL_ALT_SONAR_OFFSET = []          # 記錄補正值
R_ACTUAL_ALT_SONAR_AVE = []             # 記錄平均值
R_ACTUAL_ALT_SONAR_KF = []              # 記錄KF值

# ->氣壓高度計
R_ACTUAL_ALT_IMU_RAW = []               # 記錄原始值
R_ACTUAL_ALT_IMU_OFFSET = []            # 記錄補正值
R_ACTUAL_ALT_IMU_AVE = []               # 記錄平均值
R_ACTUAL_ALT_IMU_KF = []                # 記錄KF值

# ->飛行紀錄
R_ACTUAL_ROLL = []
R_ACTUAL_PITCH = []
R_ACTUAL_YAW = []
R_ACTUAL_ROLL_RATE = []
R_ACTUAL_PITCH_RATE = []
R_ACTUAL_YAW_RATE = []
R_ACTUAL_ROLL_ACC = []
R_ACTUAL_PITCH_ACC = []
R_ACTUAL_YAW_ACC = []
R_VEL_X = []
R_VEL_Y = []
R_VEL_Z = []
R_ACC_X = []
R_ACC_Y = []
R_ACC_Z = []


print('<啟動程序> 載入全域變數設定完畢')


# ========== 與UAV執行連線 =============================================
print('<啟動程序> 與UAV連線中，請稍候...')
# 樹莓派連線指令
uav = connect(CONNECT_PORT, wait_ready=True, baud=CONNECT_BAUD)
# 電腦連線指令
# uav = connect(CONNECT_PORT, wait_ready=True, baud=CONNECT_BAUD)
print('<啟動程序> UAV已連線')


# ====================================================================
# ========== 背景主函式函式區塊 =========================================
# ====================================================================
# 本區為主要運作的函式，以控制器為主，包含高度、位置、航向、雙高度感測器、
# 相機、命令傳送，多線程執行時能達到協作的處理功能
# (但仍受每個迴圈處理時間不一影響，為非同步處理)

# ========== 高度控制器 ================================================
# 說明：
# 為PID控制器型態，於任務期間全程作用，輸入目標高度後，依據高度感測器執行誤差PID增益，
# 最終輸出成爬升速率型態(符合MavLink指令)，設有輸出範圍限制器、中點飽和區間功能
def altitude_controller():
    global TARGET_THRUST, STATE_ALT
    # PID變數設定
    int_error = 0
    previous_error = 0
    previous_time = time.time()

    # 等待感測器啟動
    if not STATE_SENSOR_ALT:
        print('<啟動程序> 高度控制器：等待高度感測器啟動')
    while not STATE_SENSOR_ALT:
        time.sleep(0.1)

    # 回報高度控制器啟動
    STATE_ALT = True
    print('<啟動程序> 高度控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 計算誤差值
        error = TARGET_ALT - ACTUAL_ALT

        # 誤差值微積分計算
        current_time = time.time()
        delta_error = error - previous_error
        delta_time = current_time - previous_time
        int_error += (error * delta_time)
        derivative = delta_error / delta_time

        # 誤差增益處理
        gain_p = error * ALT_KP
        gain_i = int_error * ALT_KI
        gain_d = derivative * ALT_KD
        speed_out = gain_p + gain_i + gain_d

        # 將速度以比例輸出
        thrust_out = speed_out / ALT_SPEED

        # 補正至中點
        thrust_out += 0.5

        # 輸出限制(0下降max ~ 1上升max)
        if thrust_out > 1:
            thrust_out = 1
        if thrust_out < 0:
            thrust_out = 0

        # 更新推力值至全域變數
        TARGET_THRUST = thrust_out

        # 傳遞值給下次迴圈使用
        previous_error = error
        previous_time = current_time

        # 記錄數值
        R_TARGET_ALT.append(TARGET_ALT)
        R_ACTUAL_ALT_CTRL.append(ACTUAL_ALT)
        R_TARGET_THRUST.append(TARGET_THRUST)
        R_ALT_ERROR.append(error)
        R_ALT_GP.append(gain_p)
        R_ALT_GI.append(gain_i)
        R_ALT_GD.append(gain_d)

        # 迴圈執行間隔
        time.sleep(1 / FREQ_ALT)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_ALT = False
    print('<關閉程序> 高度控制器：關閉')


# =========== 高度感測器整合程式 =========================================
# 說明：
# 負責整合超音波高度計、飛控氣壓計(可能為IMU來源)，做訊號整理給控制器使用
# 各訊號先經均值濾波後，再用卡爾曼濾波處理，最後再針對所有來源做加權平均
# 本函式處理氣壓計之濾波、高度感測器整合濾波
def altitude_sensor():
    global ACTUAL_ALT, ACTUAL_ALT_IMU, STATE_SENSOR_ALT
    # 平均降噪設定
    height_register_imu = []

    # 卡爾曼濾波器初始化(IMU高度計用)
    x_o = 0                     # 前一次狀態
    p_o = 0                     # 前一次斜方差
    z_o = 0                     # 前一次量測值
    q = math.exp(-IMU_KF_Q)     # 斜方差噪音值
    r = 2.92 * math.exp(-3)     # 斜方差量測噪音值(定值不更改)

    # 執行IMU高度計校正程序
    imu_calibrate()

    # 等待超音波高度計啟動
    if not STATE_SENSOR_SONAR:
        print('<啟動程序> 高度感測器：等待超音波高度計啟動')
    while not STATE_SENSOR_SONAR:
        time.sleep(0.1)

    # 回報高度感測器啟動
    STATE_SENSOR_ALT = True
    print('<啟動程序> 高度感測器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 讀取超音波高度計資料
        height_sonar = ACTUAL_ALT_SONAR

        # IMU高度計數據處理
        # 卡爾曼濾波-預測
        x_p = x_o
        p_p = p_o + q

        # 讀值與補正
        height_raw_imu = uav.location.global_relative_frame.alt + IMU_OFFSET

        # 均值濾波器
        height_register_imu.insert(0, height_raw_imu)
        if len(height_register_imu) > IMU_SAMPLE_NUM_AVE:
            height_register_imu.pop()
        height_imu_ave = np.average(height_register_imu)

        # 卡爾曼濾波器輸入限制(防止上衝突波)
        if height_imu_ave <= IMU_KF_INPUT_MAX:
            z = height_imu_ave  # 若低於限制則使用新值
        else:
            z = z_o  # 若超出限制則使用舊值

        # 卡爾曼濾波-更新
        k = p_p / (p_p + r)  # 卡爾曼增益
        x = x_p + k * (z - x_p)  # 狀態(濾波輸出值)
        p = (1 - k) * p_p  # 斜方差

        # 更新高度值至變數
        height_imu = x
        ACTUAL_ALT_IMU = height_imu

        # 卡爾曼濾波：傳遞值給下次迴圈使用
        x_o = x
        p_o = p
        z_o = z

        # 加權平均(可調整加權值)
        alt_num = (WEIGHTED_SONAR * height_sonar + WEIGHTED_IMU * height_imu)
        alt_den = (WEIGHTED_SONAR + WEIGHTED_IMU)
        alt_ave = alt_num / alt_den

        # 更新高度值至全域變數
        ACTUAL_ALT = alt_ave

        # 記錄數值
        R_ACTUAL_ALT_SONAR.append(height_sonar)  # 記錄超音波高度計值
        R_ACTUAL_ALT_IMU_RAW.append(uav.location.global_relative_frame.alt)  # 記錄氣壓高度計原始值
        R_ACTUAL_ALT_IMU_OFFSET.append(height_raw_imu)  # 記錄補正值
        R_ACTUAL_ALT_IMU_AVE.append(height_imu_ave)  # 記錄均值濾波值
        R_ACTUAL_ALT_IMU_KF.append(height_imu)  # 記錄KF值
        R_ACTUAL_ALT.append(ACTUAL_ALT)  # 記錄總高度值

        time.sleep(1 / FREQ_SENSOR_ALT)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_ALT = False
    print('<關閉程序> 高度感測器：關閉')


# ========== 超音波高度計 ==============================================
# 說明：
# 使用HC-SR04為超音波感測模組，發出一脈波訊號使模組執行量測後傳回經過時間長的脈波
# 實際使用上雜訊相對多(可能受飛行時噪音影響)，故施加多重濾波條件：
# 超時防止(防止過長的無效等待)、角度補正、突波保護、均值濾波、卡爾曼濾波
# 經處理後的值堪用度大幅提高，對吸音面也能有效處理，使高度控制器能穩定運作
def sonar():
    global ACTUAL_ALT_SONAR, STATE_SENSOR_SONAR, SONAR_ERR_COUNT
    # 初始化腳位
    GPIO.setmode(GPIO.BCM)  # 設定為BCM模式
    GPIO.setwarnings(False)  # 忽略警告
    GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)  # TRIG腳位輸出設定
    GPIO.setup(SONAR_ECHO_PIN, GPIO.IN)  # ECHO腳位輸入設定
    GPIO.output(SONAR_TRIG_PIN, False)  # 預設輸出為0

    # 脈波變化時間值變數
    pulse_start = time.time()  # 脈波起始時間
    pulse_end = time.time()  # 脈波結束時間

    # 卡爾曼濾波器初始化
    x_o = 0  # 前一次狀態
    p_o = 0  # 前一次斜方差
    z_o = 0  # 前一次量測值
    q = math.exp(-SONAR_KF_Q)  # 斜方差噪音值
    r = 2.92 * math.exp(-3)  # 斜方差量測噪音值(定值不更改)

    # 平均降噪設定
    height_register = []
    height_ave = 0

    # 執行超音波高度計校正程序
    sonar_calibrate()

    # 回報超音波高度計啟動
    STATE_SENSOR_SONAR = True
    print('<啟動程序> 超音波高度計：啟動')

    # 程式迴圈
    while STATE_THREAD:
        # 重設超時旗標
        timeout = True

        # 卡爾曼濾波-預測
        x_p = x_o
        p_p = p_o + q

        # 量測迴圈(線程優先執行，超時防止)
        while timeout:
            # 腳位初始化(因有報錯記錄，於每次前重新初始化)
            GPIO.setmode(GPIO.BCM)  # 設定為BCM模式
            GPIO.setwarnings(False)  # 忽略警告
            GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)  # TRIG腳位輸出設定

            # 輸出啟用脈波
            GPIO.output(SONAR_TRIG_PIN, True)
            time.sleep(0.00001)  # 發出10us脈波
            GPIO.output(SONAR_TRIG_PIN, False)
            timeout_start = time.time()  # 記錄超時基準時間

            # 接收計時脈波
            # 於低電位迴圈等待，跳出時記錄脈波起始時間
            while not GPIO.input(SONAR_ECHO_PIN):
                pulse_start = time.time()  # 記錄脈波起始時間
                timeout = False  # 成功記錄脈波時間，切換旗標
                if pulse_start - timeout_start > SONAR_TIMEOUT:
                    timeout = True  # 等待接收時發生超時，跳出迴圈
                    break

            # 於高電位迴圈等待，跳出時記錄脈波結束時間(若發生超時則不執行本迴圈)
            while GPIO.input(SONAR_ECHO_PIN) and not timeout:
                pulse_end = time.time()  # 記錄脈波結束時間
                timeout = False  # 成功記錄脈波時間，切換旗標
                if pulse_end - pulse_start > SONAR_TIMEOUT:
                    timeout = True  # 接收時長發生超時，跳出迴圈
                    break

            # 若發生超時則顯示於訊息(可停用)
            if timeout:
                SONAR_ERR_COUNT += 1  # 超時次數計(統計用)
                # print('<錯誤> 超音波高度計：量測超時')

        # 計算高度值
        pulse_duration = pulse_end - pulse_start
        height_raw = (pulse_duration * (331.3 + 0.606 * SONAR_AIR_TEMP)) / 2

        # 高度角度補正
        roll_angle = uav.attitude.roll
        pitch_angle = uav.attitude.pitch
        height_offset = (height_raw * math.cos(roll_angle) * math.cos(pitch_angle)) + SONAR_OFFSET
        height_offset_r = height_offset  # 記錄用值

        # 上下降突波防止
        if ACTUAL_ALT > SONAR_FILTER_ALT:
            if math.fabs(height_offset - z_o) > SONAR_MAX_CHANGE:
                height_offset = height_ave

        # 均值濾波器(使線條滑順)
        height_register.insert(0, height_offset)
        if len(height_register) > SONAR_SAMPLE_NUM_AVE:
            height_register.pop()
        height_ave = np.average(height_register)  # 依據取樣次數計算即時平均值(會產生響應延遲)

        # 卡爾曼濾波器輸入限制(防止突波)
        if height_ave <= SONAR_KF_INPUT_MAX:
            z = height_ave  # 若低於限制則使用新值
        else:
            z = z_o  # 若超出限制則使用舊值

        # 卡爾曼濾波-更新
        k = p_p / (p_p + r)  # 卡爾曼增益
        x = x_p + k * (z - x_p)  # 狀態(濾波輸出值)
        p = (1 - k) * p_p  # 斜方差

        # 更新高度值至全域變數
        ACTUAL_ALT_SONAR = x

        # 傳遞值給下次迴圈使用
        x_o = x
        p_o = p
        z_o = z

        # 記錄數值
        R_ACTUAL_ALT_SONAR_RAW.append(height_raw)  # 記錄原始值
        R_ACTUAL_ALT_SONAR_OFFSET.append(height_offset_r)  # 記錄補正值(突波防止前)
        R_ACTUAL_ALT_SONAR_AVE.append(height_ave)  # 記錄平均值
        R_ACTUAL_ALT_SONAR_KF.append(ACTUAL_ALT_SONAR)  # 記錄KF值

        # 迴圈執行間隔
        time.sleep(1 / FREQ_SENSOR_SONAR)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_SONAR = False
    print('<關閉程序> 超音波高度計：關閉；量測超時次數共 %d 次' % SONAR_ERR_COUNT)


# ========== 命令傳輸程式 ==============================================
# 根據Ardupilot與MavLink協定規範，僅能使用set_attitude_target命令
# 另DroneKit套件提供函式封裝上也是依據上述規範編碼，故所使用的命令包括：
# 姿態四元數、各軸角速度、推力(昇降率)、操作模式(角度、角速度)
# 姿態四元數將控制器的角度值傳至函式做計算後返還、角速度採徑度制
# 根據測試及參考習慣操作，故選用姿態控制但於Yaw軸採角速度控制，符合手控習慣
# 流程上為將各控制器所輸出的命令值做統一編碼封裝後傳送
def command_transfer():
    global STATE_CMD

    # 回報命令傳輸程式啟動
    STATE_CMD = True
    print('<啟動程序> 命令傳輸程式：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 命令編碼，使用偏航角速度控制(餘採姿態控制)
        msg = uav.message_factory.set_attitude_target_encode(
            0,                              # 開機時間
            0,                              # 目標系統
            0,                              # 目標裝置
            0b00000011,                     # 命令遮罩(1忽略，0啟用)
            to_quaternion(TARGET_ROLL, TARGET_PITCH, TARGET_YAW),  # 姿態四元數
            0,                              # Roll滾轉角速度(radian/s)
            0,                              # Pitch滾轉角速度(radian/s)
            math.radians(TARGET_YAW_RATE),  # Yaw滾轉角速度(radian/s)
            TARGET_THRUST)                  # 推力升降比(0~1，中點0.5)

        # 傳輸命令
        uav.send_mavlink(msg)

        # 迴圈執行間隔
        time.sleep(1 / FREQ_CMD)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_CMD = False
    print('<關閉程序> 命令傳輸程式：關閉')


# ========== 飛行記錄器 ================================================
def flight_recoder():
    # 數值初始化
    roll_old = 0
    pitch_old = 0
    yaw_old = 0
    roll_rate_old = 0
    pitch_rate_old = 0
    yaw_rate_old = 0
    vx_old = 0
    vy_old = 0
    vz_old = 0
    t_end = time.time()

    # 數值記錄迴圈
    while STATE_THREAD:
        # 更新時間戳記
        t_start = time.time()

        # 記錄姿態值
        roll = math.degrees(uav.attitude.roll)
        pitch = math.degrees(uav.attitude.pitch)
        yaw = math.degrees(uav.attitude.yaw)

        R_ACTUAL_ROLL.append(roll)
        R_ACTUAL_PITCH.append(pitch)
        R_ACTUAL_YAW.append(yaw)

        # 記錄姿態速度
        d_t = t_end - t_start
        roll_rate = (roll - roll_old) / d_t
        pitch_rate = (pitch - pitch_old) / d_t
        yaw_rate = (yaw - yaw_old) / d_t

        R_ACTUAL_ROLL_RATE.append(roll_rate)
        R_ACTUAL_PITCH_RATE.append(pitch_rate)
        R_ACTUAL_YAW_RATE.append(yaw_rate)

        # 記錄姿態加速度
        roll_acc = (roll_rate - roll_rate_old) / d_t
        pitch_acc = (pitch_rate - pitch_rate_old) / d_t
        yaw_acc = (yaw_rate - yaw_rate_old) / d_t

        R_ACTUAL_ROLL_ACC.append(roll_acc)
        R_ACTUAL_PITCH_ACC.append(pitch_acc)
        R_ACTUAL_YAW_ACC.append(yaw_acc)

        # 記錄速度值
        vx = uav.velocity[0]
        vy = uav.velocity[1]
        vz = uav.velocity[2]

        R_VEL_X.append(vx)
        R_VEL_Y.append(vy)
        R_VEL_Z.append(vz)

        # 記錄加速度值
        ax = (vx - vx_old) / d_t
        ay = (vy - vy_old) / d_t
        az = (vz - vz_old) / d_t

        R_ACC_X.append(ax)
        R_ACC_Y.append(ay)
        R_ACC_Z.append(az)

        # 傳遞值
        roll_old = roll
        pitch_old = pitch
        yaw_old = yaw
        roll_rate_old = roll_rate
        pitch_rate_old = pitch_rate
        yaw_rate_old = yaw_rate
        vx_old = vx
        vy_old = vy
        vz_old = vz
        t_end = t_start

        # 迴圈執行間隔(精確版)
        while time.time()-t_start < 1/FREQ_FR:
            pass


# ====================================================================
# ========== 子函式區塊 ================================================
# ====================================================================
# 本區放置各主函式下會使用到的功能函式，分有校正類、計算類、裝備控制類
# 將常使用的重複功能函式化以節省程式大小(僅限無傳遞值類的函式)


# ========== 超音波高度計校正程序 ========================================
# 於啟動前自動校正高度，為避免單次量測不準確，故以多次取樣後平均為基準來補正
def sonar_calibrate():
    global SONAR_OFFSET
    # 平均降噪設定
    height_register = []

    # 脈波變化時間值變數
    pulse_start = time.time()  # 脈波起始時間
    pulse_end = time.time()  # 脈波結束時間

    # 量測迴圈(線程優先執行)
    for i in range(0, SONAR_SAMPLE_NUM_CAL):
        # 設定超時旗標
        timeout = True

        # 量測迴圈(超時防止)
        while timeout:
            # 腳位初始化(因有報錯記錄，於每次前重新初始化)
            GPIO.setmode(GPIO.BCM)  # 設定為BCM模式
            GPIO.setwarnings(False)  # 忽略警告
            GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)  # TRIG腳位輸出設定

            # 輸出啟用脈波
            GPIO.output(SONAR_TRIG_PIN, True)
            time.sleep(0.00001)  # 發出10us脈波
            GPIO.output(SONAR_TRIG_PIN, False)
            timeout_start = time.time()  # 記錄超時基準時間

            # 接收計時脈波
            # 於低電位迴圈等待，跳出時記錄脈波起始時間
            while not GPIO.input(SONAR_ECHO_PIN):
                pulse_start = time.time()  # 記錄脈波起始時間
                timeout = False  # 成功記錄脈波時間，切換旗標
                if pulse_start - timeout_start > SONAR_TIMEOUT:
                    timeout = True  # 等待接收時發生超時，跳出迴圈
                    break

            # 於高電位迴圈等待，跳出時記錄脈波結束時間(若發生超時則不執行本迴圈)
            while GPIO.input(SONAR_ECHO_PIN) and not timeout:
                pulse_end = time.time()  # 記錄脈波結束時間
                timeout = False  # 成功記錄脈波時間，切換旗標
                if pulse_end - pulse_start > SONAR_TIMEOUT:
                    timeout = True  # 接收時長發生超時，跳出迴圈
                    break

            # 若發生超時則顯示於訊息
            if timeout:
                # print('<錯誤> 超音波感測器：量測超時')
                pass

        # 計算高度值
        pulse_duration = pulse_end - pulse_start
        height = (pulse_duration * (331.3 + 0.606 * SONAR_AIR_TEMP)) / 2

        # 均值濾波器
        height_register.append(height)

        # 等待10ms執行下次量測
        time.sleep(0.1)

    # 更新高度校正偏移量
    SONAR_OFFSET = -np.average(height_register)
    print('<啟動程序> 超音波高度計校正完畢，偏移 %.3f m' % SONAR_OFFSET)


# ========== IMU高度計校正程序 =========================================
# 因IMU高度計會有波動，需做補正
def imu_calibrate():
    global IMU_OFFSET
    # 平均降噪設定
    height_register = []

    # 校正氣壓計
    uav.send_calibrate_barometer()
    uav.send_calibrate_vehicle_level()
    time.sleep(1)

    # 量測迴圈(線程優先執行)
    for i in range(0, IMU_SAMPLE_NUM_CAL):
        # 讀取氣壓計高度值
        height = uav.location.global_relative_frame.alt

        # 均值濾波器
        height_register.append(height)

        # 等待200ms執行下次量測
        time.sleep(0.2)

    # 更新高度校正偏移量
    IMU_OFFSET = -np.average(height_register)
    print('<啟動程序> 氣壓高度計校正完畢，偏移 %.3f m' % IMU_OFFSET)


# ========== 轉換尤拉角到四元數 =========================================
# 因命令格式針對姿態需使用四元數格式，因此需先把尤拉角作轉換，此格式可避免萬向節鎖產生
# 輸入格式為度(degree)，輸出為四元數(q,x,y,z)
def to_quaternion(roll=0.0, pitch=0.0, yaw=0.0):
    t0 = math.cos(math.radians(yaw * 0.5))
    t1 = math.sin(math.radians(yaw * 0.5))
    t2 = math.cos(math.radians(roll * 0.5))
    t3 = math.sin(math.radians(roll * 0.5))
    t4 = math.cos(math.radians(pitch * 0.5))
    t5 = math.sin(math.radians(pitch * 0.5))

    w = t0 * t2 * t4 + t1 * t3 * t5
    x = t0 * t3 * t4 - t1 * t2 * t5
    y = t0 * t2 * t5 + t1 * t3 * t4
    z = t1 * t2 * t4 - t0 * t3 * t5

    return [w, x, y, z]


# ========== 多線程啟動函式 ============================================
# 供主程式啟動多線程用，若連線中斷時，自動重新連線
def thread_activate():
    global STATE_THREAD, STATE_UAV, uav
    # 若未連線，重新執行連線
    if not STATE_UAV:
        print('<啟動程序> 與UAV連線中，請稍候...')
        # 樹莓派連線指令
        uav = connect(CONNECT_PORT, baud=CONNECT_BAUD)
        # 電腦連線指令
        # uav = connect(CONNECT_PORT, baud=CONNECT_BAUD)
        STATE_UAV = True
        print('<啟動程序> UAV已連線')

    # 定義多線程
    sonar_sen = threading.Thread(target=sonar)
    alt_sen = threading.Thread(target=altitude_sensor)
    alt_ctrl = threading.Thread(target=altitude_controller)
    cmd = threading.Thread(target=command_transfer)

    # 啟動多線程旗標
    STATE_THREAD = True

    # 執行多線程(依啟動順序放置)
    sonar_sen.start()
    alt_sen.start()
    alt_ctrl.start()
    cmd.start()

    # 回報多線程啟動
    print('<啟動程序> 多線程程序：啟動')


def thread_activate2():
    # 定義多線程
    fr = threading.Thread(target=flight_recoder)
    # 啟動多線程
    fr.start()


# ====================================================================
# ========== 數值紀錄區塊 ==============================================
# ====================================================================
# ========== Excel 輸出 ===============================================
def data_sheet():
    # 建立活頁簿
    db = Workbook()

    # 建立分頁
    ds1 = db.create_sheet("ALT Controller")
    ds2 = db.create_sheet("ALT Sensor")
    ds3 = db.create_sheet("Flight Recoder")

    # 高度控制器(先列在行)
    ds1.cell(1, 1, '高度控制器')
    ds1.cell(2, 1, '實際高度')
    for x in range(len(R_ACTUAL_ALT_CTRL)):
        ds1.cell(x + 3, 1, R_ACTUAL_ALT_CTRL[x])

    ds1.cell(2, 2, '目標高度')
    for x in range(len(R_TARGET_ALT)):
        ds1.cell(x + 3, 2, R_TARGET_ALT[x])

    ds1.cell(2, 3, '爬升率')
    for x in range(len(R_TARGET_THRUST)):
        ds1.cell(x + 3, 3, R_TARGET_THRUST[x])

    # 高度控制器PID響應
    ds1.cell(1, 5, '高度控制器PID響應')
    ds1.cell(2, 5, 'Error')
    for x in range(len(R_ALT_ERROR)):
        ds1.cell(x + 3, 5, R_ALT_ERROR[x])

    ds1.cell(2, 6, 'Gain P')
    for x in range(len(R_ALT_GP)):
        ds1.cell(x + 3, 6, R_ALT_GP[x])

    ds1.cell(2, 7, 'Gain I')
    for x in range(len(R_ALT_GI)):
        ds1.cell(x + 3, 7, R_ALT_GI[x])

    ds1.cell(2, 8, 'Gain D')
    for x in range(len(R_ALT_GD)):
        ds1.cell(x + 3, 8, R_ALT_GD[x])

    # 第二頁-感測器
    # 高度感測器數值處理圖
    ds2.cell(1, 1, '高度感測器數值處理圖')
    ds2.cell(2, 1, '實際高')
    for x in range(len(R_ACTUAL_ALT)):
        ds2.cell(x + 3, 1, R_ACTUAL_ALT[x])

    ds2.cell(2, 2, '超音波')
    for x in range(len(R_ACTUAL_ALT_SONAR)):
        ds2.cell(x + 3, 2, R_ACTUAL_ALT_SONAR[x])

    ds2.cell(2, 3, 'IMU')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds2.cell(x + 3, 3, R_ACTUAL_ALT_IMU_KF[x])

    # 超音波高度計數值處理圖
    ds2.cell(1, 5, '超音波高度計數值處理圖')
    ds2.cell(2, 5, '原始值')
    for x in range(len(R_ACTUAL_ALT_SONAR_RAW)):
        ds2.cell(x + 3, 5, R_ACTUAL_ALT_SONAR_RAW[x])

    ds2.cell(2, 6, '補正後')
    for x in range(len(R_ACTUAL_ALT_SONAR_OFFSET)):
        ds2.cell(x + 3, 6, R_ACTUAL_ALT_SONAR_OFFSET[x])

    ds2.cell(2, 7, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_AVE)):
        ds2.cell(x + 3, 7, R_ACTUAL_ALT_SONAR_AVE[x])

    ds2.cell(2, 8, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_KF)):
        ds2.cell(x + 3, 8, R_ACTUAL_ALT_SONAR_KF[x])

    # IMU高度計數值處理圖
    ds2.cell(1, 10, 'IMU高度計數值處理圖')
    ds2.cell(2, 10, '原始值')
    for x in range(len(R_ACTUAL_ALT_IMU_RAW)):
        ds2.cell(x + 3, 10, R_ACTUAL_ALT_IMU_RAW[x])

    ds2.cell(2, 11, '補正後')
    for x in range(len(R_ACTUAL_ALT_IMU_OFFSET)):
        ds2.cell(x + 3, 11, R_ACTUAL_ALT_IMU_OFFSET[x])

    ds2.cell(2, 12, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_AVE)):
        ds2.cell(x + 3, 12, R_ACTUAL_ALT_IMU_AVE[x])

    ds2.cell(2, 13, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds2.cell(x + 3, 13, R_ACTUAL_ALT_IMU_KF[x])

    # 第三頁
    # 姿態記錄(先列在行)
    ds3.cell(1, 1, '姿態')
    ds3.cell(2, 1, 'Roll')
    for x in range(len(R_ACTUAL_ROLL)):
        ds3.cell(x + 3, 1, R_ACTUAL_ROLL[x])

    ds3.cell(2, 2, 'Pitch')
    for x in range(len(R_ACTUAL_PITCH)):
        ds3.cell(x + 3, 2, R_ACTUAL_PITCH[x])

    ds3.cell(2, 3, 'Yaw')
    for x in range(len(R_ACTUAL_YAW)):
        ds3.cell(x + 3, 3, R_ACTUAL_YAW[x])

    # 姿態速度記錄(先列在行)
    ds3.cell(1, 5, '姿態速度')
    ds3.cell(2, 5, 'Roll Rate')
    for x in range(len(R_ACTUAL_ROLL_RATE)):
        ds3.cell(x + 3, 5, R_ACTUAL_ROLL_RATE[x])

    ds3.cell(2, 6, 'Pitch Rate')
    for x in range(len(R_ACTUAL_PITCH_RATE)):
        ds3.cell(x + 3, 6, R_ACTUAL_PITCH_RATE[x])

    ds3.cell(2, 7, 'Yaw Rate')
    for x in range(len(R_ACTUAL_YAW_RATE)):
        ds3.cell(x + 3, 7, R_ACTUAL_YAW_RATE[x])

    # 姿態加速度記錄(先列在行)
    ds3.cell(1, 9, '姿態加速度')
    ds3.cell(2, 9, 'Roll Acc')
    for x in range(len(R_ACTUAL_ROLL_ACC)):
        ds3.cell(x + 3, 9, R_ACTUAL_ROLL_ACC[x])

    ds3.cell(2, 10, 'Pitch Acc')
    for x in range(len(R_ACTUAL_PITCH_ACC)):
        ds3.cell(x + 3, 10, R_ACTUAL_PITCH_ACC[x])

    ds3.cell(2, 11, 'Yaw Acc')
    for x in range(len(R_ACTUAL_YAW_ACC)):
        ds3.cell(x + 3, 11, R_ACTUAL_YAW_ACC[x])

    # 速度記錄(先列在行)
    ds3.cell(1, 13, '速度')
    ds3.cell(2, 13, 'Vx')
    for x in range(len(R_VEL_X)):
        ds3.cell(x + 3, 13, R_VEL_X[x])

    ds3.cell(2, 14, 'Vy')
    for x in range(len(R_VEL_Y)):
        ds3.cell(x + 3, 14, R_VEL_Y[x])

    ds3.cell(2, 15, 'Vz')
    for x in range(len(R_VEL_Z)):
        ds3.cell(x + 3, 15, R_VEL_Z[x])

    # 加速度記錄(先列在行)
    ds3.cell(1, 17, '速度')
    ds3.cell(2, 17, 'Ax')
    for x in range(len(R_ACC_X)):
        ds3.cell(x + 3, 17, R_ACC_X[x])

    ds3.cell(2, 18, 'Ay')
    for x in range(len(R_ACC_Y)):
        ds3.cell(x + 3, 18, R_ACC_Y[x])

    ds3.cell(2, 19, 'Az')
    for x in range(len(R_ACC_Z)):
        ds3.cell(x + 3, 19, R_ACC_Z[x])

    # 依照時間存檔
    t_save = time.localtime()
    db.save('Flight Data %d_%d_%d_%d_%d_%d.xlsx' %
            (t_save[0], t_save[1], t_save[2], t_save[3], t_save[4], t_save[5]))


# ========== 高度控制器 ================================================
def alt_ctrl_data():
    # ======高度控制器================
    # 高度控制器響應圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_CTRL) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_CTRL, label="Actual Altitude")
    plt.plot(num_alt, R_TARGET_ALT, label="Target Altitude")
    plt.plot(num_alt, R_TARGET_THRUST, label="Climb Rate")

    plt.xlabel('Data number')
    plt.ylabel('Height(m), Velocity(m/s)')
    plt.title('Altitude Controller Data')
    plt.legend()
    plt.savefig('Altitude Controller Data.png')
    plt.show()

    # 高度控制器PID輸出圖
    num_alt = []
    for i in range(1, len(R_ALT_ERROR) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ALT_ERROR, label="Error")
    plt.plot(num_alt, R_ALT_GP, label="Gain P")
    plt.plot(num_alt, R_ALT_GI, label="Gain I")
    plt.plot(num_alt, R_ALT_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Height(m), Velocity(m/s)')
    plt.title('Altitude Controller PID Output Data')
    plt.legend()
    plt.savefig('Altitude Controller PID Output Data.png')
    plt.show()


# ========== 高度感測器 ================================================
def alsen_data():
    # 高度感測器數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT, label="Actual Height")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR, label="Sonar Data")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_KF, label="IMU Data")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Sensor Data')
    plt.legend()
    plt.savefig('Altitude Sensor Data.png')
    plt.show()

    # 超音波高度計數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_SONAR_KF) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_RAW, label="Raw Data")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_OFFSET, label="Angle & Height Offset")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_AVE, label="Average Filter")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_KF, label="Kalman Filter")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Data - Sonar')
    plt.legend()
    plt.savefig('Altitude Data - Sonar.png')
    plt.show()

    # IMU高度計數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_IMU_RAW) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_IMU_RAW, label="Raw Data")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_OFFSET, label="Height Offset")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_AVE, label="Average Filter")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_KF, label="Kalman Filter")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Data - IMU')
    plt.legend()
    plt.savefig('Altitude Data - IMU.png')
    plt.show()


# ========== 飛行記錄器 ================================================
def fr_data():
    # 姿態
    num_att = []
    for i in range(1, len(R_ACTUAL_ROLL) + 1):
        num_att.append(i)

    plt.plot(num_att, R_ACTUAL_ROLL, label="Roll")
    plt.plot(num_att, R_ACTUAL_PITCH, label="Pitch")

    plt.xlabel('Data number')
    plt.ylabel('Angle(degree)')
    plt.title('Attitude Data')
    plt.legend()
    plt.savefig('Attitude Data.png')
    plt.show()

    # 姿態速度
    plt.plot(num_att, R_ACTUAL_ROLL_RATE, label="Roll Rate")
    plt.plot(num_att, R_ACTUAL_PITCH_RATE, label="Pitch Rate")
    plt.plot(num_att, R_ACTUAL_YAW_RATE, label="Yaw Rate")

    plt.xlabel('Data number')
    plt.ylabel('Angular Velocity(degree/s)')
    plt.title('Attitude Rate Data')
    plt.legend()
    plt.savefig('Attitude Rate Data.png')
    plt.show()

    # 姿態加速度
    plt.plot(num_att, R_ACTUAL_ROLL_ACC, label="Roll Acc")
    plt.plot(num_att, R_ACTUAL_PITCH_ACC, label="Pitch Acc")
    plt.plot(num_att, R_ACTUAL_YAW_ACC, label="Yaw Acc")

    plt.xlabel('Data number')
    plt.ylabel('Angular Acceleration(degree/s^2)')
    plt.title('Attitude Acc Data')
    plt.legend()
    plt.savefig('Attitude Acc Data.png')
    plt.show()

    # 速度
    plt.plot(num_att, R_VEL_X, label="X-Velocity")
    plt.plot(num_att, R_VEL_Y, label="Y-Velocity")
    plt.plot(num_att, R_VEL_Z, label="Z-Velocity")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s)')
    plt.title('Velocity Data')
    plt.legend()
    plt.savefig('Velocity Data.png')
    plt.show()

    # 加速度
    plt.plot(num_att, R_ACC_X, label="X-Acc")
    plt.plot(num_att, R_ACC_Y, label="Y-Acc")
    plt.plot(num_att, R_ACC_Z, label="Z-Acc")

    plt.xlabel('Data number')
    plt.ylabel('Acceleration(m/s^2)')
    plt.title('Acceleration Data')
    plt.legend()
    plt.savefig('Acceleration Data.png')
    plt.show()


# ====================================================================
# ========== 飛行程序指令(條件與命令) ====================================
# ====================================================================
# 任務中會使用的區段飛行命令、模式切換命令、觸發條件


# ========== 啟動前檢查 ================================================
# 本程式為飛行前的安全檢查項目，除確認控制器是否上線外，也提醒控制員檢查安全程序
# 項目有：安全裝置移除(槳片銷)、起飛區淨空，完成後即交付起飛控制
def pre_check():
    # 執行起飛前檢查
    print('<啟動程序> 啟動前檢查程序：啟動')

    # 階段一：控制器檢查
    # 等待高度控制器啟動
    if not STATE_ALT:
        print('<啟動程序> 主程序：等待高度控制器啟動')
    while not STATE_ALT:
        time.sleep(0.1)

    # 等待命令傳輸啟動
    if not STATE_CMD:
        print('<啟動程序> 主程序：等待命令傳輸啟動')
    while not STATE_CMD:
        time.sleep(0.1)

    # 階段三：安全檢查程序
    # 確認安全裝置解除
    safe_pin = 'N'
    while safe_pin != 'Y':
        safe_pin = input('<啟動程序> 請確認安全裝置已移除(Y/N) => ')
    print('<啟動程序> 安全裝置已移除')
    time.sleep(1)

    # 確認起飛區已淨空
    lunch_clear = 'N'
    while lunch_clear != 'Y':
        lunch_clear = input('<啟動程序> 請確認起飛區已淨空(Y/N) => ')
    print('<啟動程序> 起飛區已淨空')

    # 回報起飛前檢點完成
    print('<啟動程序> 啟動前檢查程序：完成')
    time.sleep(1)

    # 飛機電量資訊
    print('無人機目前電池電壓： %.2f V，剩餘電量 %d %%\n' % (uav.battery.voltage, uav.battery.level))

    # 任務開始確認
    go_or_nogo = 'N'
    while go_or_nogo != 'Y':
        go_or_nogo = input('<啟動程序> 請求任務開始(Y/N) => ')

    # 回報任務執行開始
    print('<啟動程序> TDK飛行任務，開始!')
    time.sleep(1)


# ========== 起飛程序 =================================================
# 執行模式切換、無人機解鎖，並由高度控制器將無人機送至目標飛行高度
# 使用斜波上升法
def take_off(target_alt):
    global TARGET_ALT, MODE
    # 切換至無衛星導航模式
    uav.mode = VehicleMode("GUIDED_NOGPS")
    while not uav.mode.name == 'GUIDED_NOGPS':
        time.sleep(0.5)
    print('<啟動程序> 無人機已切換至無衛星導航模式')

    # 解鎖無人機
    uav.arm(wait=True)
    print('<啟動程序> 無人機已解鎖')

    # 設定目標高度
    print('<飛行命令> 起飛至目標高度 %.2f m' % target_alt)

    # 等待抵達目標高度(每1秒更新一次資訊，每0.5秒更新高度值，每0.1秒檢查一次)
    i = 0
    while ACTUAL_ALT < target_alt * (TAKEOFF_SETTLE / 100):
        # 更新目標高
        if i/5 == 0:
            TARGET_ALT += ALT_SPEED
            if TARGET_ALT > target_alt:
                TARGET_ALT = target_alt

        # 顯示資訊
        if i/10 == 0:
            print('<飛行狀態> 起飛中，目前高度 %.2f m，距離目標高度尚有 %.2f m' %
                  (ACTUAL_ALT, target_alt - ACTUAL_ALT))
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m\n' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))

        # 急停跳出
        if not STATE_THREAD:
            break

        # 計數器
        i += 1
        time.sleep(0.1)

    # 回報抵達目標高度
    print('<飛行狀態> 已抵達目標高度 %.2f m' % TARGET_ALT)


# ========== 姿態程序 =================================================
def attitude(roll=0, pitch=0, yaw=0, duration=0):
    global TARGET_ROLL, TARGET_PITCH, TARGET_YAW
    TARGET_ROLL = roll
    TARGET_PITCH = pitch
    TARGET_YAW = yaw
    print('<飛行命令> 目標姿態Roll:%.2f Pitch:%.2f Yaw:%.2f(度)' % (roll, pitch, yaw))

    t_start = time.time()
    i = 0
    while time.time()-t_start < duration:
        # 顯示資訊
        if i/10 == 0:
            print('<飛行狀態> 飛行中，目前高度 %.2f m' % ACTUAL_ALT)
            print('<飛行狀態> 飛行中，目前姿態Roll:%.2f Pitch:%.2f Yaw:%.2f(度)' %
                  (math.degrees(uav.attitude.roll),
                   math.degrees(uav.attitude.pitch),
                   math.degrees(uav.attitude.yaw)))

        i += 1
        time.sleep(0.1)


# ========== 降落程序 =================================================
def landing():
    global TARGET_ALT
    # 設定目標高度
    print('<飛行命令> 位置已確認，開始降落至地面')
    TARGET_ALT = -0.05  # 負值以確保不會著地後彈升

    # 等待抵達目標高度(每1秒更新一次資訊，每0.1秒檢查一次)
    i = 0
    while ACTUAL_ALT > LANDING_CUTOFF:
        if i/10 == 0:
            print('<飛行狀態> 降落中，目前高度 %.2f m' % ACTUAL_ALT)
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))

        if not STATE_THREAD:
            break

        i += 1
        time.sleep(0.1)

    # 回報降落成功
    print('<飛行狀態> 已降落地面')

    # 返回手動模式
    uav.mode = VehicleMode("STABILIZE")
    while not uav.mode.name == 'STABILIZE':
        time.sleep(0.5)
    print('<關閉程序> 無人機已切換至手動模式')


# ====================================================================
# ========== 前景主函式區塊 ============================================
# ====================================================================
# 任務主函式
def mission():
    global STATE_THREAD, STATE_UAV
    print('<啟動程序> 任務程序：啟動')
    # 啟動多線程
    thread_activate()

    # 飛行任務
    pre_check()                         # 執行起飛前檢查
    thread_activate2()                  # 開始記錄
    take_off(1.2)                       # 起飛至任務高度
    attitude(pitch=1, duration=5)       # 姿態命令
    landing()                           # 執行降落與關閉程序

    # 關閉多線程
    STATE_THREAD = False
    print('<關閉程序> 主程序：關閉')

    # 清空GPIO設定
    GPIO.cleanup()
    print('<關閉程序> GPIO已關閉')

    # 結束與無人機連線
    uav.close()
    STATE_UAV = False
    print('<關閉程序> 無人機連線：關閉')

    # 輸出圖表
    data_sheet()
    fr_data()
    alt_ctrl_data()
    alsen_data()


# 執行程式
mission()
