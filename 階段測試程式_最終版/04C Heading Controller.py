# 起飛區 1017修正版
# 高度控制器採PD控制，可再優化參數
# 位置控制器調整參數配置，視情況需降低控制器頻率
# -內迴圈回授增益需對應外迴圈增益
# -測試出界響應情形、中心截止區調整、輸出截止區調整
# 速度控制器需觀察感測器回授、飛控角度特性、慣性回補
# 航向控制器需處理角度響應過慢問題：鏡頭改放前端
# 適度降低控制器速度:最低5Hz

# 測試進度：
# 1020 完成定高(受限感測器精度暫不處理)
#      完成定位(堪用級，可從四方維持定位)
# 增加雷射飛時測距VL53L1X(未驗證)

# !/usr/bin/env python
# -*- coding: utf-8 -*-

# 載入模組
from dronekit import connect, VehicleMode
from openpyxl import Workbook
# import dronekit_sitl
import matplotlib.pyplot as plt
import RPi.GPIO as GPIO
import numpy as np
import VL53L1X
import threading
import time
import math
import cv2

# ====================================================================
# ========== 設定全域變數 ==============================================
# ====================================================================
# 加註*字為可調整之參數
print('<啟動程序> 載入全域變數設定')

# ========== UAV通訊參數 ==============================================
# 樹莓派Serial連線參數
CONNECT_PORT = '/dev/serial0'   # 傳輸口位址
CONNECT_BAUD = 921600           # 傳輸鮑率

# MacBook 3DR連線參數
# CONNECT_PORT = '/dev/tty.SLAB_USBtoUART'
# CONNECT_BAUD = 57600

# 軟體模擬連線參數
# sitl = dronekit_sitl.start_default()
# CONNECT_PORT = sitl.connection_string()
# CONNECT_BAUD = 921600

# ========== 旗標與變數 ================================================
# ->旗標
STATE_THREAD = False            # 多線程旗標
STATE_UAV = True                # 無人機連線狀態
STATE_ALT = False               # 高度控制器旗標
STATE_POS = False               # 位置控制器旗標
STATE_VEL = False               # 位置-速度控制器旗標
STATE_HDG = False               # 航向控制器旗標
STATE_SPD = False               # 飛行速度控制器旗標
STATE_CAM = False               # 影像擷取程式旗標
STATE_SENSOR_CAM = False        # 影像感測器旗標
STATE_SENSOR_ALT = False        # 高度感測器旗標
STATE_SENSOR_SONAR = False      # 超音波高度計旗標
STATE_SENSOR_SPD = False        # 飛行速度感測器旗標
STATE_CMD = False               # 命令傳輸程式旗標
STATE_BOX = False               # 貨物艙門狀態旗標
MODE = ''                       # 模式旗標(POS:定位模式，LINE:循跡模式)
MISSION_COLOR = ''              # 任務顏色(blue or green)
POS_COLOR = ''                  # 定位的色塊顏色

# ->多線程優先執行命令
THREAD_OCCUPY = threading.Lock()

# ->目標值(命令值)
TARGET_ROLL = 0                 # 目標滾轉角(degree)
TARGET_PITCH = 0                # 目標俯仰角(degree)
TARGET_YAW = 0                  # 目標偏航角(degree)，不使用
TARGET_YAW_RATE = 0             # 偏航角速度(degree/s)，需搭配控制器控制角度
TARGET_THRUST = 0               # 目標推力，表示升降速度(0.5為中點)
TARGET_ALT = 0                  # 目標高度，主程式輸入於高度控制器(m)
TARGET_VEL = [0, 0]             # 目標移動速度(px/s，相機座標)
TARGET_POS = [0, 0]             # 目標型心位置(px，相機座標)
TARGET_SPD = 0                  # 目標飛行速度

# ->實際值(量測值)
ACTUAL_ALT = 0                  # 實際高度(m)
ACTUAL_ALT_SONAR = 0            # 超音波高度計(m)
ACTUAL_ALT_IMU = 0              # IMU氣壓高度計(m)
ACTUAL_ALT_LASER = 0            # 雷射測距儀(m)
ACTUAL_ANG_LINE = 0             # 線角度(degree，Y軸為基準，最大轉角限制用)
ACTUAL_HDG = 0                  # 與線夾角(degree)
ACTUAL_POS = [0, 0]             # 色塊實際位置(px)
ACTUAL_VEL = [0, 0]             # 色塊實際速度(px/s)
ACTUAL_AREA = 0                 # 色塊面積(px^2)
ACTUAL_SPD = 0                  # 實際飛行速度

# ========== 控制器參數設定 =============================================
# ->迴圈頻率(預設20Hz)
FREQ_ALT = 20                   # 高度控制器頻率(不快於高度感測器頻率)
FREQ_POS = 10                   # 位置控制器頻率(不快於位置-速度控制器頻率)
FREQ_VEL = 20                   # 位置-速度控制器頻率(不快於影像禎率)
FREQ_HDG = 20                   # 航向控制器頻率(不快於影像禎率)
FREQ_SPD = 20                   # 飛行速度控制器頻率
FREQ_CAM = 30                   # 影像禎率(fps，依系統負載調整)
FREQ_SENSOR_CAM = 20            # 影像感測器頻率(不快於影像禎率)
FREQ_SENSOR_ALT = 20            # 高度感測器頻率(不快於超音波感測器頻率)
FREQ_SENSOR_SONAR = 20          # 超音波感測器頻率(5~50Hz，越高雜訊越多)
FREQ_SENSOR_SPD = 20            # 飛行速度感測器頻率
FREQ_CMD = 20                   # 命令傳輸頻率(不快於控制器頻率)

# ->高度控制器參數
ALT_KP = 0.1                    # P增益值*
ALT_KI = 0                      # I增益值*
ALT_KD = 0.1                    # D增益值*
ALT_GAIN_D_LIMIT = 0.3
ALT_SPEED = 0.3                 # 昇降速率(m/s，與飛控參數一致)*

# ->位置控制器參數
POS_SAMPLE_NUM_AVE = 10         # 位置控制器均值濾波取樣次數*
POS_X_KP = 0.01                 # 位置控制器X軸-P增益值
POS_X_KI = 0                    # 位置控制器X軸-I增益值
POS_X_KD = 0                    # 位置控制器X軸-D增益值
POS_Y_KP = 0.01                 # 位置控制器Y軸-P增益值
POS_Y_KI = 0                    # 位置控制器Y軸-I增益值
POS_Y_KD = 0                    # 位置控制器Y軸-D增益值
POS_VEL_LIMIT = 2000            # 位置控制器-速度輸出限制(px/s，1.2m高時1px=3.02mm，不使用)
POS_DZ = 50                     # 位置控制器截止區參數(於截止區內使速度控制器控制零速度，防震盪)
# 特殊附註：截止區調校方式=>區域內阻力可以抵消慣性、控制器解析程度內(控制器截止區)

# ->速度控制器參數
POS_VEL_X_KP = 0.1              # 速度控制器X軸-P增益值
POS_VEL_X_KI = 0                # 速度控制器X軸-I增益值
POS_VEL_X_KD = 0                # 速度控制器X軸-D增益值
POS_VEL_Y_KP = 0.1              # 速度控制器Y軸-P增益值
POS_VEL_Y_KI = 0                # 速度控制器Y軸-I增益值
POS_VEL_Y_KD = 0                # 速度控制器Y軸-D增益值
POS_ANGLE_LIMIT = 3             # 速度控制器-最大飛行角度(degree)
POS_ANGLE_CUTOFF = 1            # 速度控制器-最小飛行角度(degree)
POS_VEL_X_FEEDBACK_GAIN = 0.01  # X軸速度回授增益
POS_VEL_Y_FEEDBACK_GAIN = 0.01  # Y軸速度回授增益
POS_VEL_CUTOFF_ALT = 0.5        # 控制器最低作動高度

# ->航向控制器參數
HDG_MODE = ''                   # 航向控制器模式(R:Roll/Y:Yaw/RY:Roll&Yaw-預設)
HDG_KP = 0.1                    # P增益值*
HDG_KI = 0                      # I增益值*
HDG_KD = 0                      # D增益值*
HDG_DZ = 5                      # 截止區角度(degree)*
HDG_ROLL_ANGLE_LIMIT = 3        # 角度輸出限制(degree)*
HDG_ROLL_ANGLE_CUTOFF = 1       # 角度截止限制(degree)*
HDG_YAW_RATE_LIMIT = 10         # 角速度輸出限制(degree/s)*
HDG_ROLL_YAW_RATIO = 5          # 模式RY下滾轉偏航比(大於一以滾轉為主，小於一以偏航為主)

# ->飛行速度控制器參數
SPD_KP = 0.1                    # P增益值*
SPD_KI = 0                      # I增益值*
SPD_KD = 0                      # D增益值*
SPD_ANGLE_CUTOFF = 1            # 飛行速度控制器-最小俯仰角度(degree)
SPD_ANGLE_LIMIT = 3             # 飛行速度控制器-最大俯仰角度(degree)
SPD_LIMIT = 0.5                 # 飛行速度控制器-最大飛行速度(m/s)
SPD_OFF_ANGLE = 5               # 切換跳出角度(degree)，慣性抵銷用
SPD_SAMPLE_NUM_AVE = 5          # 飛行速度感測器均值濾波取樣次數*

# ->飛行程序參數
TAKEOFF_SETTLE = 95             # 安定高度百分比(%，高度確認門檻)*
LANDING_CUTOFF = 0.05           # 油門關閉高度(m，停止油門輸出)*
TURN_SETTLE = 95                # 安定轉角百分比(%，停止轉向輸出)*
TURN_YAW_RATE = 30              # 轉角速度(degree/s)*
TURN_AREA_LIMIT = 3000          # 號誌燈號面積閥值(px)
DROP_POS_LIMIT_X = 100                      # 空投定位X範圍限制(單向，px)*
DROP_POS_LIMIT_Y = DROP_POS_LIMIT_X*(3/4)   # 空投定位Y範圍限制(單向，px)*
AREA_ARRIVE_LIMIT = 4000        # 色塊抵達偵測門檻(px)
AREA_LEAVE_LIMIT = 3000         # 色塊離開偵測門檻(px)

# ========== 影像參數設定 ===============================================
# ->相機設定參數
CAM_VIDEO = cv2.VideoCapture(0)             # 影像來源(預設來源為0)
CAM_IMAGE = 0                               # 儲存影像
CAM_IMAGE_LIVE = 0                          # 即時影像
CAM_IMAGE_WIDTH = 480                       # 影像寬度(px，4:3設定 )
CAM_IMAGE_HEIGHT = 360                      # 影像高度(px)
VIDEO_FOURCC = cv2.VideoWriter_fourcc(*'XVID')
VIDEO_SAVE = cv2.VideoWriter('Flight Camera.avi',
                             VIDEO_FOURCC, FREQ_CAM, (CAM_IMAGE_WIDTH, CAM_IMAGE_HEIGHT))

# ->影像計算參數
IMG_BLUE_L = np.array([85, 76, 90])         # 藍色色域-下限
IMG_BLUE_H = np.array([127, 255, 255])      # 藍色色域-上限
IMG_GREEN_L = np.array([44, 104, 27])       # 綠色色域-下限
IMG_GREEN_H = np.array([87, 255, 255])      # 綠色色域-上限
IMG_RED_L = np.array([160, 111, 56])        # 紅色色域-下限
IMG_RED_H = np.array([180, 255, 255])       # 紅色色域-上限
IMG_BLACK_L = np.array([0, 0, 0])           # 黑色色域-下限
IMG_BLACK_H = np.array([180, 255, 75])      # 黑色色域-上限
IMG_MORPHOLOGY_EX_KERNEL = (5, 5)           # 閉運算矩陣大小(奇數，越大降噪能力越強)*
IMG_AREA_LIMIT = 3000                       # 取用面積下限值*
IMG_SAMPLE_NUM_AVE = 10                     # 均值濾波取樣次數*
BODY_REF_X = int(CAM_IMAGE_WIDTH / 2)       # 畫面參考X座標中心
BODY_REF_Y = int(CAM_IMAGE_HEIGHT / 2)      # 畫面位置Y座標中心
BODY_OFFSET_X = 0                           # 機身中心X偏移量(px，左負右正)*
BODY_OFFSET_Y = 0                           # 機身中心Y偏移量(px，上正下負)*

# ->線角度計算
ANG_THRESHOLD = 40              # 二值化門檻值(調高比較能看到線，但可能有雜訊)*
ANG_THRESHOLD_NEW = 255         # 取得黑白照填入最高值
ANG_GB_KERNEL_SIZE = (7, 7)     # 高斯模糊矩陣大小(奇數，越大越模糊)*
ANG_GB_SIGMA = 2                # 高斯模糊色域標準差(越大越模糊)*
ANG_LINE_TRACK_AREA_L = 450     # 線追蹤型心計算面積下限*
ANG_SAMPLE_NUM_AVE = 10         # 航向控制器均值濾波取樣次數*
ANG_LIMIT = 0                   # 最大差角限制(degree)

# ========== 貨物艙參數 ================================================
# ->沙包盒參數
BOX_PWM_PIN = 12                # 沙包盒PWM輸出腳位
BOX_PWM_FREQ = 50               # PWM頻率(方波Hz)
BOX_PWM_OPEN = 1500             # 開門位置(400~2350us)*
BOX_PWM_CLOSE = 2000            # 關門位置(400~2350us)*

# ========== 高度感測器參數 =============================================
# ->高度感測器整合參數
WEIGHTED_SONAR = 2              # 超音波高度計權重*
WEIGHTED_IMU = 1                # IMU高度計權重*
WEIGHTED_LASER = 1              # 雷射測距儀權重*

# ->超音波高度計參數
SONAR_AIR_TEMP = 28             # 氣溫(攝氏)*
SONAR_TRIG_PIN = 2              # TRIG輸出腳位
SONAR_ECHO_PIN = 3              # ECHO輸入腳位
SONAR_TIMEOUT = 0.015           # 無回應超時限制(秒，過低時遠距離偵測將會變成超時)*
SONAR_OFFSET = 0                # 高度校正偏移量(m，向上為正)
SONAR_SAMPLE_NUM_CAL = 10       # 高度校正取樣次數*
SONAR_SAMPLE_NUM_AVE = 10       # 均值濾波取樣次數*
SONAR_KF_INPUT_MAX = 3          # 卡爾曼濾波器輸入值限制(高度，m)*
SONAR_KF_Q = 5                  # 卡爾曼濾波器滑順度(越小反應快，越不平；越大反應慢，但較平滑)*
SONAR_ERR_COUNT = 0             # 量測超時次數記錄
SONAR_FILTER_ALT = 1            # 突波濾波作用起始高度(m)
SONAR_MAX_CHANGE = 0.5          # 突波容許值(m)

# ->IMU高度計參數
IMU_OFFSET = 0                  # 高度校正偏移量(m，向上為正)
IMU_SAMPLE_NUM_CAL = 10         # 高度校正取樣次數*
IMU_SAMPLE_NUM_AVE = 10         # 均值濾波取樣次數*
IMU_KF_INPUT_MAX = 3            # 卡爾曼濾波器輸入值限制(高度，m)*
IMU_KF_Q = 5                    # 卡爾曼濾波器滑順度(越小反應快，越不平；越大反應慢，但較平滑)*

# ->雷射測距儀
LASER_OFFSET = 0                # 高度校正偏移量(m，向上為正)
LASER_SAMPLE_NUM_CAL = 3        # 高度校正取樣次數*
LASER_SAMPLE_NUM_AVE = 3        # 均值濾波取樣次數*
LASER_MODE = 1                  # 雷射量測模式(1:短 2:中 3:長)

# ========== 數值記錄 ==================================================
# ->高度控制器
R_TARGET_ALT = []                       # 記錄目標高
R_ACTUAL_ALT_CTRL = []                  # 記錄實際高
R_TARGET_THRUST = []                    # 記錄目標推力
R_ALT_ERROR = []                        # 記錄高度誤差
R_ALT_GP = []                           # 記錄P增益
R_ALT_GI = []                           # 記錄I增益
R_ALT_GD = []                           # 記錄D增益

# ->位置控制器
R_POS_BLOCK_X = []                      # 記錄X座標
R_POS_BLOCK_Y = []                      # 記錄Y座標
R_POS_BLOCK_AREA = []                   # 記錄面積變化
R_POS_RECORD_SCALE = 1000               # 面積縮放比
R_POS_X_ERROR = []                      # 記錄誤差值
R_POS_Y_ERROR = []                      # 記錄誤差值
R_POS_X_GP = []                         # 記錄P增益
R_POS_X_GI = []                         # 記錄I增益
R_POS_X_GD = []                         # 記錄D增益
R_POS_Y_GP = []                         # 記錄P增益
R_POS_Y_GI = []                         # 記錄I增益
R_POS_Y_GD = []                         # 記錄D增益
R_POS_X_TARGET_VEL = []                 # 記錄輸出的目標速度值
R_POS_Y_TARGET_VEL = []                 # 記錄輸出的目標速度值

# ->速度控制器
R_POS_TARGET_ROLL = []                  # 記錄目標滾轉角
R_POS_TARGET_PITCH = []                 # 記錄目標俯仰角
R_POS_ACTUAL_ROLL = []                  # 記錄實際滾轉角
R_POS_ACTUAL_PITCH = []                 # 記錄實際俯仰角
R_POS_VEL_X_TAR = []                    # 記錄X目標速度
R_POS_VEL_Y_TAR = []                    # 記錄Y目標速度
R_POS_VEL_X_ACT = []                    # 記錄X實際速度
R_POS_VEL_Y_ACT = []                    # 記錄Y實際速度
R_POS_VEL_X_ERROR = []                  # 記錄誤差值
R_POS_VEL_Y_ERROR = []                  # 記錄誤差值
R_POS_VEL_X_GP = []                     # 記錄P增益
R_POS_VEL_X_GI = []                     # 記錄I增益
R_POS_VEL_X_GD = []                     # 記錄D增益
R_POS_VEL_Y_GP = []                     # 記錄P增益
R_POS_VEL_Y_GI = []                     # 記錄I增益
R_POS_VEL_Y_GD = []                     # 記錄D增益

# ->航向控制器
R_HDG_ACTUAL = []                       # 記錄實際角度
R_HDG_TARGET_ROLL = []                  # 記錄滾轉輸出
R_HDG_TARGET_YAW_RATE = []              # 記錄目標偏航速率
R_HDG_ERROR = []                        # 記錄角度誤差
R_HDG_GP = []                           # 記錄P增益
R_HDG_GI = []                           # 記錄I增益
R_HDG_GD = []                           # 記錄D增益

# ->飛行速度控制器
R_SPD_TARGET = []                       # 記錄目標速度
R_SPD_ACTUAL = []                       # 記錄實際速度
R_SPD_TARGET_PITCH = []                 # 記錄目標俯仰角
R_SPD_ACTUAL_PITCH = []                 # 記錄實際俯仰角
R_SPD_ERROR = []                        # 記錄速度誤差
R_SPD_GP = []                           # 記錄P增益
R_SPD_GI = []                           # 記錄I增益
R_SPD_GD = []                           # 記錄D增益

# ->飛行高度感測器
R_ACTUAL_ALT = []                       # 記錄加權值
R_ACTUAL_ALT_SONAR = []                 # 記錄超音波高度計值
R_ACTUAL_ALT_BAROM = []                 # 記錄氣壓高度計值
R_ACTUAL_ALT_LASER = []                 # 記錄雷射測距高度值

# ->超音波高度計
R_ACTUAL_ALT_SONAR_RAW = []             # 記錄原始值
R_ACTUAL_ALT_SONAR_OFFSET = []          # 記錄補正值
R_ACTUAL_ALT_SONAR_AVE = []             # 記錄平均值
R_ACTUAL_ALT_SONAR_KF = []              # 記錄KF值

# ->氣壓高度計
R_ACTUAL_ALT_IMU_RAW = []               # 記錄原始值
R_ACTUAL_ALT_IMU_OFFSET = []            # 記錄補正值
R_ACTUAL_ALT_IMU_AVE = []               # 記錄平均值
R_ACTUAL_ALT_IMU_KF = []                # 記錄KF值

print('<啟動程序> 載入全域變數設定完畢')

# ========== 與UAV執行連線 =============================================
print('<啟動程序> 與UAV連線中，請稍候...')
# 樹莓派連線指令
uav = connect(CONNECT_PORT, wait_ready=True, baud=CONNECT_BAUD)
# 電腦連線指令
# uav = connect(CONNECT_PORT, baud=CONNECT_BAUD)
print('<啟動程序> UAV已連線')

# ====================================================================
# ========== 背景主函式函式區塊 =========================================
# ====================================================================
# 本區為主要運作的函式，以控制器為主，包含高度、位置、航向、雙高度感測器、
# 相機、命令傳送，多線程執行時能達到協作的處理功能
# (但仍受每個迴圈處理時間不一影響，為非同步處理)


# ========== 高度控制器 ================================================
# 說明：
# 為PID控制器型態，於任務期間全程作用，輸入目標高度後，依據高度感測器執行誤差PID增益，
# 最終輸出成爬升速率型態(符合MavLink指令)，設有輸出範圍限制器、中點飽和區間功能
def altitude_controller():
    global TARGET_THRUST, STATE_ALT
    # PID變數設定
    int_error = 0
    previous_error = 0
    previous_time = time.time()

    # 等待感測器啟動
    if not STATE_SENSOR_ALT:
        print('<啟動程序> 高度控制器：等待高度感測器啟動')
    while not STATE_SENSOR_ALT:
        time.sleep(0.1)

    # 回報高度控制器啟動
    STATE_ALT = True
    print('<啟動程序> 高度控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 同捆執行起點
        THREAD_OCCUPY.acquire()

        # 計算誤差值
        error = TARGET_ALT - ACTUAL_ALT

        # 誤差值微積分計算
        current_time = time.time()
        delta_error = error - previous_error
        delta_time = current_time - previous_time
        int_error += (error * delta_time)
        derivative = delta_error / delta_time

        # 誤差增益處理
        gain_p = error * ALT_KP
        gain_i = int_error * ALT_KI
        gain_d = derivative * ALT_KD

        if gain_d > ALT_GAIN_D_LIMIT:
            gain_d = ALT_GAIN_D_LIMIT
        elif gain_d < -ALT_GAIN_D_LIMIT:
            gain_d = -ALT_GAIN_D_LIMIT
        speed_out = gain_p + gain_i + gain_d

        # 將速度以比例輸出
        thrust_out = speed_out / ALT_SPEED

        # 補正至中點
        thrust_out += 0.5

        # 輸出限制(0下降max ~ 1上升max)
        if thrust_out > 1:
            thrust_out = 1
        if thrust_out < 0:
            thrust_out = 0

        # 更新推力值至全域變數
        TARGET_THRUST = thrust_out

        # 傳遞值給下次迴圈使用
        previous_error = error
        previous_time = current_time

        # 記錄數值
        R_TARGET_ALT.append(TARGET_ALT)
        R_ACTUAL_ALT_CTRL.append(ACTUAL_ALT)
        R_TARGET_THRUST.append(TARGET_THRUST)
        R_ALT_ERROR.append(error)
        R_ALT_GP.append(gain_p)
        R_ALT_GI.append(gain_i)
        R_ALT_GD.append(gain_d)

        # 同捆執行終點
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        time.sleep(1 / FREQ_ALT)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_ALT = False
    print('<關閉程序> 高度控制器：關閉')


# ========== 位置控制器 ================================================
# 說明：
# 外迴圈的位置控制
# 先以影像取得的色塊座標計算速度增益值，在輸出至速度控制器做角度-加速度控制
def position_controller():
    global TARGET_VEL, STATE_POS
    # PID變數設定
    pos_int_error_x = 0
    pos_int_error_y = 0
    pos_previous_error_x = 0
    pos_previous_error_y = 0
    previous_time = time.time()

    # 均值濾波設定
    vel_x_cmd_register = []
    vel_y_cmd_register = []

    # 等待影像感測器啟動
    if not STATE_SENSOR_CAM:
        print('<啟動程序> 位置控制器：等待影像感測器啟動')
    while not STATE_SENSOR_CAM:
        time.sleep(0.1)

    # 回報位置控制器啟用
    STATE_POS = True
    print('<啟動程序> 位置控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        while MODE == 'POS' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 讀取回授值
            pos_actual_x = ACTUAL_POS[0]
            pos_actual_y = ACTUAL_POS[1]

            # 回授值截止區修正
            if TARGET_POS[0] - POS_DZ < pos_actual_x < TARGET_POS[0] + POS_DZ:
                pos_actual_x = TARGET_POS[0]
            if TARGET_POS[1] - POS_DZ < pos_actual_y < TARGET_POS[1] + POS_DZ:
                pos_actual_y = TARGET_POS[1]

            # 計算誤差值
            pos_error_x = TARGET_POS[0] - pos_actual_x
            pos_error_y = TARGET_POS[1] - pos_actual_y

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            pos_delta_error_x = pos_error_x - pos_previous_error_x
            pos_int_error_x += (pos_error_x * delta_time)
            pos_derivative_x = pos_delta_error_x / delta_time

            pos_delta_error_y = pos_error_y - pos_previous_error_y
            pos_int_error_y += (pos_error_y * delta_time)
            pos_derivative_y = pos_delta_error_y / delta_time

            # X軸誤差增益處理
            pos_gain_p_x = pos_error_x * POS_X_KP
            pos_gain_i_x = pos_int_error_x * POS_X_KI
            pos_gain_d_x = pos_derivative_x * POS_X_KD
            vel_x_out = pos_gain_p_x + pos_gain_i_x + pos_gain_d_x

            # Y軸誤差增益處理
            pos_gain_p_y = pos_error_y * POS_Y_KP
            pos_gain_i_y = pos_int_error_y * POS_Y_KI
            pos_gain_d_y = pos_derivative_y * POS_Y_KD
            vel_y_out = pos_gain_p_y + pos_gain_i_y + pos_gain_d_y

            # X速度命令均值濾波器
            vel_x_cmd_register.insert(0, vel_x_out)
            if len(vel_x_cmd_register) > POS_SAMPLE_NUM_AVE:
                vel_x_cmd_register.pop()
            vel_x_cmd = np.average(vel_x_cmd_register)

            # Y速度命令均值濾波器
            vel_y_cmd_register.insert(0, vel_y_out)
            if len(vel_y_cmd_register) > POS_SAMPLE_NUM_AVE:
                vel_y_cmd_register.pop()
            vel_y_cmd = np.average(vel_y_cmd_register)

            # 速度輸出限制(原則上不使用)
            if vel_x_cmd > POS_VEL_LIMIT:
                vel_x_cmd = POS_VEL_LIMIT
            elif vel_x_cmd < -POS_VEL_LIMIT:
                vel_x_cmd = -POS_VEL_LIMIT

            if vel_y_cmd > POS_VEL_LIMIT:
                vel_y_cmd = POS_VEL_LIMIT
            elif vel_y_cmd < -POS_VEL_LIMIT:
                vel_y_cmd = -POS_VEL_LIMIT

            # 更新速度命令值至全域變數
            TARGET_VEL = [vel_x_cmd, vel_y_cmd]

            # 傳遞值給下次迴圈使用
            pos_previous_error_x = pos_error_x
            pos_previous_error_y = pos_error_y
            previous_time = current_time

            # 記錄數值
            R_POS_BLOCK_X.append(ACTUAL_POS[0])
            R_POS_BLOCK_Y.append(ACTUAL_POS[1])
            R_POS_BLOCK_AREA.append(ACTUAL_AREA / R_POS_RECORD_SCALE)

            R_POS_X_ERROR.append(pos_error_x)
            R_POS_Y_ERROR.append(pos_error_y)
            R_POS_X_GP.append(pos_gain_p_x)
            R_POS_X_GI.append(pos_gain_i_x)
            R_POS_X_GD.append(pos_gain_d_x)
            R_POS_Y_GP.append(pos_gain_p_y)
            R_POS_Y_GI.append(pos_gain_i_y)
            R_POS_Y_GD.append(pos_gain_d_y)

            R_POS_X_TARGET_VEL.append(vel_x_cmd)
            R_POS_Y_TARGET_VEL.append(vel_y_cmd)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            time.sleep(1 / FREQ_POS)

            # 當飛行模式不為定位時，跳出迴圈並暫停函式
            if not MODE == 'POS':
                # 歸零速度命令數值
                TARGET_VEL = [0, 0]
                print('<飛行狀態> 位置控制器：暫停')

        # 重設暫存器
        vel_x_cmd_register = []
        vel_y_cmd_register = []
        pos_int_error_x = 0
        pos_int_error_y = 0
        pos_previous_error_x = 0
        pos_previous_error_y = 0
        previous_time = time.time()

        # 記錄數值
        R_POS_BLOCK_X.append(0)
        R_POS_BLOCK_Y.append(0)
        R_POS_BLOCK_AREA.append(0)

        R_POS_X_ERROR.append(0)
        R_POS_Y_ERROR.append(0)
        R_POS_X_GP.append(0)
        R_POS_X_GI.append(0)
        R_POS_X_GD.append(0)
        R_POS_Y_GP.append(0)
        R_POS_Y_GI.append(0)
        R_POS_Y_GD.append(0)

        R_POS_X_TARGET_VEL.append(0)
        R_POS_Y_TARGET_VEL.append(0)

        # 迴圈執行間隔
        time.sleep(1 / FREQ_POS)

        # 當飛行模式為定位時，跳入迴圈並啟動函式
        if MODE == 'POS' and STATE_THREAD:
            print('<飛行狀態> 位置控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_POS = False
    print('<關閉程序> 位置控制器：關閉')


# ========== 位置-速度控制器 ============================================
# 說明：
# 位置控制器的內迴圈，負責控制與色塊速度
def position_velocity_controller():
    global TARGET_PITCH, TARGET_ROLL, TARGET_VEL, ACTUAL_POS, ACTUAL_AREA, STATE_VEL
    # PID變數設定
    vel_int_error_x = 0
    vel_int_error_y = 0
    vel_previous_error_x = 0
    vel_previous_error_y = 0
    previous_time = time.time()

    # 等待位置感測器啟動
    if not STATE_POS:
        print('<啟動程序> 速度控制器：等待位置控制器啟動')
    while not STATE_POS:
        time.sleep(0.1)

    # 回報速度控制器啟用
    STATE_VEL = True
    print('<啟動程序> 速度控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        while MODE == 'POS' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 誤差增益
            vel_actual_x = ACTUAL_VEL[0] * POS_VEL_X_FEEDBACK_GAIN
            vel_actual_y = ACTUAL_VEL[1] * POS_VEL_Y_FEEDBACK_GAIN

            # 計算誤差值
            vel_error_x = TARGET_VEL[0] - vel_actual_x
            vel_error_y = TARGET_VEL[1] - vel_actual_y

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            vel_delta_error_x = vel_error_x - vel_previous_error_x
            vel_int_error_x += (vel_error_x * delta_time)
            vel_derivative_x = vel_delta_error_x / delta_time

            vel_delta_error_y = vel_error_y - vel_previous_error_y
            vel_int_error_y += (vel_error_y * delta_time)
            vel_derivative_y = vel_delta_error_y / delta_time

            # X軸誤差增益處理
            vel_gain_p_x = vel_error_x * POS_VEL_X_KP
            vel_gain_i_x = vel_int_error_x * POS_VEL_X_KI
            vel_gain_d_x = vel_derivative_x * POS_VEL_X_KD
            roll_out = vel_gain_p_x + vel_gain_i_x + vel_gain_d_x

            # Y軸誤差增益處理
            vel_gain_p_y = vel_error_y * POS_VEL_Y_KP
            vel_gain_i_y = vel_int_error_y * POS_VEL_Y_KI
            vel_gain_d_y = vel_derivative_y * POS_VEL_Y_KD
            pitch_out = vel_gain_p_y + vel_gain_i_y + vel_gain_d_y

            # 角度限制(飽和區限幅與截止區處理)
            if roll_out > POS_ANGLE_LIMIT:
                roll_out = POS_ANGLE_LIMIT
            elif roll_out < -POS_ANGLE_LIMIT:
                roll_out = -POS_ANGLE_LIMIT
            elif -POS_ANGLE_CUTOFF < roll_out < POS_ANGLE_CUTOFF:
                roll_out = 0

            if pitch_out > POS_ANGLE_LIMIT:
                pitch_out = POS_ANGLE_LIMIT
            elif pitch_out < -POS_ANGLE_LIMIT:
                pitch_out = -POS_ANGLE_LIMIT
            elif -POS_ANGLE_CUTOFF < pitch_out < POS_ANGLE_CUTOFF:
                pitch_out = 0

            # 輸出至全域變數
            if ACTUAL_ALT < POS_VEL_CUTOFF_ALT:
                TARGET_ROLL = 0
                TARGET_PITCH = 0
            else:
                TARGET_ROLL = -roll_out
                TARGET_PITCH = pitch_out

            # 傳遞值給下次迴圈使用
            vel_previous_error_x = vel_error_x
            vel_previous_error_y = vel_error_y
            previous_time = current_time

            # 記錄數值
            R_POS_VEL_X_TAR.append(TARGET_VEL[0])
            R_POS_VEL_Y_TAR.append(TARGET_VEL[1])
            R_POS_VEL_X_ACT.append(vel_actual_x)
            R_POS_VEL_Y_ACT.append(vel_actual_y)

            R_POS_VEL_X_ERROR.append(vel_error_x)
            R_POS_VEL_Y_ERROR.append(vel_error_y)
            R_POS_VEL_X_GP.append(vel_gain_p_x)
            R_POS_VEL_X_GI.append(vel_gain_i_x)
            R_POS_VEL_X_GD.append(vel_gain_d_x)
            R_POS_VEL_Y_GP.append(vel_gain_p_y)
            R_POS_VEL_Y_GI.append(vel_gain_i_y)
            R_POS_VEL_Y_GD.append(vel_gain_d_y)

            R_POS_TARGET_ROLL.append(TARGET_ROLL)
            R_POS_TARGET_PITCH.append(TARGET_PITCH)
            R_POS_ACTUAL_ROLL.append(math.degrees(uav.attitude.roll))
            R_POS_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            time.sleep(1 / FREQ_VEL)

            # 當飛行模式不為定位時，跳出迴圈並暫停函式
            if not MODE == 'POS':
                # 歸零姿態數值
                TARGET_ROLL = 0
                TARGET_PITCH = 0
                print('<飛行狀態> 速度控制器：暫停')

        # 重設暫存器
        vel_int_error_x = 0
        vel_int_error_y = 0
        vel_previous_error_x = 0
        vel_previous_error_y = 0
        previous_time = time.time()

        # 記錄值
        R_POS_VEL_X_TAR.append(0)
        R_POS_VEL_Y_TAR.append(0)
        R_POS_VEL_X_ACT.append(0)
        R_POS_VEL_Y_ACT.append(0)

        R_POS_VEL_X_ERROR.append(0)
        R_POS_VEL_Y_ERROR.append(0)
        R_POS_VEL_X_GP.append(0)
        R_POS_VEL_X_GI.append(0)
        R_POS_VEL_X_GD.append(0)
        R_POS_VEL_Y_GP.append(0)
        R_POS_VEL_Y_GI.append(0)
        R_POS_VEL_Y_GD.append(0)

        R_POS_TARGET_ROLL.append(TARGET_ROLL)
        R_POS_TARGET_PITCH.append(TARGET_PITCH)
        R_POS_ACTUAL_ROLL.append(math.degrees(uav.attitude.roll))
        R_POS_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))

        # 迴圈執行間隔
        time.sleep(1 / FREQ_VEL)

        # 當飛行模式為定位時，跳入迴圈並啟動函式
        if MODE == 'POS' and STATE_THREAD:
            print('<飛行狀態> 速度控制器：啟動')

        # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_VEL = False
    print('<關閉程序> 速度控制器：關閉')


# =========== 航向控制器 ===============================================
# 說明：
# 採用PID控制器型態，每一迴圈讀取一次影像計算線條型心，並將型心值做均值綠波
# 濾波完畢後再以該值計算出型心至畫面下方中心之夾角，並以此值為控制器的誤差值
# 以該值來做PID增益處理，輸出角速度來追尋角度，此控制器可處理平行、轉角、Ｔ路口轉向
# 為不可逆型循跡控制器，參數可調最大角速度輸出、切斷區間、濾波等級、面積門檻值
# 使用Global-Frame參照，以地面標線為參照座標，預測為PD控制器(前進速度控制分離)
def heading_controller():
    global TARGET_ROLL, TARGET_YAW_RATE, STATE_HDG
    # PID變數設定
    int_error = 0
    previous_error = 0
    previous_time = time.time()

    # 等待相機啟動
    if not STATE_SENSOR_CAM:
        print('<啟動程序> 位置控制器：等待影像感測器啟動')
    while not STATE_SENSOR_CAM:
        time.sleep(0.1)

    # 回報定位控制器啟用
    STATE_HDG = True
    print('<啟動程序> 航向控制器：啟動')

    # 程式迴圈
    while STATE_THREAD:
        # 定位作動迴圈(同捆執行模式，於循跡模式作動)
        while MODE == 'LINE' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 讀取回授值
            hdg_actual = ACTUAL_HDG

            # 回授值截止區修正
            if HDG_DZ > hdg_actual > - HDG_DZ:
                hdg_actual = 0

            # 計算誤差值
            error = -hdg_actual

            # 誤差值微積分計算
            current_time = time.time()
            delta_error = error - previous_error
            delta_time = current_time - previous_time
            int_error += (error * delta_time)
            derivative = delta_error / delta_time

            # 誤差增益處理
            gain_p = error * HDG_KP
            gain_i = int_error * HDG_KI
            gain_d = derivative * HDG_KD
            gain_out = gain_p + gain_i + gain_d

            # 輸出限制
            if math.fabs(ACTUAL_ANG_LINE) > ANG_LIMIT:
                gain_out = 0

            # 輸出模式
            if HDG_MODE == 'R':
                roll_out = gain_out
                angular_rate = 0
            elif HDG_MODE == 'Y':
                roll_out = 0
                angular_rate = gain_out
            else:
                roll_out = gain_out * HDG_ROLL_YAW_RATIO
                angular_rate = gain_out

            # 角度上下限制(包含截止區設定)
            if roll_out > HDG_ROLL_ANGLE_LIMIT:
                roll_out = HDG_ROLL_ANGLE_LIMIT
            elif roll_out < -HDG_ROLL_ANGLE_LIMIT:
                roll_out = -HDG_ROLL_ANGLE_LIMIT
            elif -HDG_ROLL_ANGLE_CUTOFF < roll_out < HDG_ROLL_ANGLE_CUTOFF:
                roll_out = 0

            # 角速度上下限制
            if angular_rate > HDG_YAW_RATE_LIMIT:
                angular_rate = HDG_YAW_RATE_LIMIT
            elif angular_rate < -HDG_YAW_RATE_LIMIT:
                angular_rate = -HDG_YAW_RATE_LIMIT

            # 更新目標值至全域變數
            TARGET_ROLL = roll_out
            TARGET_YAW_RATE = angular_rate

            # 傳遞值給下次迴圈使用
            previous_error = error
            previous_time = current_time

            # 記錄數值
            R_HDG_ACTUAL.append(hdg_actual)
            R_HDG_TARGET_ROLL.append(TARGET_ROLL)
            R_HDG_TARGET_YAW_RATE.append(TARGET_YAW_RATE)

            R_HDG_ERROR.append(error)
            R_HDG_GP.append(gain_p)
            R_HDG_GI.append(gain_i)
            R_HDG_GD.append(gain_d)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            time.sleep(1 / FREQ_HDG)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if not MODE == 'LINE':
                # 歸零姿態數值
                TARGET_ROLL = 0
                TARGET_YAW_RATE = 0
                STATE_HDG = False
                print('<飛行狀態> 航向控制器：暫停')

        # 無循跡作動區間
        # 回報迴圈休眠狀態
        if MODE == 'POS' and STATE_HDG:
            STATE_HDG = False
            print('<飛行狀態> 航向控制器：暫停')

        # 重設暫存器
        int_error = 0
        previous_error = 0
        previous_time = time.time()

        # 記錄數值
        R_HDG_ACTUAL.append(0)
        R_HDG_TARGET_ROLL.append(0)
        R_HDG_TARGET_YAW_RATE.append(0)

        R_HDG_ERROR.append(0)
        R_HDG_GP.append(0)
        R_HDG_GI.append(0)
        R_HDG_GD.append(0)

        # 迴圈執行間隔
        time.sleep(1 / FREQ_HDG)

        # 當飛行模式為循跡時，跳入迴圈並啟動函式
        if MODE == 'LINE' and STATE_THREAD:
            STATE_HDG = True
            print('<飛行狀態> 航向控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_HDG = False
    print('<關閉程序> 航向控制器：關閉')


# =========== 飛行速度控制器 ============================================
# 說明：
# 於循線期間控制往前飛行的速度
# 不一定會使用(視感測器響應情形而定)
def flight_speed_controller():
    global TARGET_PITCH, STATE_SPD
    # PID變數設定
    int_error = 0
    previous_error = 0
    previous_time = time.time()

    # 等待飛行速度感測器啟動
    if not STATE_SENSOR_SPD:
        print('<啟動程序> 飛行速度控制器：等待飛行速度感測器啟動')
    while not STATE_SENSOR_SPD:
        time.sleep(0.1)

    # 回報飛行速度控制器啟用
    STATE_SPD = True
    print('<啟動程序> 飛行速度控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        while MODE == 'LINE' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 讀取目標值
            spd_target = TARGET_SPD

            # 目標輸入限制
            if spd_target > SPD_LIMIT:
                spd_target = SPD_LIMIT
            if spd_target < -SPD_LIMIT:
                spd_target = -SPD_LIMIT

            # 計算誤差值
            error = spd_target - ACTUAL_SPD

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            delta_error = error - previous_error
            int_error += (error * delta_time)
            derivative = delta_error / delta_time

            # 誤差增益處理
            gain_p = error * SPD_KP
            gain_i = int_error * SPD_KI
            gain_d = derivative * SPD_KD
            pitch_out = gain_p + gain_i + gain_d

            # 角度限制(包含截止區設定)
            if pitch_out > SPD_ANGLE_LIMIT:
                pitch_out = SPD_ANGLE_LIMIT
            elif pitch_out < -SPD_ANGLE_LIMIT:
                pitch_out = -SPD_ANGLE_LIMIT
            elif -SPD_ANGLE_CUTOFF < pitch_out < SPD_ANGLE_CUTOFF:
                pitch_out = 0

            # 輸出至全域變數
            TARGET_PITCH = -pitch_out

            # 傳遞值給下次迴圈使用
            previous_error = error
            previous_time = current_time

            # 記錄數值
            R_SPD_TARGET.append(spd_target)
            R_SPD_ACTUAL.append(ACTUAL_SPD)
            R_SPD_TARGET_PITCH.append(TARGET_PITCH)
            R_SPD_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))

            R_SPD_ERROR.append(error)
            R_SPD_GP.append(gain_p)
            R_SPD_GI.append(gain_i)
            R_SPD_GD.append(gain_d)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            time.sleep(1 / FREQ_SPD)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if not MODE == 'LINE':
                # 切換跳出姿態數值(慣性抵銷用)
                TARGET_PITCH = SPD_OFF_ANGLE
                STATE_SPD = False
                print('<飛行狀態> 飛行速度控制器：暫停')

        # 無循跡作動區間
        # 回報迴圈休眠狀態
        if MODE == 'POS' and STATE_SPD:
            STATE_SPD = False
            print('<飛行狀態> 飛行速度控制器：暫停')

        # 重設暫存器
        int_error = 0
        previous_error = 0
        previous_time = time.time()

        # 記錄值
        R_SPD_TARGET.append(0)
        R_SPD_ACTUAL.append(0)
        R_SPD_TARGET_PITCH.append(0)
        R_SPD_ACTUAL_PITCH.append(0)

        R_SPD_ERROR.append(0)
        R_SPD_GP.append(0)
        R_SPD_GI.append(0)
        R_SPD_GD.append(0)

        # 迴圈執行間隔
        time.sleep(1 / FREQ_SPD)

        # 當飛行模式為循跡時，跳入迴圈並啟動函式
        if MODE == 'LINE' and STATE_THREAD:
            print('<飛行狀態> 飛行速度控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_SPD = False
    print('<關閉程序> 飛行速度控制器：關閉')


# ========== 影像擷取程式 ==============================================
# 負責截取機載影像畫面，儲存至全域值供控制器使用
# 因主線程衝突故將畫面顯示整合至GUI介面中方便瀏覽
# 另將飛行影像存成影片以供記錄(處理後之影像若需錄下，須經格式轉換處理，暫不增加)
def camera():
    global CAM_IMAGE, STATE_CAM, STATE_THREAD
    # 程式迴圈(同捆執行模式)
    while CAM_VIDEO.isOpened() and STATE_THREAD:
        # 確認影像擷取狀態
        state_grab = CAM_VIDEO.grab()

        # 擷取成功
        if state_grab:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 擷取影像並儲存
            state_retrieve, image = CAM_VIDEO.retrieve()
            CAM_IMAGE = cv2.resize(image, (CAM_IMAGE_WIDTH, CAM_IMAGE_HEIGHT))

            # 寫入錄影檔
            if MODE != '':
                # 從灰階轉換回BGR色域
                img_origin = cv2.cvtColor(CAM_IMAGE_LIVE, cv2.COLOR_GRAY2BGR)
                VIDEO_SAVE.write(img_origin)
            else:
                VIDEO_SAVE.write(CAM_IMAGE)

            # 回報影像擷取程式啟動
            if not STATE_CAM:
                STATE_CAM = True
                print('<啟動程序> 影像擷取程式：啟動\n')

            # 顯示影像視窗
            if MODE == 'POS' or MODE == 'LINE':
                cv2.imshow("Flight Camera", CAM_IMAGE_LIVE)
            else:
                cv2.imshow("Flight Camera", CAM_IMAGE)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 手動關閉影像顯示
            if cv2.waitKey(int(1000 / FREQ_CAM)) & 0xff == ord("q"):
                STATE_THREAD = False
                print('<關閉程序> 已手動關閉相機，請手動重啟程式')
                break

        # 擷取失敗
        else:
            print('<錯誤> 影像擷取程式：無法擷取影像')
            STATE_THREAD = False
            print('<嚴重錯誤> 請檢查相機並手動重啟程式!!!!')
            break

    # 當系統狀態為否時，跳出迴圈並關閉函式
    CAM_VIDEO.release()
    VIDEO_SAVE.release()
    cv2.destroyAllWindows()  # 關閉影像顯示
    STATE_CAM = False
    print('<關閉程序> 影像擷取程式：關閉')


# ========== 影像感測器 ===============================================
# 負責計算影像的位置、速度、角度值，結合位置與循線模式
def camera_sensor():
    global STATE_SENSOR_CAM, ACTUAL_POS, ACTUAL_VEL, ACTUAL_AREA, ACTUAL_HDG, ACTUAL_ANG_LINE
    # 均值濾波器設定
    pos_x_register = []
    pos_y_register = []
    vel_x_register = []
    vel_y_register = []
    hdg_register = []

    # 速度值計算初始設定
    pos_x_old = 0
    pos_y_old = 0
    vel_x = 0
    vel_y = 0
    previous_time = time.time()

    # 等待影像擷取程式啟動
    if not STATE_CAM:
        print('<啟動程序> 影像感測器：等待影像擷取程式啟動')
    while not STATE_CAM:
        time.sleep(0.1)

    # 回報影像感測器啟用
    STATE_SENSOR_CAM = True
    print('<啟動程序> 影像感測器：啟動')

    # 程式迴圈
    while STATE_THREAD:
        # 定位模式計算迴圈(同捆執行模式)
        while MODE == 'POS' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 取得色塊XY座標
            x_raw, y_raw, block_area = image_moment(POS_COLOR)

            # 位置出界狀態值保留
            if x_raw == 0 and math.fabs(pos_x_old) > CAM_IMAGE_WIDTH*0.25:
                x_raw = pos_x_old
                outbound_x = True
            else:
                outbound_x = False

            if y_raw == 0 and math.fabs(pos_y_old) > CAM_IMAGE_HEIGHT*0.25:
                y_raw = pos_y_old
                outbound_y = True
            else:
                outbound_y = False

            # X座標均值濾波器
            pos_x_register.insert(0, x_raw)
            if len(pos_x_register) > IMG_SAMPLE_NUM_AVE:
                pos_x_register.pop()
            pos_x = np.average(pos_x_register)

            # Y座標均值濾波器
            pos_y_register.insert(0, y_raw)
            if len(pos_y_register) > IMG_SAMPLE_NUM_AVE:
                pos_y_register.pop()
            pos_y = np.average(pos_y_register)

            # 計算型心移動速度
            current_time = time.time()
            delta_time = current_time - previous_time

            pos_delta_x = pos_x - pos_x_old
            pos_delta_y = pos_y - pos_y_old
            vel_x_raw = pos_delta_x / delta_time
            vel_y_raw = pos_delta_y / delta_time

            # 速度出界狀態值保留
            if outbound_x:
                vel_x_raw = vel_x

            if outbound_y:
                vel_y_raw = vel_y

            # X速度均值濾波器
            vel_x_register.insert(0, vel_x_raw)
            if len(vel_x_register) > IMG_SAMPLE_NUM_AVE:
                vel_x_register.pop()
            vel_x = np.average(vel_x_register)

            # Y速度均值濾波器
            vel_y_register.insert(0, vel_y_raw)
            if len(vel_y_register) > IMG_SAMPLE_NUM_AVE:
                vel_y_register.pop()
            vel_y = np.average(vel_y_register)

            # 輸出數值至全域變數
            ACTUAL_POS = [pos_x, pos_y]
            ACTUAL_VEL = [vel_x, vel_y]
            ACTUAL_AREA = block_area

            # 傳遞值給下次迴圈使用
            pos_x_old = pos_x
            pos_y_old = pos_y
            previous_time = current_time

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            time.sleep(1 / FREQ_SENSOR_CAM)

            # 當飛行模式不為定位時，跳出迴圈並暫停函式
            if MODE == 'LINE':
                # 歸零數值
                ACTUAL_POS = [0, 0]
                ACTUAL_VEL = [0, 0]
                ACTUAL_AREA = 0
                print('<飛行狀態> 影像感測器：切換至角度輸出')

        # 循跡模式計算迴圈(同捆執行模式)
        while MODE == 'LINE' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 取得線條型心位置
            pos_x_raw, pos_y_raw, line_angle = line_moment()

            ACTUAL_ANG_LINE = line_angle

            # 型心位置均值濾波器
            # X座標均值濾波器
            pos_x_register.insert(0, pos_x_raw)
            if len(pos_x_register) > ANG_SAMPLE_NUM_AVE:
                pos_x_register.pop()
            pos_x = np.average(pos_x_register)

            # Y座標均值濾波器
            pos_y_register.insert(0, pos_y_raw)
            if len(pos_y_register) > ANG_SAMPLE_NUM_AVE:
                pos_y_register.pop()
            pos_y = np.average(pos_y_register)

            # 取得現在角度值(中心為目標值，朝右正，朝左負，使用global Frame參照)
            hdg_raw = angle_calculate(pos_x, pos_y)

            # 誤差角均值濾波器
            hdg_register.insert(0, hdg_raw)
            if len(hdg_register) > ANG_SAMPLE_NUM_AVE:
                hdg_register.pop()
            hdg = np.average(hdg_register)

            # 輸出角度誤差
            ACTUAL_HDG = hdg

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            time.sleep(1 / FREQ_SENSOR_CAM)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if MODE == 'POS':
                # 歸零數值
                ACTUAL_HDG = 0
                print('<飛行狀態> 影像感測器：切換至位置-速度輸出')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_CAM = False
    print('<關閉程序> 影像感測器：關閉')


# =========== 高度感測器整合程式 =========================================
# 說明：
# 負責整合超音波高度計、飛控氣壓計(可能為IMU來源)，做訊號整理給控制器使用
# 各訊號先經均值濾波後，再用卡爾曼濾波處理，最後再針對所有來源做加權平均
# 本函式處理氣壓計之濾波、高度感測器整合濾波
# 1021 增加VL53L1X雷射飛時測距(ToF)
def altitude_sensor():
    global ACTUAL_ALT, ACTUAL_ALT_IMU, ACTUAL_ALT_LASER, STATE_SENSOR_ALT
    # 平均降噪設定
    height_register_imu = []
    height_register_laser = []

    # 卡爾曼濾波器初始化(IMU高度計用)
    x_o = 0                     # 前一次狀態
    p_o = 0                     # 前一次斜方差
    z_o = 0                     # 前一次量測值
    q = math.exp(-IMU_KF_Q)     # 斜方差噪音值
    r = 2.92 * math.exp(-3)     # 斜方差量測噪音值(定值不更改)

    # 執行IMU高度計校正程序
    imu_calibrate()

    # 初始化雷射測距儀
    laser_calibrate()

    # 重新建立雷射量測物件
    laser = VL53L1X.VL53L1X(i2c_bus=1, i2c_address=0x29)
    laser.open()
    laser.start_ranging(LASER_MODE)

    # 等待超音波高度計啟動
    if not STATE_SENSOR_SONAR:
        print('<啟動程序> 高度感測器：等待超音波高度計啟動\n')
    while not STATE_SENSOR_SONAR:
        time.sleep(0.1)

    # 回報高度感測器啟動
    STATE_SENSOR_ALT = True
    print('<啟動程序> 高度感測器：啟動\n')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 同捆執行起點
        THREAD_OCCUPY.acquire()

        # 讀取超音波高度計資料
        height_sonar = ACTUAL_ALT_SONAR

        # IMU高度計數據處理
        # 卡爾曼濾波-預測
        x_p = x_o
        p_p = p_o + q

        # IMU讀值與補正
        height_raw_imu = uav.location.global_relative_frame.alt + IMU_OFFSET

        # 雷射測距儀讀值與補正
        height_raw_laser = laser.get_distance() / 1000 + LASER_OFFSET

        # IMU均值濾波器
        height_register_imu.insert(0, height_raw_imu)
        if len(height_register_imu) > IMU_SAMPLE_NUM_AVE:
            height_register_imu.pop()
        height_imu_ave = np.average(height_register_imu)

        # 雷射測距儀均值濾波器
        height_register_laser.insert(0, height_raw_laser)
        if len(height_register_laser) > LASER_SAMPLE_NUM_AVE:
            height_register_laser.pop()
        height_laser_ave = np.average(height_register_laser)

        # 卡爾曼濾波器輸入限制(防止上衝突波)
        if height_imu_ave <= IMU_KF_INPUT_MAX:
            z = height_imu_ave          # 若低於限制則使用新值
        else:
            z = z_o                     # 若超出限制則使用舊值

        # 卡爾曼濾波-更新
        k = p_p / (p_p + r)             # 卡爾曼增益
        x = x_p + k * (z - x_p)         # 狀態(濾波輸出值)
        p = (1 - k) * p_p               # 斜方差

        # 更新高度值至變數
        height_imu = x
        ACTUAL_ALT_IMU = height_imu
        ACTUAL_ALT_LASER = height_laser_ave

        # 卡爾曼濾波：傳遞值給下次迴圈使用
        x_o = x
        p_o = p
        z_o = z

        # 加權平均(可調整加權值)
        alt_num = (WEIGHTED_SONAR * height_sonar + WEIGHTED_IMU * height_imu
                   + WEIGHTED_LASER * height_laser_ave)
        alt_den = (WEIGHTED_SONAR + WEIGHTED_IMU + WEIGHTED_LASER)
        alt_ave = alt_num / alt_den

        # 更新高度值至全域變數
        ACTUAL_ALT = alt_ave

        # 記錄數值
        R_ACTUAL_ALT_SONAR.append(height_sonar)         # 記錄超音波高度計值
        R_ACTUAL_ALT_LASER.append(height_laser_ave)     # 記錄雷射測距高度
        R_ACTUAL_ALT_IMU_RAW.append(uav.location.global_relative_frame.alt)  # 記錄氣壓高度計原始值
        R_ACTUAL_ALT_IMU_OFFSET.append(height_raw_imu)  # 記錄補正值
        R_ACTUAL_ALT_IMU_AVE.append(height_imu_ave)     # 記錄均值濾波值
        R_ACTUAL_ALT_IMU_KF.append(height_imu)          # 記錄KF值
        R_ACTUAL_ALT.append(ACTUAL_ALT)                 # 記錄總高度值

        # 同捆執行終點
        THREAD_OCCUPY.release()

        time.sleep(1 / FREQ_SENSOR_ALT)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_ALT = False
    print('<關閉程序> 高度感測器：關閉\n')


# ========== 超音波高度計 ==============================================
# 說明：
# 使用HC-SR04為超音波感測模組，發出一脈波訊號使模組執行量測後傳回經過時間長的脈波
# 實際使用上雜訊相對多(可能受飛行時噪音影響)，故施加多重濾波條件：
# 超時防止(防止過長的無效等待)、角度補正、突波保護、均值濾波、卡爾曼濾波
# 經處理後的值堪用度大幅提高，對吸音面也能有效處理，使高度控制器能穩定運作
# 於過多線程執行中因線程排序問題需做應對處置
def altitude_sensor_sonar():
    global ACTUAL_ALT_SONAR, STATE_SENSOR_SONAR, SONAR_ERR_COUNT
    # 初始化腳位
    GPIO.setmode(GPIO.BCM)                  # 設定為BCM模式
    GPIO.setwarnings(False)                 # 忽略警告
    GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)    # TRIG腳位輸出設定
    GPIO.setup(SONAR_ECHO_PIN, GPIO.IN)     # ECHO腳位輸入設定
    GPIO.output(SONAR_TRIG_PIN, False)      # 預設輸出為0

    # 脈波變化時間值變數
    pulse_start = time.time()               # 脈波起始時間
    pulse_end = time.time()                 # 脈波結束時間

    # 卡爾曼濾波器初始化
    x_o = 0                                 # 前一次狀態
    p_o = 0                                 # 前一次斜方差
    z_o = 0                                 # 前一次量測值
    q = math.exp(-SONAR_KF_Q)               # 斜方差噪音值
    r = 2.92 * math.exp(-3)                 # 斜方差量測噪音值(定值不更改)

    # 平均降噪設定
    height_register = []
    height_ave = 0

    # 執行超音波高度計校正程序
    sonar_calibrate()

    # 回報超音波高度計啟動
    STATE_SENSOR_SONAR = True
    print('<啟動程序> 超音波高度計：啟動\n')

    # 程式迴圈
    while STATE_THREAD:
        # 線程優先執行起點
        THREAD_OCCUPY.acquire()

        # 重設超時旗標
        timeout = True

        # 卡爾曼濾波-預測
        x_p = x_o
        p_p = p_o + q

        # 腳位初始化(因有報錯記錄，於每次前重新初始化)
        GPIO.setmode(GPIO.BCM)                  # 設定為BCM模式
        GPIO.setwarnings(False)                 # 忽略警告
        GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)    # TRIG腳位輸出設定

        # 輸出啟用脈波
        GPIO.output(SONAR_TRIG_PIN, True)
        time.sleep(0.00001)                     # 發出10us脈波
        GPIO.output(SONAR_TRIG_PIN, False)

        # 記錄超時基準時間
        timeout_start = time.time()

        # 接收計時脈波
        # 於低電位迴圈等待，跳出時記錄脈波起始時間
        while not GPIO.input(SONAR_ECHO_PIN):
            pulse_start = time.time()           # 記錄脈波起始時間
            timeout = False                     # 成功記錄脈波時間，切換旗標
            if pulse_start - timeout_start > SONAR_TIMEOUT:
                timeout = True                  # 等待接收時發生超時，跳出迴圈
                break

        # 於高電位迴圈等待，跳出時記錄脈波結束時間(若發生超時則不執行本迴圈)
        while GPIO.input(SONAR_ECHO_PIN) and not timeout:
            pulse_end = time.time()             # 記錄脈波結束時間
            if pulse_end - pulse_start > SONAR_TIMEOUT:
                timeout = True                  # 接收時長發生超時，跳出迴圈
                break

        # 若發生超時則顯示於訊息(可停用)
        if timeout:
            SONAR_ERR_COUNT += 1        # 超時次數計(統計用)
            # print('<錯誤> 超音波高度計：量測超時')

        if not timeout:
            # 計算高度值
            pulse_duration = pulse_end - pulse_start
            height_raw = (pulse_duration * (331.3 + 0.606 * SONAR_AIR_TEMP)) / 2

            # 高度角度補正
            roll_angle = uav.attitude.roll
            pitch_angle = uav.attitude.pitch
            height_offset = (height_raw * math.cos(roll_angle) * math.cos(pitch_angle)) + SONAR_OFFSET
            height_offset_r = height_offset  # 記錄用值

            # 上下降突波防止
            if ACTUAL_ALT > SONAR_FILTER_ALT:
                if math.fabs(height_offset - z_o) > SONAR_MAX_CHANGE:
                    height_offset = height_ave

            # 均值濾波器(使線條滑順)
            height_register.insert(0, height_offset)
            if len(height_register) > SONAR_SAMPLE_NUM_AVE:
                height_register.pop()
            height_ave = np.average(height_register)  # 依據取樣次數計算即時平均值(會產生響應延遲)

            # 卡爾曼濾波器輸入限制(防止突波)
            if height_ave <= SONAR_KF_INPUT_MAX:
                z = height_ave  # 若低於限制則使用新值
            else:
                z = z_o  # 若超出限制則使用舊值

            # 卡爾曼濾波-更新
            k = p_p / (p_p + r)  # 卡爾曼增益
            x = x_p + k * (z - x_p)  # 狀態(濾波輸出值)
            p = (1 - k) * p_p  # 斜方差

            # 更新高度值至全域變數
            ACTUAL_ALT_SONAR = x

            # 傳遞值給下次迴圈使用
            x_o = x
            p_o = p
            z_o = z

            # 記錄數值
            R_ACTUAL_ALT_SONAR_RAW.append(height_raw)  # 記錄原始值
            R_ACTUAL_ALT_SONAR_OFFSET.append(height_offset_r)  # 記錄補正值(突波防止前)
            R_ACTUAL_ALT_SONAR_AVE.append(height_ave)  # 記錄平均值
            R_ACTUAL_ALT_SONAR_KF.append(ACTUAL_ALT_SONAR)  # 記錄KF值

        # 結束優先執行
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        time.sleep(1 / FREQ_SENSOR_SONAR)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_SONAR = False
    print('<關閉程序> 超音波高度計：關閉；量測超時次數共 %d 次\n' % SONAR_ERR_COUNT)


# =========== 飛行速度感測器 ============================================
# 說明：
# 計算循跡時的飛行速度，利用IMU的速度值計算合速度
def speed_sensor():
    global ACTUAL_SPD, STATE_SENSOR_SPD
    # 均值濾波器設定
    speed_register = []

    # 等待無人機連線
    if not STATE_UAV:
        print('<啟動程序> 飛行速度感測器：等待無人機連線')
    while not STATE_UAV:
        time.sleep(0.1)

    # 回報飛行速度控制器啟用
    STATE_SENSOR_SPD = True
    print('<啟動程序> 飛行速度感測器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 讀取速度值
        vx = uav.velocity[0]
        vy = uav.velocity[1]

        # 計算和速度
        if MODE == 'POS':
            speed_raw = 0
        else:
            speed_raw = math.sqrt(math.pow(vx, 2) + math.pow(vy, 2))

        # 均值濾波處理
        speed_register.insert(0, speed_raw)
        if len(speed_register) > SPD_SAMPLE_NUM_AVE:
            speed_register.pop()
        speed = np.average(speed_register)

        # 方向判斷
        hdg = uav.heading
        if vx > 0 and (hdg < 90 or hdg > 270):
            speed = speed
        elif vx > 0 and (90 < hdg < 270):
            speed = -speed
        elif vx < 0 and (hdg < 90 or hdg > 270):
            speed = -speed
        elif vx < 0 and (90 < hdg < 270):
            speed = speed
        elif vx == 0:
            if vy > 0 and hdg == 90:
                speed = speed
            elif vy > 0 and hdg == 270:
                speed = -speed
            elif vy < 0 and hdg == 270:
                speed = speed
            elif vy < 0 and hdg == 90:
                speed = -speed

        # 輸出至全域變數
        ACTUAL_SPD = speed

        # 迴圈執行間隔
        time.sleep(1 / FREQ_SENSOR_SPD)

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_SPD = False
    print('<關閉程序> 飛行速度感測器：關閉')


# ========== 命令傳輸程式 ==============================================
# 根據Ardupilot與MavLink協定規範，僅能使用set_attitude_target命令
# 另DroneKit套件提供函式封裝上也是依據上述規範編碼，故所使用的命令包括：
# 姿態四元數、各軸角速度、推力(昇降率)、操作模式(角度、角速度)
# 姿態四元數將控制器的角度值傳至函式做計算後返還、角速度採徑度制
# 根據測試及參考習慣操作，故選用姿態控制但於Yaw軸採角速度控制，符合手控習慣
# 流程上為將各控制器所輸出的命令值做統一編碼封裝後傳送
def command_transfer():
    global STATE_CMD

    # 回報命令傳輸程式啟動
    STATE_CMD = True
    print('<啟動程序> 命令傳輸程式：啟動\n')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 同捆執行起點
        THREAD_OCCUPY.acquire()

        # 命令編碼，使用偏航角速度控制(餘採姿態控制)
        msg = uav.message_factory.set_attitude_target_encode(
            0,                  # 開機時間
            0,                  # 目標系統
            0,                  # 目標裝置
            0b00000011,         # 命令遮罩(1忽略，0啟用)
            to_quaternion(TARGET_ROLL, TARGET_PITCH, TARGET_YAW),  # 姿態四元數
            0,                  # Roll滾轉角速度(radian/s)
            0,                  # Pitch滾轉角速度(radian/s)
            math.radians(TARGET_YAW_RATE),          # Yaw滾轉角速度(radian/s)
            TARGET_THRUST)      # 推力升降比(0~1，中點0.5)

        # 傳輸命令
        uav.send_mavlink(msg)

        # 同捆執行終點
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        time.sleep(1 / FREQ_CMD)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_CMD = False
    print('<關閉程序> 命令傳輸程式：關閉')


# ====================================================================
# ========== 子函式區塊 ================================================
# ====================================================================
# 本區放置各主函式下會使用到的功能函式，分有校正類、計算類、裝備控制類
# 將常使用的重複功能函式化以節省程式大小(僅限無傳遞值類的函式)


# ========== 超音波高度計校正程序 ========================================
# 於啟動前自動校正高度，為避免單次量測不準確，故以多次取樣後平均為基準來補正
def sonar_calibrate():
    global SONAR_OFFSET
    # 平均降噪設定
    height_register = []

    # 脈波變化時間值變數
    pulse_start = time.time()           # 脈波起始時間
    pulse_end = time.time()             # 脈波結束時間

    # 量測迴圈(線程優先執行)
    THREAD_OCCUPY.acquire()
    for i in range(0, SONAR_SAMPLE_NUM_CAL):
        # 設定超時旗標
        timeout = True

        # 量測迴圈(超時防止)
        while timeout:
            # 腳位初始化(因有報錯記錄，於每次前重新初始化)
            GPIO.setmode(GPIO.BCM)                  # 設定為BCM模式
            GPIO.setwarnings(False)                 # 忽略警告
            GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)    # TRIG腳位輸出設定

            # 輸出啟用脈波
            GPIO.output(SONAR_TRIG_PIN, True)
            time.sleep(0.00001)                     # 發出10us脈波
            GPIO.output(SONAR_TRIG_PIN, False)
            timeout_start = time.time()             # 記錄超時基準時間

            # 接收計時脈波
            # 於低電位迴圈等待，跳出時記錄脈波起始時間
            while not GPIO.input(SONAR_ECHO_PIN):
                pulse_start = time.time()           # 記錄脈波起始時間
                timeout = False                     # 成功記錄脈波時間，切換旗標
                if pulse_start - timeout_start > SONAR_TIMEOUT:
                    timeout = True                  # 等待接收時發生超時，跳出迴圈
                    break

            # 於高電位迴圈等待，跳出時記錄脈波結束時間(若發生超時則不執行本迴圈)
            while GPIO.input(SONAR_ECHO_PIN) and not timeout:
                pulse_end = time.time()             # 記錄脈波結束時間
                timeout = False                     # 成功記錄脈波時間，切換旗標
                if pulse_end - pulse_start > SONAR_TIMEOUT:
                    timeout = True                  # 接收時長發生超時，跳出迴圈
                    break

            # 若發生超時則顯示於訊息
            if timeout:
                # print('<錯誤> 超音波感測器：量測超時')
                pass

        # 計算高度值
        pulse_duration = pulse_end - pulse_start
        height = (pulse_duration * (331.3 + 0.606 * SONAR_AIR_TEMP)) / 2

        # 均值濾波器
        height_register.append(height)

        # 等待100ms執行下次量測
        time.sleep(0.1)

    # 結束優先執行
    THREAD_OCCUPY.release()

    # 更新高度校正偏移量
    SONAR_OFFSET = -np.average(height_register)
    print('<啟動程序> 超音波高度計校正完畢，偏移 %.3f m\n' % SONAR_OFFSET)


# ========== IMU高度計校正程序 =========================================
# 因IMU高度計會有波動，需做補正
def imu_calibrate():
    global IMU_OFFSET
    # 平均降噪設定
    height_register = []

    # 校正氣壓計
    uav.send_calibrate_barometer()
    uav.send_calibrate_vehicle_level()
    time.sleep(1)

    # 量測迴圈(線程優先執行)
    THREAD_OCCUPY.acquire()
    for i in range(0, IMU_SAMPLE_NUM_CAL):
        # 讀取氣壓計高度值
        height = uav.location.global_relative_frame.alt

        # 均值濾波器
        height_register.append(height)

        # 等待200ms執行下次量測
        time.sleep(0.2)

    # 結束優先執行
    THREAD_OCCUPY.release()

    # 更新高度校正偏移量
    IMU_OFFSET = -np.average(height_register)
    print('<啟動程序> 氣壓高度計校正完畢，偏移 %.3f m\n' % IMU_OFFSET)


# ========== 雷射測距儀校正程序 =========================================
def laser_calibrate():
    global LASER_OFFSET
    # 建立雷射量測物件
    tof = VL53L1X.VL53L1X(i2c_bus=1, i2c_address=0x29)
    tof.open()

    # 設定量測模式-近距模式
    tof.start_ranging(1)

    # 平均降噪設定
    height_register = []

    # 量測迴圈(線程優先執行)
    THREAD_OCCUPY.acquire()
    for i in range(0, LASER_SAMPLE_NUM_CAL):
        # 量測迴圈
        height_raw = tof.get_distance() / 1000

        # 均值濾波器
        height_register.append(height_raw)

        # 等待50ms執行下次量測
        time.sleep(0.05)

    # 關閉量測物件(防止衝突)
    tof.stop_ranging()

    # 結束優先執行
    THREAD_OCCUPY.release()

    # 更新高度校正偏移量
    LASER_OFFSET = -np.average(height_register)
    print('<啟動程序> 雷射測距儀校正完畢，偏移 %.3f m\n' % LASER_OFFSET)


# ========== 色塊型心計算程序 ==========================================
# 為計算當禎影像中，目標顏色的型心位置、色塊面積大小，主要供位置控制器、飛行條件判斷使用
# 為防止雜訊可設置最低面積值，因此若色塊大小在此之下將無法被偵測到
# 處理流程為將影像先依目標顏色做過濾，再依此做型心、面積運算
def image_moment(color):
    global CAM_IMAGE_LIVE
    # 色域轉換BGR->HSV
    image_hsv = cv2.cvtColor(CAM_IMAGE, cv2.COLOR_BGR2HSV)
    # 閉運算
    image_close = cv2.morphologyEx(image_hsv, cv2.MORPH_CLOSE,
                                   np.ones(IMG_MORPHOLOGY_EX_KERNEL, np.uint8))

    # 依照目標色抓該顏色輪廓
    if color == 'red':          # 紅色
        image_mask = cv2.inRange(image_close, IMG_RED_L, IMG_RED_H)
    elif color == 'green':      # 綠色
        image_mask = cv2.inRange(image_close, IMG_GREEN_L, IMG_GREEN_H)
    elif color == 'blue':       # 藍色
        image_mask = cv2.inRange(image_close, IMG_BLUE_L, IMG_BLUE_H)
    else:  # 黑色(預設顏色)
        image_mask = cv2.inRange(image_close, IMG_BLACK_L, IMG_BLACK_H)

    # 取得影像、輪廓點與索引(OpenCV Ver.3)
    # image_c, contours, hierarchy = cv2.findContours(image_mask, cv2.RETR_TREE,
    #                                                cv2.CHAIN_APPROX_SIMPLE)
    # 取得輪廓點與索引(OpenCV Ver.2&4)
    contours, hierarchy = cv2.findContours(image_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # 設定型心暫存列表
    moment_x_register = []
    moment_y_register = []
    area_register = []

    # 計算型心與面積
    for c in contours:
        # 計算色塊面積
        block_area = cv2.contourArea(c)

        # 面積符合使用門檻方取用(降噪處理)
        if block_area >= IMG_AREA_LIMIT:
            # 取得型心位置(左上為原點，加法增量)
            moment = cv2.moments(c)
            if moment["m00"] != 0:
                moment_x_raw = int(moment["m10"] / moment["m00"])
                moment_y_raw = int(moment["m01"] / moment["m00"])
            else:
                moment_x_raw = 0
                moment_y_raw = 0

            # 計算色塊位置(相對於畫面中心)
            block_x_offset = moment_x_raw - BODY_REF_X - BODY_OFFSET_X
            block_y_offset = BODY_REF_Y - moment_y_raw - BODY_OFFSET_Y

            moment_x_register.append(block_x_offset)
            moment_y_register.append(block_y_offset)
            area_register.append(block_area)

    # 設定型心運算暫存列表
    moment_area_x = 0
    moment_area_y = 0
    area_sum = 0

    # 計算面積一次矩
    for i in range(len(moment_x_register)):
        moment_area_x += moment_x_register[i] * area_register[i]
        moment_area_y += moment_y_register[i] * area_register[i]
        area_sum += area_register[i]

    # 求得組合型心位置
    if len(moment_x_register) != 0:
        block_x = moment_area_x / area_sum
        block_y = moment_area_y / area_sum
    else:
        block_x = 0
        block_y = 0

    # 輸出影像(GUI顯示用)
    if MODE == 'POS':
        CAM_IMAGE_LIVE = image_mask

    return block_x, block_y, area_sum


# ========== 線條型心計算 =============================================
# 將影像做二值化後再做型心與面積計算，除影響去用過程與畫面中心設定點不同外，其餘與前函式相同
# 但因中間有程式碼不同，為避免增加樹狀複雜度，故不再做副程式呼叫
# 本函式供航向控制器使用，僅回傳型心位置值，面積閥值亦獨立設定
def line_moment():
    global CAM_IMAGE_LIVE
    # 色域轉換BGR->GRAY
    img_gray = cv2.cvtColor(CAM_IMAGE, cv2.COLOR_BGR2GRAY)
    # 高斯模糊
    img_blur = cv2.GaussianBlur(img_gray, ANG_GB_KERNEL_SIZE, ANG_GB_SIGMA)
    # 二值化
    ret, img_th = cv2.threshold(img_blur, ANG_THRESHOLD,
                                ANG_THRESHOLD_NEW, cv2.THRESH_BINARY_INV)

    # 取得影像、輪廓點與索引(OpenCV Ver.3)
    # image_c, contours, hierarchy = cv2.findContours(img_th, cv2.RETR_TREE,
    #                                                cv2.CHAIN_APPROX_SIMPLE)

    # 取得輪廓點與索引(OpenCV Ver.2&4)
    contours, hierarchy = cv2.findContours(img_th, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # 設定型心暫存列表
    moment_x_register = []
    moment_y_register = []
    area_register = []
    angle_line = 0

    # 計算型心與面積
    for c in contours:
        # 計算色塊面積
        block_area = cv2.contourArea(c)

        # 面積符合使用門檻方取用(降噪處理)
        if block_area >= ANG_LINE_TRACK_AREA_L:
            # 取得型心位置(左上為原點，加法增量)
            moment = cv2.moments(c)
            if moment["m00"] != 0:
                moment_x_raw = int(moment["m10"] / moment["m00"])
                moment_y_raw = int(moment["m01"] / moment["m00"])
            else:
                moment_x_raw = 0
                moment_y_raw = 0

            # 矩形框選
            (x, y, w, h) = cv2.boundingRect(c)

            # 計算角度
            if h == 0 and img_th[y, x] == 255:
                angle_line = -90

            elif h == 0 and img_th[y, x] == 255:
                angle_line = 90

            else:
                angle_line = math.degrees(math.atan(w / h + 0.01))
                if img_th[y, x] == 255:
                    angle_line *= -1

            # 計算色塊位置(相對於畫面下方中心)
            block_x_offset = moment_x_raw - BODY_REF_X - BODY_OFFSET_X
            block_y_offset = 2 * BODY_REF_Y - moment_y_raw - BODY_OFFSET_Y

            moment_x_register.append(block_x_offset)
            moment_y_register.append(block_y_offset)
            area_register.append(block_area)

    # 設定型心運算暫存列表
    moment_area_x = 0
    moment_area_y = 0
    area_sum = 0

    # 計算面積一次矩
    for i in range(len(moment_x_register)):
        moment_area_x += moment_x_register[i] * area_register[i]
        moment_area_y += moment_y_register[i] * area_register[i]
        area_sum += area_register[i]

    # 求得組合型心位置
    if len(moment_x_register) != 0:
        block_x = moment_area_x / area_sum
        block_y = moment_area_y / area_sum
    else:
        block_x = 0
        block_y = 0

    # 輸出影像(GUI顯示用)
    if MODE == 'LINE':
        CAM_IMAGE_LIVE = img_th

    return block_x, block_y, angle_line


# ========== 角度計算 =================================================
# 本函式供計算型心點與原點連線和畫面Y軸夾角計算，於航向控制器使用
# 可解決平行線夾角為零的問題，所得的計算值相當於誤差值為本式特點，整合目標與回授
# 輸出值右側為正，左側為負(使用global Frame參照)
def angle_calculate(block_x, block_y):
    # 計算與中心線夾角
    if block_x == 0:
        angle = 0
    elif block_y == 0 and block_x > 0:
        angle = 90
    elif block_y == 0 and block_x < 0:
        angle = -90
    else:
        angle = math.degrees(math.atan(block_x/block_y))

    return -angle


# ========== 轉換尤拉角到四元數 =========================================
# 因命令格式針對姿態需使用四元數格式，因此需先把尤拉角作轉換，此格式可避免萬向節鎖產生
# 輸入格式為度(degree)，輸出為四元數(q,x,y,z)
def to_quaternion(roll=0.0, pitch=0.0, yaw=0.0):
    t0 = math.cos(math.radians(yaw * 0.5))
    t1 = math.sin(math.radians(yaw * 0.5))
    t2 = math.cos(math.radians(roll * 0.5))
    t3 = math.sin(math.radians(roll * 0.5))
    t4 = math.cos(math.radians(pitch * 0.5))
    t5 = math.sin(math.radians(pitch * 0.5))

    w = t0 * t2 * t4 + t1 * t3 * t5
    x = t0 * t3 * t4 - t1 * t2 * t5
    y = t0 * t2 * t5 + t1 * t3 * t4
    z = t1 * t2 * t4 - t0 * t3 * t5

    return [w, x, y, z]


# ========= 沙包盒初始化 ===============================================
# 於起飛時送出脈波訊號固定艙門，防止貨物掉出，並維持較佳的空氣動力學外型
def box_initialize():
    global STATE_BOX, BOX
    # 初始化腳位
    GPIO.setmode(GPIO.BCM)                      # 設定為BCM模式
    GPIO.setwarnings(False)                     # 忽略警告
    GPIO.setup(BOX_PWM_PIN, GPIO.OUT)           # PWM腳位輸出設定
    BOX = GPIO.PWM(BOX_PWM_PIN, BOX_PWM_FREQ)   # 建立PWM控制物件

    # 啟動PWM控制，使門維持關閉
    dc = BOX_PWM_FREQ * (BOX_PWM_CLOSE / 1000000) * 100
    BOX.start(dc)

    # 回報艙門狀態
    STATE_BOX = True
    print('<飛行狀態> 沙包盒艙門：關閉\n')


# ========= 沙包盒關門 =================================================
# 於起飛時送出脈波訊號固定艙門，防止貨物掉出，並維持較佳的空氣動力學外型
def box_close():
    global STATE_BOX
    # 啟動PWM控制，使門維持關閉
    dc = BOX_PWM_FREQ * (BOX_PWM_CLOSE / 1000000) * 100
    BOX.ChangeDutyCycle(dc)

    # 回報艙門狀態
    STATE_BOX = True
    print('<飛行狀態> 沙包盒艙門：關閉\n')


# ======== 沙包盒開門 =================================================
# 輸出訊號使艙門開啟，使貨物卸出
def box_open():
    global STATE_BOX
    # 啟動PWM控制，使門打開
    dc = BOX_PWM_FREQ * (BOX_PWM_OPEN / 1000000) * 100
    BOX.ChangeDutyCycle(dc)

    # 回報艙門狀態
    STATE_BOX = False
    print('<飛行狀態> 沙包盒艙門：開啟\n')


# ========== 多線程啟動函式 ============================================
# 供主程式啟動多線程用，若連線中斷時，自動重新連線
def thread_activate():
    global STATE_THREAD, STATE_UAV, uav
    # 若未連線，重新執行連線
    if not STATE_UAV:
        print('<啟動程序> 與UAV連線中，請稍候...')
        # 樹莓派連線指令
        uav = connect(CONNECT_PORT, baud=CONNECT_BAUD)
        # 電腦連線指令
        # uav = connect(CONNECT_PORT, baud=CONNECT_BAUD)
        STATE_UAV = True
        print('<啟動程序> UAV已連線')

    # 定義多線程
    cam = threading.Thread(target=camera)
    cam_sen = threading.Thread(target=camera_sensor)
    sonar_sen = threading.Thread(target=altitude_sensor_sonar)
    alt_sen = threading.Thread(target=altitude_sensor)
    spd_sen = threading.Thread(target=speed_sensor)
    alt_ctrl = threading.Thread(target=altitude_controller)
    pos_ctrl = threading.Thread(target=position_controller)
    pos_vel_ctrl = threading.Thread(target=position_velocity_controller)
    hdg_ctrl = threading.Thread(target=heading_controller)
    spd_ctrl = threading.Thread(target=flight_speed_controller)
    cmd = threading.Thread(target=command_transfer)

    # 啟動多線程旗標
    STATE_THREAD = True

    # 執行多線程(依啟動順序放置)
    cam.start()
    cam_sen.start()
    sonar_sen.start()
    alt_sen.start()
    spd_sen.start()
    alt_ctrl.start()
    pos_ctrl.start()
    pos_vel_ctrl.start()
    hdg_ctrl.start()
    spd_ctrl.start()
    cmd.start()

    # 回報多線程啟動
    time.sleep(5)
    print('<啟動程序> 多線程程序：啟動')

# ====================================================================
# ========== 數值紀錄區塊 ==============================================
# ====================================================================


# ========== Excel 輸出 ===============================================
def data_sheet():
    # 建立活頁簿
    db = Workbook()

    # 建立分頁
    ds1 = db.create_sheet("ALT Controller")
    ds2 = db.create_sheet("POS Controller")
    ds2b = db.create_sheet("POS-VEL Controller")
    ds3 = db.create_sheet("HDG Controller")
    ds4 = db.create_sheet("ALT Sensor")

    # 高度控制器(先列在行)
    ds1.cell(1, 1, '高度控制器')
    ds1.cell(2, 1, '實際高度')
    for x in range(len(R_ACTUAL_ALT_CTRL)):
        ds1.cell(x + 3, 1, R_ACTUAL_ALT_CTRL[x])

    ds1.cell(2, 2, '目標高度')
    for x in range(len(R_TARGET_ALT)):
        ds1.cell(x + 3, 2, R_TARGET_ALT[x])

    ds1.cell(2, 3, '爬升率')
    for x in range(len(R_TARGET_THRUST)):
        ds1.cell(x + 3, 3, R_TARGET_THRUST[x])

    # 高度控制器PID響應
    ds1.cell(1, 5, '高度控制器PID響應')
    ds1.cell(2, 5, 'Error')
    for x in range(len(R_ALT_ERROR)):
        ds1.cell(x + 3, 5, R_ALT_ERROR[x])

    ds1.cell(2, 6, 'Gain P')
    for x in range(len(R_ALT_GP)):
        ds1.cell(x + 3, 6, R_ALT_GP[x])

    ds1.cell(2, 7, 'Gain I')
    for x in range(len(R_ALT_GI)):
        ds1.cell(x + 3, 7, R_ALT_GI[x])

    ds1.cell(2, 8, 'Gain D')
    for x in range(len(R_ALT_GD)):
        ds1.cell(x + 3, 8, R_ALT_GD[x])

    # 位置控制器-第二頁
    # 位置控制器X響應圖
    ds2.cell(1, 1, '位置控制器X響應圖')
    ds2.cell(2, 1, 'X位置')
    for x in range(len(R_POS_BLOCK_X)):
        ds2.cell(x + 3, 1, R_POS_BLOCK_X[x])

    ds2.cell(2, 2, 'X速度輸出')
    for x in range(len(R_POS_X_TARGET_VEL)):
        ds2.cell(x + 3, 2, R_POS_X_TARGET_VEL[x])

    ds2.cell(2, 3, '面積大小')
    for x in range(len(R_POS_BLOCK_AREA)):
        ds2.cell(x + 3, 3, R_POS_BLOCK_AREA[x])

    # 位置控制器Y響應圖
    ds2.cell(1, 5, '位置控制器Y響應圖')
    ds2.cell(2, 5, 'Y位置')
    for x in range(len(R_POS_BLOCK_Y)):
        ds2.cell(x + 3, 5, R_POS_BLOCK_Y[x])

    ds2.cell(2, 6, 'Y速度輸出')
    for x in range(len(R_POS_Y_TARGET_VEL)):
        ds2.cell(x + 3, 6, R_POS_Y_TARGET_VEL[x])

    ds2.cell(2, 7, '面積大小')
    for x in range(len(R_POS_BLOCK_AREA)):
        ds2.cell(x + 3, 7, R_POS_BLOCK_AREA[x])

    # 軌跡記錄
    ds2.cell(1, 9, '位置控制器Y響應圖')
    ds2.cell(2, 9, 'X位置')
    for x in range(len(R_POS_BLOCK_X)):
        ds2.cell(x + 3, 9, R_POS_BLOCK_X[x])

    ds2.cell(2, 10, 'Y位置')
    for x in range(len(R_POS_BLOCK_Y)):
        ds2.cell(x + 3, 10, R_POS_BLOCK_Y[x])

    # 位置控制器X-PID響應圖
    ds2.cell(1, 12, '位置控制器X-PID響應圖')
    ds2.cell(2, 12, 'Error')
    for x in range(len(R_POS_X_ERROR)):
        ds2.cell(x + 3, 12, R_POS_X_ERROR[x])

    ds2.cell(2, 13, 'Gain P')
    for x in range(len(R_POS_X_GP)):
        ds2.cell(x + 3, 13, R_POS_X_GP[x])

    ds2.cell(2, 14, 'Gain I')
    for x in range(len(R_POS_X_GI)):
        ds2.cell(x + 3, 14, R_POS_X_GI[x])

    ds2.cell(2, 15, 'Gain D')
    for x in range(len(R_POS_X_GD)):
        ds2.cell(x + 3, 15, R_POS_X_GD[x])

    # 位置控制器Y-PID響應圖
    ds2.cell(1, 17, '位置控制器Y-PID響應圖')
    ds2.cell(2, 17, 'Error')
    for x in range(len(R_POS_Y_ERROR)):
        ds2.cell(x + 3, 17, R_POS_Y_ERROR[x])

    ds2.cell(2, 18, 'Gain P')
    for x in range(len(R_POS_Y_GP)):
        ds2.cell(x + 3, 18, R_POS_Y_GP[x])

    ds2.cell(2, 19, 'Gain I')
    for x in range(len(R_POS_Y_GI)):
        ds2.cell(x + 3, 19, R_POS_Y_GI[x])

    ds2.cell(2, 20, 'Gain D')
    for x in range(len(R_POS_Y_GD)):
        ds2.cell(x + 3, 20, R_POS_Y_GD[x])

    # 速度控制器-第三頁
    # 速度控制器X響應圖
    ds2b.cell(1, 1, '速度控制器X響應圖')
    ds2b.cell(2, 1, 'X目標速度')
    for x in range(len(R_POS_VEL_X_TAR)):
        ds2b.cell(x + 3, 1, R_POS_VEL_X_TAR[x])

    ds2b.cell(2, 2, 'X實際速度')
    for x in range(len(R_POS_VEL_X_ACT)):
        ds2b.cell(x + 3, 2, R_POS_VEL_X_ACT[x])

    ds2b.cell(2, 3, '目標Roll')
    for x in range(len(R_POS_TARGET_ROLL)):
        ds2b.cell(x + 3, 3, R_POS_TARGET_ROLL[x])

    ds2b.cell(2, 4, '實際Roll')
    for x in range(len(R_POS_ACTUAL_ROLL)):
        ds2b.cell(x + 3, 4, R_POS_ACTUAL_ROLL[x])

    # 速度控制器Y響應圖
    ds2b.cell(1, 6, '速度控制器Y響應圖')
    ds2b.cell(2, 6, 'Y目標速度')
    for x in range(len(R_POS_VEL_Y_TAR)):
        ds2b.cell(x + 3, 6, R_POS_VEL_Y_TAR[x])

    ds2b.cell(2, 7, 'Y實際速度')
    for x in range(len(R_POS_VEL_Y_ACT)):
        ds2b.cell(x + 3, 7, R_POS_VEL_Y_ACT[x])

    ds2b.cell(2, 8, '目標Pitch')
    for x in range(len(R_POS_TARGET_PITCH)):
        ds2b.cell(x + 3, 8, R_POS_TARGET_PITCH[x])

    ds2b.cell(2, 9, '實際Pitch')
    for x in range(len(R_POS_ACTUAL_PITCH)):
        ds2b.cell(x + 3, 9, R_POS_ACTUAL_PITCH[x])

    # 速度控制器X-PID響應圖
    ds2b.cell(1, 11, '速度控制器X-PID響應圖')
    ds2b.cell(2, 11, 'Error')
    for x in range(len(R_POS_VEL_X_ERROR)):
        ds2b.cell(x + 3, 11, R_POS_VEL_X_ERROR[x])

    ds2b.cell(2, 12, 'Gain P')
    for x in range(len(R_POS_VEL_X_GP)):
        ds2b.cell(x + 3, 12, R_POS_VEL_X_GP[x])

    ds2b.cell(2, 13, 'Gain I')
    for x in range(len(R_POS_VEL_X_GI)):
        ds2b.cell(x + 3, 13, R_POS_VEL_X_GI[x])

    ds2b.cell(2, 14, 'Gain D')
    for x in range(len(R_POS_VEL_X_GD)):
        ds2b.cell(x + 3, 14, R_POS_VEL_X_GD[x])

    # 速度控制器Y-PID響應圖
    ds2b.cell(1, 16, '速度控制器Y-PID響應圖')
    ds2b.cell(2, 16, 'Error')
    for x in range(len(R_POS_VEL_Y_ERROR)):
        ds2b.cell(x + 3, 16, R_POS_VEL_Y_ERROR[x])

    ds2b.cell(2, 17, 'Gain P')
    for x in range(len(R_POS_VEL_Y_GP)):
        ds2b.cell(x + 3, 17, R_POS_VEL_Y_GP[x])

    ds2b.cell(2, 18, 'Gain I')
    for x in range(len(R_POS_VEL_Y_GI)):
        ds2b.cell(x + 3, 18, R_POS_VEL_Y_GI[x])

    ds2b.cell(2, 19, 'Gain D')
    for x in range(len(R_POS_VEL_Y_GD)):
        ds2b.cell(x + 3, 19, R_POS_VEL_Y_GD[x])

    # 第四頁-循線控制器
    # 航向控制器數值處理圖
    ds3.cell(1, 1, '航向控制器數值處理圖')
    ds3.cell(2, 1, '實際角度差')
    for x in range(len(R_HDG_ACTUAL)):
        ds3.cell(x + 3, 1, R_HDG_ACTUAL[x])

    ds3.cell(2, 2, '輸出滾轉角')
    for x in range(len(R_HDG_TARGET_ROLL)):
        ds3.cell(x + 3, 2, R_HDG_TARGET_ROLL[x])

    ds3.cell(2, 3, '輸出偏航角速率')
    for x in range(len(R_HDG_TARGET_YAW_RATE)):
        ds3.cell(x + 3, 3, R_HDG_TARGET_YAW_RATE[x])

    # 航向控制器PID數值處理圖
    ds3.cell(1, 5, '航向控制器PID數值處理圖')
    ds3.cell(2, 5, 'Error')
    for x in range(len(R_HDG_ERROR)):
        ds3.cell(x + 3, 5, R_HDG_ERROR[x])

    ds3.cell(2, 6, 'Gain P')
    for x in range(len(R_HDG_GP)):
        ds3.cell(x + 3, 6, R_HDG_GP[x])

    ds3.cell(2, 7, 'Gain I')
    for x in range(len(R_HDG_GI)):
        ds3.cell(x + 3, 7, R_HDG_GI[x])

    ds3.cell(2, 8, 'Gain D')
    for x in range(len(R_HDG_GP)):
        ds3.cell(x + 3, 8, R_HDG_GP[x])

    # 飛行速度控制器
    ds3.cell(1, 10, '飛行速度控制器數值處理圖')
    ds3.cell(2, 10, '目標速度')
    for x in range(len(R_SPD_TARGET)):
        ds3.cell(x + 3, 10, R_SPD_TARGET[x])

    ds3.cell(2, 11, '實際速度')
    for x in range(len(R_SPD_ACTUAL)):
        ds3.cell(x + 3, 11, R_SPD_ACTUAL[x])

    ds3.cell(2, 12, '目標傾角')
    for x in range(len(R_SPD_TARGET_PITCH)):
        ds3.cell(x + 3, 12, R_SPD_TARGET_PITCH[x])

    ds3.cell(2, 13, '實際傾角')
    for x in range(len(R_SPD_ACTUAL_PITCH)):
        ds3.cell(x + 3, 13, R_SPD_ACTUAL_PITCH[x])

    # 飛行速度控制器PID數值處理圖
    ds3.cell(1, 15, '飛行速度控制器PID數值處理圖')
    ds3.cell(2, 15, 'Error')
    for x in range(len(R_SPD_ERROR)):
        ds3.cell(x + 3, 15, R_SPD_ERROR[x])

    ds3.cell(2, 16, 'Gain P')
    for x in range(len(R_SPD_GP)):
        ds3.cell(x + 3, 16, R_SPD_GP[x])

    ds3.cell(2, 17, 'Gain I')
    for x in range(len(R_SPD_GI)):
        ds3.cell(x + 3, 17, R_SPD_GI[x])

    ds3.cell(2, 18, 'Gain D')
    for x in range(len(R_SPD_GP)):
        ds3.cell(x + 3, 18, R_SPD_GP[x])

    # 第五頁-感測器
    # 高度感測器數值處理圖
    ds4.cell(1, 1, '高度感測器數值處理圖')
    ds4.cell(2, 1, '實際高')
    for x in range(len(R_ACTUAL_ALT)):
        ds4.cell(x + 3, 1, R_ACTUAL_ALT[x])

    ds4.cell(2, 2, '超音波')
    for x in range(len(R_ACTUAL_ALT_SONAR)):
        ds4.cell(x + 3, 2, R_ACTUAL_ALT_SONAR[x])

    ds4.cell(2, 3, 'IMU')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds4.cell(x + 3, 3, R_ACTUAL_ALT_IMU_KF[x])

    ds4.cell(2, 4, 'IMU')
    for x in range(len(R_ACTUAL_ALT_LASER)):
        ds4.cell(x + 3, 4, R_ACTUAL_ALT_LASER[x])

    # 超音波高度計數值處理圖
    ds4.cell(1, 6, '超音波高度計數值處理圖')
    ds4.cell(2, 6, '原始值')
    for x in range(len(R_ACTUAL_ALT_SONAR_RAW)):
        ds4.cell(x + 3, 6, R_ACTUAL_ALT_SONAR_RAW[x])

    ds4.cell(2, 7, '補正後')
    for x in range(len(R_ACTUAL_ALT_SONAR_OFFSET)):
        ds4.cell(x + 3, 7, R_ACTUAL_ALT_SONAR_OFFSET[x])

    ds4.cell(2, 8, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_AVE)):
        ds4.cell(x + 3, 8, R_ACTUAL_ALT_SONAR_AVE[x])

    ds4.cell(2, 9, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_KF)):
        ds4.cell(x + 3, 9, R_ACTUAL_ALT_SONAR_KF[x])

    # IMU高度計數值處理圖
    ds4.cell(1, 11, 'IMU高度計數值處理圖')
    ds4.cell(2, 11, '原始值')
    for x in range(len(R_ACTUAL_ALT_IMU_RAW)):
        ds4.cell(x + 3, 11, R_ACTUAL_ALT_IMU_RAW[x])

    ds4.cell(2, 12, '補正後')
    for x in range(len(R_ACTUAL_ALT_IMU_OFFSET)):
        ds4.cell(x + 3, 12, R_ACTUAL_ALT_IMU_OFFSET[x])

    ds4.cell(2, 13, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_AVE)):
        ds4.cell(x + 3, 13, R_ACTUAL_ALT_IMU_AVE[x])

    ds4.cell(2, 14, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds4.cell(x + 3, 14, R_ACTUAL_ALT_IMU_KF[x])

    # 依照時間存檔
    t_save = time.localtime()
    db.save('Flight Data %d_%d_%d_%d_%d_%d.xlsx' %
            (t_save[0], t_save[1], t_save[2], t_save[3], t_save[4], t_save[5]))


# ========== 高度控制器 ================================================
def alt_ctrl_data():
    # ======高度控制器================
    # 高度控制器響應圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_CTRL) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_CTRL, label="Actual Altitude")
    plt.plot(num_alt, R_TARGET_ALT, label="Target Altitude")
    plt.plot(num_alt, R_TARGET_THRUST, label="Climb Rate")

    plt.xlabel('Data number')
    plt.ylabel('Height(m), Velocity(m/s)')
    plt.title('Altitude Controller Data')
    plt.legend()
    plt.savefig('Altitude Controller Data.png')
    plt.show()

    # 高度控制器PID輸出圖
    num_alt = []
    for i in range(1, len(R_ALT_ERROR) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ALT_ERROR, label="Error")
    plt.plot(num_alt, R_ALT_GP, label="Gain P")
    plt.plot(num_alt, R_ALT_GI, label="Gain I")
    plt.plot(num_alt, R_ALT_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Height(m), Velocity(m/s)')
    plt.title('Altitude Controller PID Output Data')
    plt.legend()
    plt.savefig('Altitude Controller PID Output Data.png')
    plt.show()


# ========== 位置控制器 ================================================
def pos_ctrl_data():
    # 位置控制器X響應圖
    num_pos = []
    for i in range(1, len(R_POS_BLOCK_X) + 1):
        num_pos.append(i)

    plt.plot(num_pos, R_POS_BLOCK_X, label="X Position")
    plt.plot(num_pos, R_POS_X_TARGET_VEL, label="X Velocity Output")
    plt.plot(num_pos, R_POS_BLOCK_AREA, label="Block Area")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller Data - X axis')
    plt.legend()
    plt.savefig('Position Controller Data - X axis.png')
    plt.show()

    # 位置控制器Y響應圖
    plt.plot(num_pos, R_POS_BLOCK_Y, label="Y Position")
    plt.plot(num_pos, R_POS_Y_TARGET_VEL, label="Y Velocity Output")
    plt.plot(num_pos, R_POS_BLOCK_AREA, label="Block Area")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller Data - Y axis')
    plt.legend()
    plt.savefig('Position Controller Data - Y axis.png')
    plt.show()

    # 位置控制器X-PID響應圖
    plt.plot(num_pos, R_POS_X_ERROR, label="Error")
    plt.plot(num_pos, R_POS_X_GP, label="Gain P")
    plt.plot(num_pos, R_POS_X_GI, label="Gain I")
    plt.plot(num_pos, R_POS_X_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller PID Output Data - X axis')
    plt.legend()
    plt.savefig('Position Controller PID Output Data - X axis.png')
    plt.show()

    # 位置控制器Y-PID響應圖
    plt.plot(num_pos, R_POS_Y_ERROR, label="Error")
    plt.plot(num_pos, R_POS_Y_GP, label="Gain P")
    plt.plot(num_pos, R_POS_Y_GI, label="Gain I")
    plt.plot(num_pos, R_POS_Y_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller PID Output Data - Y axis')
    plt.legend()
    plt.savefig('Position Controller PID Output Data - Y axis.png')
    plt.show()

    # 位置控制器軌跡響應圖
    plt.plot(R_POS_BLOCK_X, R_POS_BLOCK_Y, label="Block Trajectory")

    plt.xlabel('X-Position(px)')
    plt.ylabel('Y-Position(px)')
    plt.title('Position Controller Trajectory Data')
    plt.legend()
    plt.savefig('Position Controller Trajectory Data.png')
    plt.show()


# ========== 速度控制器 ================================================
def vel_ctrl_data():
    # 位置控制器X響應圖
    num_vel = []
    for i in range(1, len(R_POS_VEL_X_TAR) + 1):
        num_vel.append(i)

    # 速度控制器X響應圖
    plt.plot(num_vel, R_POS_VEL_X_TAR, label="Target X Velocity")
    plt.plot(num_vel, R_POS_VEL_X_ACT, label="Actual X Velocity")
    plt.plot(num_vel, R_POS_TARGET_ROLL, label="Target Roll Angle")
    plt.plot(num_vel, R_POS_ACTUAL_ROLL, label="Actual Roll Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller Data - X axis')
    plt.legend()
    plt.savefig('Velocity Controller Data - X axis.png')
    plt.show()

    # 速度控制器Y響應圖
    plt.plot(num_vel, R_POS_VEL_Y_TAR, label="Target Y Velocity")
    plt.plot(num_vel, R_POS_VEL_Y_ACT, label="Actual Y Velocity")
    plt.plot(num_vel, R_POS_TARGET_PITCH, label="Target Pitch Angle")
    plt.plot(num_vel, R_POS_ACTUAL_PITCH, label="Actual Pitch Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller Data - Y axis')
    plt.legend()
    plt.savefig('Velocity Controller Data - Y axis.png')
    plt.show()

    # 速度控制器X-PID響應圖
    plt.plot(num_vel, R_POS_VEL_X_ERROR, label="Error")
    plt.plot(num_vel, R_POS_VEL_X_GP, label="Gain P")
    plt.plot(num_vel, R_POS_VEL_X_GI, label="Gain I")
    plt.plot(num_vel, R_POS_VEL_X_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller PID Output Data - X axis')
    plt.legend()
    plt.savefig('Velocity Controller PID Output Data - X axis.png')
    plt.show()

    # 速度控制器Y-PID響應圖
    plt.plot(num_vel, R_POS_VEL_Y_ERROR, label="Error")
    plt.plot(num_vel, R_POS_VEL_Y_GP, label="Gain P")
    plt.plot(num_vel, R_POS_VEL_Y_GI, label="Gain I")
    plt.plot(num_vel, R_POS_VEL_Y_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller PID Output Data - Y axis')
    plt.legend()
    plt.savefig('Velocity Controller PID Output Data - Y axis.png')
    plt.show()


# ========== 高度感測器 ================================================
def alsen_data():
    # 高度感測器數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT, label="Actual Height")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR, label="Sonar Data")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_KF, label="IMU Data")
    plt.plot(num_alt, R_ACTUAL_ALT_LASER, label="Laser Data")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Sensor Data')
    plt.legend()
    plt.savefig('Altitude Sensor Data.png')
    plt.show()

    # 超音波高度計數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_SONAR_KF) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_RAW, label="Raw Data")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_OFFSET, label="Angle & Height Offset")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_AVE, label="Average Filter")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_KF, label="Kalman Filter")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Data - Sonar')
    plt.legend()
    plt.savefig('Altitude Data - Sonar.png')
    plt.show()

    # IMU高度計數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_IMU_RAW) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_IMU_RAW, label="Raw Data")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_OFFSET, label="Height Offset")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_AVE, label="Average Filter")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_KF, label="Kalman Filter")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Data - IMU')
    plt.legend()
    plt.savefig('Altitude Data - IMU.png')
    plt.show()


# ========== 航向控制器 ================================================
def hdg_ctrl_data():
    # 航向控制器響應圖
    num_hdg = []
    for i in range(1, len(R_HDG_ACTUAL) + 1):
        num_hdg.append(i)

    # 航向控制器響應圖
    plt.plot(num_hdg, R_HDG_ACTUAL, label="Actual Heading")
    plt.plot(num_hdg, R_HDG_TARGET_ROLL, label="Target Roll")
    plt.plot(num_hdg, R_HDG_TARGET_YAW_RATE, label="Target Yaw Rate")

    plt.xlabel('Data number')
    plt.ylabel('Angle(degree), Angle Rate(degree/s)')
    plt.title('Heading Controller Data')
    plt.legend()
    plt.savefig('Heading Controller Data.png')
    plt.show()

    # 航向控制器PID響應圖
    plt.plot(num_hdg, R_HDG_ERROR, label="Error")
    plt.plot(num_hdg, R_HDG_GP, label="Gain P")
    plt.plot(num_hdg, R_HDG_GI, label="Gain I")
    plt.plot(num_hdg, R_HDG_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Angle(degree), Angle Rate(degree/s)')
    plt.title('Velocity Controller PID Output Data - Y axis')
    plt.legend()
    plt.savefig('Velocity Controller PID Output Data - Y axis.png')
    plt.show()


# ========== 飛行速度控制器 =============================================
def spd_ctrl_data():
    # 飛行速度控制器響應圖
    num_spd = []
    for i in range(1, len(R_SPD_TARGET) + 1):
        num_spd.append(i)

    # 速度控制器X響應圖
    plt.plot(num_spd, R_SPD_TARGET, label="Target Speed")
    plt.plot(num_spd, R_SPD_ACTUAL, label="Actual Speed")
    plt.plot(num_spd, R_SPD_TARGET_PITCH, label="Target Pitch Angle")
    plt.plot(num_spd, R_SPD_ACTUAL_PITCH, label="Actual Pitch Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s), Angle(degree)')
    plt.title('Flight Speed Controller Data')
    plt.legend()
    plt.savefig('Flight Speed Controller Data.png')
    plt.show()

    # 速度控制器X-PID響應圖
    plt.plot(num_spd, R_SPD_ERROR, label="Error")
    plt.plot(num_spd, R_SPD_GP, label="Gain P")
    plt.plot(num_spd, R_SPD_GI, label="Gain I")
    plt.plot(num_spd, R_SPD_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s)')
    plt.title('Flight Speed Controller PID Output Data')
    plt.legend()
    plt.savefig('Flight Speed Controller PID Output Data.png')
    plt.show()


# ====================================================================
# ========== 飛行程序指令(條件與命令) ====================================
# ====================================================================
# 任務中會使用的區段飛行命令、模式切換命令、觸發條件


# ========== 啟動前檢查 ================================================
# 本程式為飛行前的安全檢查項目，除確認控制器是否上線外，也提醒控制員檢查安全程序
# 項目有：安全裝置移除(槳片銷)、起飛區淨空，完成後即交付起飛控制
def pre_check():
    # 執行起飛前檢查
    print('<啟動程序> 啟動前檢查程序：啟動')

    # 階段一：控制器檢查
    # 等待高度控制器啟動
    if not STATE_ALT:
        print('<啟動程序> 主程序：等待高度控制器啟動')
    while not STATE_ALT:
        time.sleep(0.1)

    # 等待命令傳輸啟動
    if not STATE_CMD:
        print('<啟動程序> 主程序：等待命令傳輸啟動')
    while not STATE_CMD:
        time.sleep(0.1)

    # 初始化儲物倉
    box_initialize()

    # 階段三：安全檢查程序
    # 確認安全裝置解除
    safe_pin = 'N'
    while safe_pin != 'Y':
        safe_pin = input('<啟動程序> 請確認安全裝置已移除(Y/N) => ')
    print('<啟動程序> 安全裝置已移除')
    time.sleep(1)

    # 確認起飛區已淨空
    lunch_clear = 'N'
    while lunch_clear != 'Y':
        lunch_clear = input('<啟動程序> 請確認起飛區已淨空(Y/N) => ')
    print('<啟動程序> 起飛區已淨空')

    # 回報起飛前檢點完成
    print('<啟動程序> 啟動前檢查程序：完成')
    time.sleep(1)

    # 任務開始確認
    go_or_nogo = 'N'
    while go_or_nogo != 'Y':
        go_or_nogo = input('<啟動程序> 請求任務開始(Y/N) => ')

    # 回報任務執行開始
    print('<啟動程序> TDK飛行任務，開始!')
    time.sleep(1)


# ========== 起飛程序 =================================================
# 執行模式切換、無人機解鎖，並由高度控制器將無人機送至目標飛行高度
def take_off(target_alt=1.0, color=''):
    global TARGET_ALT, MODE
    # 切換至無衛星導航模式
    uav.mode = VehicleMode("GUIDED_NOGPS")
    while not uav.mode.name == 'GUIDED_NOGPS':
        time.sleep(0.5)
    print('<啟動程序> 無人機已切換至無衛星導航模式')

    # 解鎖無人機
    uav.arm(wait=True)
    print('<啟動程序> 無人機已解鎖')

    # 設定飛行模式
    if color == '':
        pass
    else:
        pos_mode(color)

    # 設定目標高度
    print('<飛行命令> 起飛至目標高度 %.2f m' % target_alt)
    TARGET_ALT = target_alt

    # 等待抵達目標高度
    i = 0
    while ACTUAL_ALT < TARGET_ALT * (TAKEOFF_SETTLE / 100):
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 起飛中，目前高度 %.2f m，距離目標高度尚有 %.2f m' %
                  (ACTUAL_ALT, TARGET_ALT - ACTUAL_ALT))
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
            print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px^2' %
                  (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
            print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' %
                  (ACTUAL_VEL[0], ACTUAL_VEL[1]))

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報抵達目標高度
    print('<飛行狀態> 已抵達目標高度 %.2f m' % TARGET_ALT)


# ========== 懸停程序 =================================================
def hover(duration=1):
    # 設定飛行命令
    duration = int(duration)
    print('<飛行命令> 懸停於 %.2f m，維持 %d sec' % (TARGET_ALT, duration))

    # 等待時間到達
    i = 0
    for x in range(duration*10):
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 懸停中，目前高度 %.2f m，剩餘時間 %d sec' % (ACTUAL_ALT, duration - x/10))
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
            print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px^2' %
                  (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
            print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' %
                  (ACTUAL_VEL[0], ACTUAL_VEL[1]))

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報懸停結束
    print('<飛行狀態> 懸停狀態結束，目前高度 %.2f m' % ACTUAL_ALT)


# ========== 降落程序 =================================================
def landing(color=''):
    global TARGET_ALT, MODE
    # 設定目標高度
    print('<飛行命令> 位置已確認，開始降落至地面')
    TARGET_ALT = -0.05  # 負值以確保不會著地後彈升

    # 設定定位顏色
    if color != '':
        pos_mode(color)

    # 等待抵達目標高度
    i = 0
    while ACTUAL_ALT > LANDING_CUTOFF:
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 降落中，目前高度 %.2f m' % ACTUAL_ALT)
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
            print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px^2' %
                  (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
            print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' %
                  (ACTUAL_VEL[0], ACTUAL_VEL[1]))

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報降落成功
    print('<飛行狀態> 已降落地面')

    # 返回手動模式
    uav.mode = VehicleMode("STABILIZE")
    while not uav.mode.name == 'STABILIZE':
        time.sleep(0.5)
    print('<關閉程序> 無人機已切換至手動模式')


# ========== 設定為定位模式 ============================================
# 切換至定位模式，並輸入目標顏色與追蹤位置來定位目標色塊
def pos_mode(color, target=(0, 0)):
    global MODE, POS_COLOR, TARGET_POS
    MODE = 'POS'                # 模式旗標切為定位
    POS_COLOR = color           # 定位目標顏色設置
    TARGET_POS = list(target)
    # 回報模式啟用狀態
    print('<飛行狀態> 切換至定位模式，顏色追蹤：%s\n' % POS_COLOR)


# ========== 設定為循跡模式 ===========================================
# 切換至循跡模式，輸入飛行速度做控制
def line_mode(flight_speed):
    global MODE, POS_COLOR, TARGET_SPD
    MODE = 'LINE'               # 模式旗標切為循跡
    POS_COLOR = ''              # 清空顏色目標
    TARGET_SPD = flight_speed   # 設定飛行速度
    # 回報模式啟用狀態
    print('<飛行狀態> 切換至循跡模式，目標飛行速度 %.2f m/s，開始往前飛行\n' % TARGET_SPD)


# ========= 顏色抵達偵測 ==============================================
# 條件式，佔用目前程序，直到偵測抵達目標顏色為止，需在循跡模式中使用
# 為避免發生離開後瞬間偵測抵達情形發生，故將面積閥值拉高作預防
def wait_arrive(color):
    # 初始化顏色面積
    img_area = 0

    # 色塊面積偵測
    i = 0
    while img_area < AREA_ARRIVE_LIMIT:
        # 更新面積數值
        x_raw, y_raw, img_area = image_moment(color)

        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 循跡中，差角 %.2f 度，角速度 %.2f 度/秒' % (ACTUAL_HDG, TARGET_YAW_RATE))
            print('<飛行狀態> 循跡中，飛行速度 %.2f m/s，俯仰角 %.2f 度' %
                  (ACTUAL_SPD, math.degrees(uav.attitude.pitch)))
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m\n' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報偵測到顏色
    print('<飛行狀態> 抵達%s色塊上方\n' % color)


# ========== 顏色離開偵測 =============================================
# 條件式，佔用目前程序，直到偵測到離開目標色塊為止，需在循跡模式中使用
def wait_leave(color):
    # 初始化顏色面積
    img_area = 100000

    # 色塊面積偵測
    i = 0
    while img_area > AREA_LEAVE_LIMIT:
        # 更新面積數值
        x_raw, y_raw, img_area = image_moment(color)

        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 循跡中，差角 %.2f 度，角速度 %.2f 度/秒' % (ACTUAL_HDG, TARGET_YAW_RATE))
            print('<飛行狀態> 循跡中，飛行速度 %.2f m/s，俯仰角 %.2f 度' %
                  (ACTUAL_SPD, math.degrees(uav.attitude.pitch)))
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m\n' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報離開的顏色
    print('<飛行狀態> 已離開%s色塊上方\n' % color)


# ========== 號誌轉向控制 =============================================
# 抵達號誌區後，進行燈號等待，並於偵測到顏色變化後做紀錄，並作相對應轉向
def turn():
    global TARGET_YAW_RATE, MISSION_COLOR, POS_COLOR
    # 轉彎等待與執行迴圈
    i = 0
    while 1:      # 藍右，綠左
        # 讀取顏色面積值做判斷
        block_x, block_y, image_area_blue = image_moment('blue')
        block_x, block_y, image_area_green = image_moment('green')

        # 顯示飛行資訊
        if i % 5 == 0:
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
            print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px^2' %
                  (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
            print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' %
                  (ACTUAL_VEL[0], ACTUAL_VEL[1]))

        # 讀取偏航角
        yaw_start = math.degrees(uav.attitude.yaw)
        yaw_current = math.degrees(uav.attitude.yaw)

        # 若偵測到藍色
        if image_area_blue >= TURN_AREA_LIMIT:
            # 記錄任務顏色並更改定位顏色
            MISSION_COLOR = 'blue'
            POS_COLOR = MISSION_COLOR
            print('<飛行狀態> 偵測到藍色為任務顏色')

            # 右轉
            print('<飛行命令> 往右旋轉90度\n')
            TARGET_YAW_RATE = TURN_YAW_RATE
            while math.fabs(yaw_current - yaw_start) < 90 * (TURN_SETTLE / 100):
                # 更新數值
                yaw_current = math.degrees(uav.attitude.yaw)

                # 顯示飛行資訊
                if i % 5 == 0:
                    print('<飛行狀態> 旋轉中，已轉 %d 度' % math.fabs(yaw_current - yaw_start))
                    print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
                    print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px^2' %
                          (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
                    print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' %
                          (ACTUAL_VEL[0], ACTUAL_VEL[1]))

                time.sleep(0.1)
                i += 1

                # 急停跳出
                if not STATE_THREAD:
                    break

            # 結束旋轉
            TARGET_YAW_RATE = 0
            break

        # 若偵測到綠色
        elif image_area_green >= TURN_AREA_LIMIT:
            # 記錄任務顏色並更改定位顏色
            MISSION_COLOR = 'green'
            POS_COLOR = MISSION_COLOR
            print('<飛行狀態> 偵測到綠色為任務顏色')

            # 左轉
            print('<飛行命令> 往左旋轉90度\n')
            TARGET_YAW_RATE = -TURN_YAW_RATE
            while math.fabs(yaw_current - yaw_start) < 90 * (TURN_SETTLE / 100):
                # 更新數值
                yaw_current = math.degrees(uav.attitude.yaw)

                # 顯示飛行資訊
                if i % 5 == 0:
                    print('<飛行狀態> 旋轉中，已轉 %d 度' % math.fabs(yaw_current - yaw_start))
                    print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
                    print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px^2' %
                          (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
                    print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' %
                          (ACTUAL_VEL[0], ACTUAL_VEL[1]))

                time.sleep(0.1)
                i += 1

                # 急停跳出
                if not STATE_THREAD:
                    break

            # 結束旋轉
            TARGET_YAW_RATE = 0
            break

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報轉向完成
    print('<飛行狀態> 轉向完成\n')


# ========== 空投程序 ================================================
# 於投擲區上方做定位穩定，並執行空投程序
def drop():
    # 切換定位模式
    pos_mode(MISSION_COLOR)

    # 定位迴圈，等待移至空投點上方
    print('<飛行命令> 移動至空投點上方')
    i = 0
    while 1:
        # 取得目前型心位置
        block_x, block_y, image_area = image_moment(MISSION_COLOR)

        # 顯示飛行資訊
        if i % 5 == 0:
            print('<高度感測器狀態> 超音波：%.2f m，IMU：%.2f m' % (ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU))
            print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px' % (ACTUAL_POS[0], ACTUAL_POS[1]))
            print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s\n' % (ACTUAL_VEL[0], ACTUAL_VEL[1]))

        # 判斷是否已抵達中心
        if -DROP_POS_LIMIT_X < block_x < DROP_POS_LIMIT_X:
            if -DROP_POS_LIMIT_Y < block_y < DROP_POS_LIMIT_Y:
                break

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 打開沙包盒空投
    print('<飛行命令> 位置已確認，執行空投\n')
    box_open()
    time.sleep(2)

    # 回報空投成功
    print('<飛行狀態> 沙包已成功空投\n')

    # 關閉沙包盒
    box_close()


# ====================================================================
# ========== 前景主函式區塊 ============================================
# ====================================================================
# 任務主函式
def mission():
    global STATE_THREAD, STATE_UAV, MISSION_COLOR
    print('<啟動程序> 任務程序：啟動')
    # 啟動多線程
    thread_activate()

    # 選擇任務模式
    mission_mode = int(input('請輸入任務模式1/2/3=>'))

    # 飛行任務1-定高測試
    if mission_mode == 1:
        pre_check()             # 執行起飛前檢查
        take_off(1)             # 起飛至任務高度
        hover(10)               # 懸停
        landing()               # 執行降落與關閉程序

    # 飛行任務1-定高&定位測試
    if mission_mode == 2:
        pre_check()             # 執行起飛前檢查
        take_off(1, 'red')      # 起飛至任務高度
        hover(10)               # 懸停
        landing()               # 執行降落與關閉程序

    # 飛行任務3-起飛區完整測試
    if mission_mode == 3:
        pre_check()             # 執行起飛前檢查
        take_off(1)             # 起飛至任務高度
        line_mode(0.3)          # 循跡模式
        wait_arrive('red')      # 等待抵達號誌區
        pos_mode('red')         # 紅燈定位保持
        turn()                  # 轉彎等待模式
        drop()                  # 空投測試
        landing()               # 執行降落與關閉程序

    # 飛行任務4-循線A區測試
    if mission_mode == 4:
        pre_check()
        take_off(1, 'red')
        hover(5)
        line_mode(0.3)          # 經過Ｌ轉角
        wait_leave('red')
        wait_arrive('blue')
        landing()

    # 飛行任務5-循線中色塊空投測試
    if mission_mode == 5:
        pre_check()
        take_off(1)
        line_mode(0.3)
        wait_arrive('green')
        pos_mode('green')
        MISSION_COLOR = 'green'
        drop()
        line_mode(0.3)
        wait_leave(MISSION_COLOR)
        wait_arrive('blue')
        landing('blue')

    # 完整版(未整理)
    if mission_mode == 9:
        # 起飛與紅燈區
        pre_check()             # 執行起飛前檢查
        take_off(1)             # 起飛至任務高度
        line_mode(0.3)          # 循跡模式
        wait_arrive('red')      # 等待抵達號誌區
        pos_mode('red')         # 紅燈定位保持

        # 循線A區
        turn()                  # 轉彎等待模式
        line_mode(0.3)
        wait_leave(MISSION_COLOR)

        # 投擲與循線B區
        wait_arrive(MISSION_COLOR)  # 偵測抵達投擲色塊上方
        drop()                      # 空投程序
        line_mode(0.3)              # 往前循跡飛行
        wait_leave(MISSION_COLOR)   # 偵測已離開投擲區

        # T路口與降落區
        wait_arrive(MISSION_COLOR)      # 偵測抵達降落區
        landing()                       # 執行降落

    # 關閉多線程
    STATE_THREAD = False
    print('<關閉程序> 主程序：關閉')

    # 清空GPIO設定
    GPIO.cleanup()
    print('<關閉程序> GPIO已關閉')

    # 結束與無人機連線
    uav.close()
    STATE_UAV = False
    print('<關閉程序> 無人機連線：關閉')

    # 輸出圖表
    data_sheet()
    alt_ctrl_data()
    alsen_data()
    pos_ctrl_data()
    vel_ctrl_data()
    hdg_ctrl_data()
    spd_ctrl_data()


# 執行程式
mission()
