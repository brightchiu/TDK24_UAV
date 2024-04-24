# 0最終版程式碼
# 依照日期分檔，過期無效

# 一般寫法
# 程式結構：參數宣告區->物件類函式->背景主函式->呼叫用函式->記錄函式->飛行功能函式->任務區塊

# 檔案日期 1101
# 測試情形：
# 定位(三色、距離遷入、速度遷入)、定高(感測器穩定前提無過衝)、轉彎定位能力(右轉通過)
# GUI版本-測試中

# 1101測試循線(面積與影像顯示測試、低空循線測試、顏色轉換模式更改)、T路口轉彎邏輯確認、下降階段應對規劃
# 沙包盒需做修正包裝

# !/usr/bin/env python
# -*- coding: utf-8 -*-

# 載入模組
from dronekit import connect, VehicleMode
from openpyxl import Workbook
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import RPi.GPIO as GPIO
import numpy as np
import threading
import time
import math
import cv2

# ====================================================================
# ========== 設定全域變數 ==============================================
# ====================================================================

# ========== UAV通訊參數 ==============================================
# 樹莓派Serial連線參數
UAV_CONNECT_PORT = '/dev/serial0'   # 傳輸口位址
UAV_CONNECT_BAUD = 921600           # 傳輸鮑率

# ========== 旗標與變數 ================================================
# ->旗標
STATE_THREAD = False            # 多線程旗標
STATE_UAV = True                # 無人機連線狀態
STATE_ALT = False               # 高度控制器旗標
STATE_POS = False               # 位置控制器旗標
STATE_VEL = False               # 位置-速度控制器旗標
STATE_HDG_OUT = False           # 航向-外迴圈控制器旗標
STATE_HDG_IN = False            # 航向-內迴圈控制器旗標
STATE_SPD = False               # 飛行速度控制器旗標
STATE_CAM = False               # 影像擷取程式旗標
STATE_SENSOR_CAM = False        # 影像感測器旗標
STATE_SENSOR_ALT = False        # 高度感測器旗標
STATE_SENSOR_SONAR = False      # 超音波高度計旗標
STATE_SENSOR_SPD = False        # 飛行速度感測器旗標
STATE_CMD = False               # 命令傳輸程式旗標
STATE_BOX = False               # 貨物艙門狀態旗標
MODE = ''                       # 模式旗標(POS:定位模式，LINE:循跡模式)
MISSION_COLOR = ''              # 任務顏色(blue or green)
POS_COLOR = ''                  # 定位的色塊顏色

# ->多線程優先執行命令
THREAD_OCCUPY = threading.Lock()

# ->目標值(命令值)
TARGET_ROLL = 0                 # 目標滾轉角(degree)
TARGET_PITCH = 0                # 目標俯仰角(degree)
TARGET_YAW = 0                  # 目標偏航角(degree)，不使用
TARGET_YAW_RATE = 0             # 偏航角速度(degree/s)，需搭配控制器控制角度
TARGET_THRUST = 0               # 目標推力，表示升降速度(0.5為中點)
TARGET_ALT = 0                  # 目標高度，主程式輸入於高度控制器(m)
TARGET_HDG = 0                  # 目標偏航角(degree)
TARGET_VEL = [0, 0]             # 目標移動速度(px/s，相機座標)
TARGET_POS = [0, 0]             # 目標型心位置(px，相機座標)
TARGET_SPD = 0                  # 目標飛行速度
TARGET_LINE_POS = 0             # 線型心目標座標(px，相機座標)
TARGET_LINE_VEL = 0             # 線型心目標速度(px/s，相機座標)

# ->實際值(量測值)
ACTUAL_ALT = 0                  # 實際高度(m)
ACTUAL_ALT_SONAR = 0            # 超音波高度計(m)
ACTUAL_ALT_IMU = 0              # IMU氣壓高度計(m)
ACTUAL_LINE_POS = 0             # 線距離(px，Y軸為基準，Roll控制器回授用)
ACTUAL_LINE_VEL = 0             # 線速度(px/s，Y軸為基準，Roll控制器回授用)
ACTUAL_LINE_AREA  = 0           # 線面積(px)
ACTUAL_HDG = 0                  # 線角(degree，Y軸為基準)
ACTUAL_POS = [0, 0]             # 色塊實際位置(px)
ACTUAL_VEL = [0, 0]             # 色塊實際速度(px/s)
ACTUAL_AREA = 0                 # 色塊面積(px)
ACTUAL_SPD = 0                  # 實際飛行速度

# ========== 控制器參數設定 =============================================
# ->迴圈頻率(預設20Hz)
FREQ_ALT = 20                   # 高度控制器頻率(不快於高度感測器頻率)
FREQ_POS = 10                   # 位置控制器頻率(不快於位置-速度控制器頻率)
FREQ_VEL = 20                   # 位置-速度控制器頻率(不快於影像禎率)
FREQ_HDG_OUT = 10               # 航向控制器-外迴圈頻率(不快於內迴圈頻率)
FREQ_HDG_IN = 20                # 航向控制器-內迴圈頻率(不快於影像禎率)
FREQ_SPD = 15                   # 飛行速度控制器頻率
FREQ_CAM = 30                   # 影像禎率(fps，依系統負載調整)
FREQ_SENSOR = 20                # 感測器資料更新率(外部資料)
FREQ_SENSOR_CAM = 30            # 影像感測器頻率(不快於影像禎率)
FREQ_SENSOR_ALT = 20            # 高度感測器頻率(不快於超音波感測器頻率)
FREQ_SENSOR_SONAR = 15          # 超音波感測器頻率(5~50Hz，越高雜訊越多)
FREQ_SENSOR_SPD = 20            # 飛行速度感測器頻率
FREQ_CMD = 20                   # 命令傳輸頻率(不快於控制器頻率)

# ->高度控制器參數(第一階段完成，待外部感測建置完成後再優化)
ALT_KP = 0.15                   # P增益值*
ALT_KI = 0                      # I增益值*
ALT_KD = 0.15                   # D增益值*
ALT_GAIN_D_LIMIT = 0.5          # D增益上限值*
ALT_GD_SAMPLE_NUM_AVE = 5       # D增益濾波器(取樣次數，最少2)*
ALT_SPEED = 0.5                 # 昇降速率(m/s，與飛控參數一致)*

# ->位置控制器參數(1031已完成)
POS_X_KP = 0.4                  # 位置控制器X軸-P增益值
POS_X_KI = 0                    # 位置控制器X軸-I增益值
POS_X_KD = 0                    # 位置控制器X軸-D增益值
POS_Y_KP = 0.4                  # 位置控制器Y軸-P增益值
POS_Y_KI = 0                    # 位置控制器Y軸-I增益值
POS_Y_KD = 0                    # 位置控制器Y軸-D增益值
POS_X_DZ = 30                   # 位置控制器輸入截止區參數(於截止區內使速度控制器控制零速度，防震盪)
POS_Y_DZ = 30                   # 位置控制器輸入截止區參數(於截止區內使速度控制器控制零速度，防震盪)
POS_SAMPLE_NUM_AVE = 10         # 位置控制器均值濾波取樣次數*
POS_GD_SAMPLE_NUM_AVE = 5       # D增益濾波器(取樣次數，最少2)*
# 特殊附註：截止區調校方式=>區域內阻力可以抵消慣性、控制器解析程度內(控制器截止區)

# ->位置-速度控制器參數(1031已完成)
VEL_X_KP = 0.1                  # 速度控制器X軸-P增益值
VEL_X_KI = 0                    # 速度控制器X軸-I增益值
VEL_X_KD = 0.1                  # 速度控制器X軸-D增益值
VEL_Y_KP = 0.1                  # 速度控制器Y軸-P增益值
VEL_Y_KI = 0                    # 速度控制器Y軸-I增益值
VEL_Y_KD = 0.1                  # 速度控制器Y軸-D增益值
VEL_GD_SAMPLE_NUM_AVE = 5       # D增益濾波器(取樣次數，最少2)*
VEL_ANGLE_LIMIT = 3             # 速度控制器-最大飛行角度(degree，Min>3)
VEL_ANGLE_CUTOFF = 0.5          # 速度控制器-最小飛行角度(degree)
VEL_X_KH = 0.25                 # X軸回授增益(與控制器增益有相依性)
VEL_Y_KH = 0.25                 # Y軸回授增益(與控制器增益有相依性)
VEL_X_DZ = 0                    # 輸入截止區參數(於截止區內使速度控制器輸出零角度，防震盪)
VEL_Y_DZ = 0                    # 輸入截止區參數(於截止區內使速度控制器輸出零角度，防震盪)
VEL_CUTOFF_ALT = 0.2            # 控制器最低作動高度(m，飽和高度點)

# ->循跡-航向控制器參數
HDG_YAW_KP = 0.6                # P增益值*
HDG_YAW_KI = 0                  # I增益值*
HDG_YAW_KD = 0                  # D增益值*
HDG_YAW_GD_SAMPLE_NUM_AVE = 5   # D增益濾波器(取樣次數，最少2)*
HDG_YAW_DZ = 5                  # 輸入截止區角度(degree)*
HDG_YAW_RATE_LIMIT = 30         # 角速度輸出限制(degree/s)*
HDG_YAW_RATE_CUTOFF = 5         # 角速度輸出截止(degree/s)*

# ->循跡-滾轉控制器參數(比照定位X軸處理)
# 外迴圈
HDG_X_KP = 0.4                  # 位置控制器P增益值(回至中心線速度)*
HDG_X_KI = 0                    # 位置控制器I增益值(有穩差才使用)*
HDG_X_KD = 0                    # 位置控制器D增益值(減少震盪並拉快安定時間)*
HDG_X_DZ = 30                   # 位置控制器輸入截止區參數(於截止區內使速度控制器控制零速度，防震盪)
HDG_SAMPLE_NUM_AVE = 10         # 位置控制器均值濾波取樣次數*
HDG_X_GD_SAMPLE_NUM_AVE = 5     # D增益濾波器(取樣次數，最少2)*
# 內迴圈
HDG_VX_KP = 0.1                 # 速度控制器P增益值(回至中心線速度)*
HDG_VX_KI = 0                   # 速度控制器I增益值(有穩差才使用)*
HDG_VX_KD = 0.1                 # 速度控制器D增益值(減少震盪並拉快安定時間)*
HDG_VX_GD_SAMPLE_NUM_AVE = 5    # D增益濾波器(取樣次數，最少2)*
HDG_ROLL_LIMIT = 3              # 速度控制器-最大飛行角度(degree)
HDG_ROLL_CUTOFF = 0.5           # 速度控制器-最小飛行角度(degree)
HDG_VX_KH = 0.25                # X軸回授增益(與控制器增益有相依性)
HDG_VX_DZ = 0                   # 輸入截止區參數(於截止區內使速度控制器輸出零角度，防震盪)

# ->飛行速度控制器參數-閉迴路
SPD_MODE = 1                    # 飛行模式(1:閉迴路模式 0:開迴路模式)
SPD_KP = 0                      # P增益值*
SPD_KI = 0                      # I增益值*
SPD_KD = 0                      # D增益值*
SPD_GD_SAMPLE_NUM_AVE = 5       # D增益濾波器(取樣次數，最少2)*
SPD_ANGLE_CUTOFF = 0.5          # 飛行速度控制器-最小俯仰角度(degree)
SPD_ANGLE_LIMIT = 1             # 飛行速度控制器-最大俯仰角度(degree)
SPD_OFF_ANGLE = 15              # 切換跳出角度(degree)，慣性抵銷用
SPD_SAMPLE_NUM_AVE = 5          # 飛行速度感測器均值濾波取樣次數*

# ->飛行速度控制器參數-開迴路
SPD_OL_ACTIVE_TIME = 0          # 開迴路控制模式，角度作動時長(s)
SPD_OL_OFF_TIME = 0             # 開迴路控制模式，角度關閉時長(s)
SPD_OL_ANGLE = 3                # 開迴路控制模式，角度作動量(degree)

# ->飛行程序參數
TAKEOFF_SETTLE = 95             # 安定高度百分比(%，高度確認門檻)*
LANDING_CUTOFF = 0.05           # 油門關閉高度(m，停止油門輸出)*
LANDING_POS_LIMIT = 50          # 降落定位設定
TURN_SETTLE = 95                # 安定轉角百分比(%，停止轉向輸出)*
TURN_YAW_RATE = 30              # 轉角速度(degree/s)*
TURN_AREA_LIMIT = 3000          # 號誌燈號面積閥值(px)
DROP_POS_LIMIT_X = 50           # 空投定位X範圍限制(單向，px)*
DROP_POS_LIMIT_Y = 50           # 空投定位Y範圍限制(單向，px)*
AREA_ARRIVE_LIMIT = 4000        # 色塊抵達偵測門檻(px)
AREA_LEAVE_LIMIT = 3000         # 色塊離開偵測門檻(px)
T_JUDGE_TIMES = 5               # T字路口判斷閥值(次數)

# ========== 影像參數設定 ===============================================
# ->相機設定參數
CAM_VIDEO = cv2.VideoCapture(0)             # 影像來源(預設來源為0)
CAM_IMAGE = 0                               # 儲存影像
CAM_IMAGE_LIVE = 0                          # 即時影像
CAM_IMAGE_WIDTH = 480                       # 影像寬度(px，4:3設定 )
CAM_IMAGE_HEIGHT = 360                      # 影像高度(px)
VIDEO_FOURCC = cv2.VideoWriter_fourcc(*'XVID')
VIDEO_SAVE = cv2.VideoWriter('Flight Camera.avi',
                             VIDEO_FOURCC, FREQ_CAM, (CAM_IMAGE_WIDTH, CAM_IMAGE_HEIGHT))

# ->影像計算參數
IMG_BLUE_L = np.array([85, 76, 90])         # 藍色色域-下限
IMG_BLUE_H = np.array([127, 255, 255])      # 藍色色域-上限
IMG_GREEN_L = np.array([44, 104, 17])       # 綠色色域-下限
IMG_GREEN_H = np.array([87, 255, 255])      # 綠色色域-上限
IMG_RED_L = np.array([160, 111, 56])        # 紅色色域-下限
IMG_RED_H = np.array([180, 255, 255])       # 紅色色域-上限
IMG_BLACK_L = np.array([0, 0, 0])           # 黑色色域-下限
IMG_BLACK_H = np.array([180, 255, 75])      # 黑色色域-上限
IMG_MORPHOLOGY_EX_KERNEL = (3, 3)           # 閉運算矩陣大小(奇數，越大降噪能力越強)*
IMG_AREA_LIMIT = 3000                       # 取用面積下限值*
IMG_SAMPLE_NUM_AVE = 5                      # 均值濾波取樣次數*
BODY_REF_X = int(CAM_IMAGE_WIDTH / 2)       # 畫面參考X座標中心
BODY_REF_Y = int(CAM_IMAGE_HEIGHT / 2)      # 畫面位置Y座標中心
BODY_OFFSET_X = 0                           # 機身中心X偏移量(px，中心往左移正)*
BODY_OFFSET_Y = 20                          # 機身中心Y偏移量(px，中心往下移正)*

# ->線角度計算
ANG_THRESHOLD = 35              # 二值化門檻值(調高比較能看到線，但可能有雜訊)*
ANG_THRESHOLD_NEW = 255         # 取得黑白照填入最高值
ANG_GB_KERNEL_SIZE = (7, 7)     # 高斯模糊矩陣大小(奇數，越大越模糊)*
ANG_GB_SIGMA = 2                # 高斯模糊色域標準差(越大越模糊)*
ANG_LINE_TRACK_AREA_L = 500     # 線追蹤型心計算面積下限*
ANG_LINE_TRACK_AREA_H = 5000    # 線追蹤型心計算面積下限(僅限制角度輸出)
ANG_SAMPLE_NUM_AVE = 5          # 航向控制器均值濾波取樣次數*
ANG_T_INTERSECT_THRESHOLD_DN = 5    # T路口邊界判斷閥值下限(px，依據線寬調整)
ANG_T_INTERSECT_THRESHOLD_UP = 30   # T路口邊界判斷閥值上限(px，依據線寬調整)
ANG_AFTER_DROP = False          # 是否為空投後狀態
ANG_T_TURN = ''                 # T路口狀態保留

# ========== 貨物艙參數 ================================================
# ->沙包盒參數
BOX_PWM_PIN = 18                # 沙包盒PWM輸出腳位
BOX_PWM_FREQ = 50               # PWM頻率(方波Hz)
BOX_PWM_OPEN = 1000             # 開門位置(400~2350us)*
BOX_PWM_CLOSE = 2000            # 關門位置(400~2350us)*

# ========== 高度感測器參數 =============================================
# ->高度感測器整合參數
WEIGHTED_SONAR = 8              # 超音波高度計權重*
WEIGHTED_IMU = 1                # IMU高度計權重*

# ->超音波高度計參數
SONAR_AIR_TEMP = 25             # 氣溫(攝氏)*
SONAR_TRIG_PIN = 20             # TRIG輸出腳位
SONAR_ECHO_PIN = 21             # ECHO輸入腳位
SONAR_TIMEOUT = 0.015           # 無回應超時限制(秒，過低時遠距離偵測將會變成超時)*
SONAR_OFFSET = 0                # 高度校正偏移量(m，向上為正)
SONAR_SAMPLE_NUM_CAL = 10       # 高度校正取樣次數*
SONAR_SAMPLE_NUM_AVE = 10       # 均值濾波取樣次數*
SONAR_KF_INPUT_MAX = 2          # 卡爾曼濾波器輸入限制(m)*
SONAR_KF_Q = 5                  # 卡爾曼濾波器滑順度(越小反應快，越不平；越大反應慢，但較平滑)*
SONAR_ERR_COUNT = 0             # 量測超時次數記錄
SONAR_FILTER_ALT = 1            # 突波濾波作用起始高度(m)
SONAR_MAX_CHANGE = 0.5          # 突波容許值(m)

# ->IMU高度計參數
IMU_OFFSET = 0                  # 高度校正偏移量(m，向上為正)
IMU_SAMPLE_NUM_CAL = 5          # 高度校正取樣次數*
IMU_SAMPLE_NUM_AVE = 10         # 均值濾波取樣次數*
IMU_KF_INPUT_MAX = 2            # 卡爾曼濾波器輸入限制(m)*
IMU_KF_Q = 5                    # 卡爾曼濾波器滑順度(越小反應快，越不平；越大反應慢，但較平滑)*

# ========== 數值記錄 ==================================================
# ->高度控制器
R_TARGET_ALT = []                       # 記錄目標高
R_ACTUAL_ALT_CTRL = []                  # 記錄實際高
R_TARGET_THRUST = []                    # 記錄目標推力
R_ALT_ERROR = []                        # 記錄高度誤差
R_ALT_GP = []                           # 記錄P增益
R_ALT_GI = []                           # 記錄I增益
R_ALT_GD = []                           # 記錄D增益

# ->位置控制器
R_POS_X = []                            # 記錄X座標
R_POS_Y = []                            # 記錄Y座標
R_POS_AREA = []                         # 記錄面積變化
R_POS_RECORD_SCALE = 1000               # 面積縮放比
R_POS_X_ERROR = []                      # 記錄誤差值
R_POS_Y_ERROR = []                      # 記錄誤差值
R_POS_X_GP = []                         # 記錄P增益
R_POS_X_GI = []                         # 記錄I增益
R_POS_X_GD = []                         # 記錄D增益
R_POS_Y_GP = []                         # 記錄P增益
R_POS_Y_GI = []                         # 記錄I增益
R_POS_Y_GD = []                         # 記錄D增益
R_POS_X_TARGET_VEL = []                 # 記錄輸出的目標速度值
R_POS_Y_TARGET_VEL = []                 # 記錄輸出的目標速度值

# ->速度控制器
R_POS_TARGET_ROLL = []                  # 記錄目標滾轉角
R_POS_TARGET_PITCH = []                 # 記錄目標俯仰角
R_POS_ACTUAL_ROLL = []                  # 記錄實際滾轉角
R_POS_ACTUAL_PITCH = []                 # 記錄實際俯仰角
R_VEL_X_TAR = []                        # 記錄X目標速度
R_VEL_Y_TAR = []                        # 記錄Y目標速度
R_POS_VEL_X_ACT = []                    # 記錄X實際速度
R_POS_VEL_Y_ACT = []                    # 記錄Y實際速度
R_POS_VEL_X_ERROR = []                  # 記錄誤差值
R_POS_VEL_Y_ERROR = []                  # 記錄誤差值
R_POS_VEL_X_GP = []                     # 記錄P增益
R_POS_VEL_X_GI = []                     # 記錄I增益
R_POS_VEL_X_GD = []                     # 記錄D增益
R_POS_VEL_Y_GP = []                     # 記錄P增益
R_POS_VEL_Y_GI = []                     # 記錄I增益
R_POS_VEL_Y_GD = []                     # 記錄D增益

# ->航向控制器
# 外迴圈
R_HDG_LINE_ACTUAL_X = []                # 線X座標
R_HDG_LINE_TARGET_VX = []               # 目標X速度
R_HDG_LINE_X_ERROR = []                 # 記錄距離誤差
R_HDG_LINE_X_GP = []                    # 記錄P增益
R_HDG_LINE_X_GI = []                    # 記錄I增益
R_HDG_LINE_X_GD = []                    # 記錄D增益

# 內迴圈Roll
R_HDG_LINE_VX_TARGET = []               # 目標X速度
R_HDG_LINE_VX_ACTUAL = []               # 目標X速度
R_HDG_LINE_VX_TARGET_ROLL = []          # 輸出Roll角
R_HDG_LINE_VX_ACTUAL_ROLL = []          # 實際Roll角
R_HDG_LINE_VX_ERROR = []                # 記錄距離誤差
R_HDG_LINE_VX_GP = []                   # 記錄P增益
R_HDG_LINE_VX_GI = []                   # 記錄I增益
R_HDG_LINE_VX_GD = []                   # 記錄D增益

# 內迴圈Yaw
R_HDG_YAW_ACTUAL = []                   # 記錄實際角度
R_HDG_YAW_TARGET_YAW_RATE = []          # 記錄目標偏航速率
R_HDG_YAW_ERROR = []                    # 記錄角度誤差
R_HDG_YAW_GP = []                       # 記錄P增益
R_HDG_YAW_GI = []                       # 記錄I增益
R_HDG_YAW_GD = []                       # 記錄D增益

# ->飛行速度控制器
R_SPD_TARGET = []                       # 記錄目標速度
R_SPD_ACTUAL = []                       # 記錄實際速度
R_SPD_TARGET_PITCH = []                 # 記錄目標俯仰角
R_SPD_ACTUAL_PITCH = []                 # 記錄實際俯仰角
R_SPD_ERROR = []                        # 記錄速度誤差
R_SPD_GP = []                           # 記錄P增益
R_SPD_GI = []                           # 記錄I增益
R_SPD_GD = []                           # 記錄D增益
R_SPD_OL_TARGET_PITCH = []              # 開迴路目標俯仰角
R_SPD_OL_ACTUAL_PITCH = []              # 開迴路實際俯仰角
R_SPD_OL_ACTUAL_SPD = []                # 開迴路實際速度

# ->飛行高度感測器
R_ACTUAL_ALT = []                       # 記錄加權值
R_ACTUAL_ALT_SONAR = []                 # 記錄超音波高度計值

# ->超音波高度計
R_ACTUAL_ALT_SONAR_RAW = []             # 記錄原始值
R_ACTUAL_ALT_SONAR_OFFSET = []          # 記錄補正值
R_ACTUAL_ALT_SONAR_AVE = []             # 記錄平均值
R_ACTUAL_ALT_SONAR_KF = []              # 記錄KF值

# ->氣壓高度計
R_ACTUAL_ALT_IMU_RAW = []               # 記錄原始值
R_ACTUAL_ALT_IMU_OFFSET = []            # 記錄補正值
R_ACTUAL_ALT_IMU_AVE = []               # 記錄平均值
R_ACTUAL_ALT_IMU_KF = []                # 記錄KF值

print('<啟動程序> 載入全域變數設定完畢')


# ========== 與UAV執行連線 =============================================
print('<啟動程序> 與UAV連線中，請稍候...')

# 連線指令
t_start = time.time()
uav = connect(UAV_CONNECT_PORT, wait_ready=True, baud=UAV_CONNECT_BAUD)
t_end = time.time()
print('<啟動程序> UAV已連線，花費%.2d秒' % (t_end-t_start))

# ====================================================================
# ========== 背景主函式函式區塊 =========================================
# ====================================================================
# 本區為主要運作的函式，以控制器為主，包含高度、位置、航向、雙高度感測器、
# 相機、命令傳送，多線程執行時能達到協作的處理功能
# (但仍受每個迴圈處理時間不一影響，為非同步處理)


# ========== 高度控制器 ================================================
# 說明：
# 為PID控制器型態，於任務期間全程作用，輸入目標高度後，依據高度感測器執行誤差PID增益，
# 最終輸出成爬升速率型態(符合MAVLink指令)，設有輸出範圍限制器、中點飽和區間功能
def altitude_controller():
    global TARGET_THRUST, STATE_ALT
    # PID變數設定
    int_error = 0
    previous_error = 0
    previous_time = time.time()

    # 均值濾波器
    gain_d_register = []

    # 等待感測器啟動
    if not STATE_SENSOR_ALT:
        print('<啟動程序> 高度控制器：等待高度感測器啟動')
    while not STATE_SENSOR_ALT:
        time.sleep(0.1)

    # 回報高度控制器啟動
    STATE_ALT = True
    print('<啟動程序> 高度控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 同捆執行起點
        THREAD_OCCUPY.acquire()
        t_loop = time.time()

        # 計算誤差值
        error = TARGET_ALT - ACTUAL_ALT

        # 誤差值微積分計算
        current_time = time.time()
        delta_error = error - previous_error
        delta_time = current_time - previous_time
        int_error += (error * delta_time)
        derivative = delta_error / delta_time

        # 誤差增益處理
        gain_p = error * ALT_KP
        gain_i = int_error * ALT_KI
        gain_d = derivative * ALT_KD

        # 微分增益限制
        if gain_d > ALT_GAIN_D_LIMIT:
            gain_d = ALT_GAIN_D_LIMIT
        elif gain_d < -ALT_GAIN_D_LIMIT:
            gain_d = -ALT_GAIN_D_LIMIT

        # 微分項均值濾波器
        gain_d_register.insert(0, gain_d)
        if len(gain_d_register) > ALT_GD_SAMPLE_NUM_AVE:
            gain_d_register.pop()
        gain_d = np.average(gain_d_register)

        speed_out = gain_p + gain_i + gain_d

        # 將速度以比例輸出
        thrust_out = speed_out / ALT_SPEED

        # 補正至中點
        thrust_out += 0.5

        # 輸出限制(0下降max ~ 1上升max)
        if thrust_out > 1:
            thrust_out = 1
        if thrust_out < 0:
            thrust_out = 0

        # 更新推力值至全域變數
        TARGET_THRUST = thrust_out
        
        # 傳遞值給下次迴圈使用
        previous_error = error
        previous_time = current_time

        # 記錄數值
        R_TARGET_ALT.append(TARGET_ALT)
        R_ACTUAL_ALT_CTRL.append(ACTUAL_ALT)
        R_TARGET_THRUST.append(TARGET_THRUST)
        R_ALT_ERROR.append(error)
        R_ALT_GP.append(gain_p)
        R_ALT_GI.append(gain_i)
        R_ALT_GD.append(gain_d)

        # 同捆執行終點
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        t_wait = (1 / FREQ_ALT) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_ALT = False
    print('<關閉程序> 高度控制器：關閉')


# ========== 位置控制器 ================================================
# 說明：
# 外迴圈的位置控制
# 先以影像取得的色塊座標計算速度增益值，在輸出至速度控制器做角度-加速度控制
def position_controller():
    global TARGET_VEL, STATE_POS
    # PID變數設定
    pos_int_error_x = 0
    pos_int_error_y = 0
    pos_previous_error_x = 0
    pos_previous_error_y = 0
    previous_time = time.time()

    # 均值濾波設定
    pos_gain_d_x_register = []
    pos_gain_d_y_register = []

    # 等待影像感測器啟動
    if not STATE_SENSOR_CAM:
        print('<啟動程序> 位置控制器：等待影像感測器啟動')
    while not STATE_SENSOR_CAM:
        time.sleep(0.1)

    # 回報位置控制器啟用
    STATE_POS = True
    reset = False           # 重設旗標(暫存清除安全標記)
    print('<啟動程序> 位置控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        while MODE == 'POS' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 讀取回授值
            pos_actual_x = ACTUAL_POS[0]
            pos_actual_y = ACTUAL_POS[1]

            # 回授值截止區修正
            if TARGET_POS[0] - POS_X_DZ < pos_actual_x < TARGET_POS[0] + POS_X_DZ:
                pos_actual_x = TARGET_POS[0]
            if TARGET_POS[1] - POS_Y_DZ < pos_actual_y < TARGET_POS[1] + POS_Y_DZ:
                pos_actual_y = TARGET_POS[1]

            # 計算誤差值
            pos_error_x = TARGET_POS[0] - pos_actual_x
            pos_error_y = TARGET_POS[1] - pos_actual_y

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            pos_delta_error_x = pos_error_x - pos_previous_error_x
            pos_int_error_x += (pos_error_x * delta_time)
            pos_derivative_x = pos_delta_error_x / delta_time

            pos_delta_error_y = pos_error_y - pos_previous_error_y
            pos_int_error_y += (pos_error_y * delta_time)
            pos_derivative_y = pos_delta_error_y / delta_time

            # X軸誤差增益處理
            pos_gain_p_x = pos_error_x * POS_X_KP
            pos_gain_i_x = pos_int_error_x * POS_X_KI
            pos_gain_d_x = pos_derivative_x * POS_X_KD

            # 微分項均值濾波器
            pos_gain_d_x_register.insert(0, pos_gain_d_x)
            if len(pos_gain_d_x_register) > POS_GD_SAMPLE_NUM_AVE:
                pos_gain_d_x_register.pop()
            pos_gain_d_x = np.average(pos_gain_d_x_register)

            vel_x_out = pos_gain_p_x + pos_gain_i_x + pos_gain_d_x

            # Y軸誤差增益處理
            pos_gain_p_y = pos_error_y * POS_Y_KP
            pos_gain_i_y = pos_int_error_y * POS_Y_KI
            pos_gain_d_y = pos_derivative_y * POS_Y_KD

            # 微分項均值濾波器
            pos_gain_d_y_register.insert(0, pos_gain_d_y)
            if len(pos_gain_d_y_register) > POS_GD_SAMPLE_NUM_AVE:
                pos_gain_d_y_register.pop()
            pos_gain_d_y = np.average(pos_gain_d_y_register)

            vel_y_out = pos_gain_p_y + pos_gain_i_y + pos_gain_d_y

            # 更新速度命令值至全域變數
            TARGET_VEL = [vel_x_out, vel_y_out]

            # 傳遞值給下次迴圈使用
            pos_previous_error_x = pos_error_x
            pos_previous_error_y = pos_error_y
            previous_time = current_time

            # 記錄數值
            R_POS_X.append(ACTUAL_POS[0])
            R_POS_Y.append(ACTUAL_POS[1])
            R_POS_AREA.append(ACTUAL_AREA / R_POS_RECORD_SCALE)

            R_POS_X_ERROR.append(pos_error_x)
            R_POS_Y_ERROR.append(pos_error_y)
            R_POS_X_GP.append(pos_gain_p_x)
            R_POS_X_GI.append(pos_gain_i_x)
            R_POS_X_GD.append(pos_gain_d_x)
            R_POS_Y_GP.append(pos_gain_p_y)
            R_POS_Y_GI.append(pos_gain_i_y)
            R_POS_Y_GD.append(pos_gain_d_y)

            R_POS_X_TARGET_VEL.append(vel_x_out)
            R_POS_Y_TARGET_VEL.append(vel_y_out)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1/FREQ_POS)-(time.time()-t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為定位時，跳出迴圈並暫停控制器
            if not MODE == 'POS':
                # 重設暫存器
                pos_gain_d_x_register = []
                pos_gain_d_y_register = []
                pos_int_error_x = 0
                pos_int_error_y = 0
                pos_previous_error_x = 0
                pos_previous_error_y = 0
                previous_time = time.time()

                # 歸零速度命令數值
                TARGET_VEL = [0, 0]

                # 回報控制器狀態
                STATE_POS = False
                reset = True
                print('<飛行狀態> 位置控制器：暫停')

        # 未成功執行跳出重設時，執行重設
        if not reset:
            pos_gain_d_x_register = []
            pos_gain_d_y_register = []
            pos_int_error_x = 0
            pos_int_error_y = 0
            pos_previous_error_x = 0
            pos_previous_error_y = 0
            previous_time = time.time()

            # 歸零速度命令數值
            TARGET_VEL = [0, 0]

            # 回報控制器狀態
            STATE_POS = False
            reset = True
            print('<飛行狀態> 位置控制器：暫停')

        # 記錄數值
        t_loop = time.time()
        R_POS_X.append(0)
        R_POS_Y.append(0)
        R_POS_AREA.append(0)

        R_POS_X_ERROR.append(0)
        R_POS_Y_ERROR.append(0)
        R_POS_X_GP.append(0)
        R_POS_X_GI.append(0)
        R_POS_X_GD.append(0)
        R_POS_Y_GP.append(0)
        R_POS_Y_GI.append(0)
        R_POS_Y_GD.append(0)

        R_POS_X_TARGET_VEL.append(0)
        R_POS_Y_TARGET_VEL.append(0)

        # 迴圈執行間隔
        t_wait = (1/FREQ_POS)-(time.time()-t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

        # 當飛行模式為定位時，跳入迴圈並啟動函式
        if MODE == 'POS' and STATE_THREAD:
            # 回報控制器狀態
            STATE_POS = True
            reset = False
            print('<飛行狀態> 位置控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_POS = False
    print('<關閉程序> 位置控制器：關閉')


# ========== 位置-速度控制器 ============================================
# 說明：
# 位置控制器的內迴圈，負責控制與色塊相對速度
def position_velocity_controller():
    global TARGET_PITCH, TARGET_ROLL, TARGET_VEL, ACTUAL_POS, ACTUAL_AREA, STATE_VEL
    # PID變數設定
    vel_int_error_x = 0
    vel_int_error_y = 0
    vel_previous_error_x = 0
    vel_previous_error_y = 0
    previous_time = time.time()

    # 均值濾波器設定
    vel_gain_d_x_register = []
    vel_gain_d_y_register = []

    # 等待位置感測器啟動
    if not STATE_POS:
        print('<啟動程序> 速度控制器：等待位置控制器啟動')
    while not STATE_POS:
        time.sleep(0.1)

    # 回報速度控制器啟用
    STATE_VEL = True
    reset = False           # 重設旗標(暫存清除安全標記)
    print('<啟動程序> 速度控制器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        while MODE == 'POS' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 誤差增益
            vel_actual_x = ACTUAL_VEL[0] * VEL_X_KH
            vel_actual_y = ACTUAL_VEL[1] * VEL_Y_KH
            
            # 回授值截止區修正
            if TARGET_VEL[0] - VEL_X_DZ < ACTUAL_VEL[0] < TARGET_VEL[0] + VEL_X_DZ:
                vel_actual_x = TARGET_VEL[0]
            if TARGET_VEL[1] - VEL_Y_DZ < ACTUAL_VEL[1] < TARGET_VEL[1] + VEL_Y_DZ:
                vel_actual_y = TARGET_VEL[1]
                
            # 計算誤差值
            vel_error_x = TARGET_VEL[0] - vel_actual_x
            vel_error_y = TARGET_VEL[1] - vel_actual_y

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            vel_delta_error_x = vel_error_x - vel_previous_error_x
            vel_int_error_x += (vel_error_x * delta_time)
            vel_derivative_x = vel_delta_error_x / delta_time

            vel_delta_error_y = vel_error_y - vel_previous_error_y
            vel_int_error_y += (vel_error_y * delta_time)
            vel_derivative_y = vel_delta_error_y / delta_time

            # X軸誤差增益處理
            vel_gain_p_x = vel_error_x * VEL_X_KP
            vel_gain_i_x = vel_int_error_x * VEL_X_KI
            vel_gain_d_x = vel_derivative_x * VEL_X_KD

            # 微分項均值濾波器
            vel_gain_d_x_register.insert(0, vel_gain_d_x)
            if len(vel_gain_d_x_register) > VEL_GD_SAMPLE_NUM_AVE:
                vel_gain_d_x_register.pop()
            vel_gain_d_x = np.average(vel_gain_d_x_register)

            roll_out = vel_gain_p_x + vel_gain_i_x + vel_gain_d_x

            # Y軸誤差增益處理
            vel_gain_p_y = vel_error_y * VEL_Y_KP
            vel_gain_i_y = vel_int_error_y * VEL_Y_KI
            vel_gain_d_y = vel_derivative_y * VEL_Y_KD

            # 微分項均值濾波器
            vel_gain_d_y_register.insert(0, vel_gain_d_y)
            if len(vel_gain_d_y_register) > VEL_GD_SAMPLE_NUM_AVE:
                vel_gain_d_y_register.pop()
            vel_gain_d_y = np.average(vel_gain_d_y_register)

            pitch_out = vel_gain_p_y + vel_gain_i_y + vel_gain_d_y

            # 角度限制(飽和區限幅與截止區處理)
            if roll_out > VEL_ANGLE_LIMIT:
                roll_out = VEL_ANGLE_LIMIT
            elif roll_out < -VEL_ANGLE_LIMIT:
                roll_out = -VEL_ANGLE_LIMIT
            elif -VEL_ANGLE_CUTOFF < roll_out < VEL_ANGLE_CUTOFF:
                roll_out = 0

            if pitch_out > VEL_ANGLE_LIMIT:
                pitch_out = VEL_ANGLE_LIMIT
            elif pitch_out < -VEL_ANGLE_LIMIT:
                pitch_out = -VEL_ANGLE_LIMIT
            elif -VEL_ANGLE_CUTOFF < pitch_out < VEL_ANGLE_CUTOFF:
                pitch_out = 0

            # 輸出至全域變數
            if ACTUAL_ALT < VEL_CUTOFF_ALT:    # 控制器最低作動高度(影像飽和高度)
                TARGET_ROLL = 0
                TARGET_PITCH = 0
            else:
                TARGET_ROLL = -roll_out
                TARGET_PITCH = pitch_out

            # 傳遞值給下次迴圈使用
            vel_previous_error_x = vel_error_x
            vel_previous_error_y = vel_error_y
            previous_time = current_time

            # 記錄數值
            R_VEL_X_TAR.append(TARGET_VEL[0])
            R_VEL_Y_TAR.append(TARGET_VEL[1])
            R_POS_VEL_X_ACT.append(ACTUAL_VEL[0])
            R_POS_VEL_Y_ACT.append(ACTUAL_VEL[1])

            R_POS_VEL_X_ERROR.append(vel_error_x)
            R_POS_VEL_Y_ERROR.append(vel_error_y)
            R_POS_VEL_X_GP.append(vel_gain_p_x)
            R_POS_VEL_X_GI.append(vel_gain_i_x)
            R_POS_VEL_X_GD.append(vel_gain_d_x)
            R_POS_VEL_Y_GP.append(vel_gain_p_y)
            R_POS_VEL_Y_GI.append(vel_gain_i_y)
            R_POS_VEL_Y_GD.append(vel_gain_d_y)

            R_POS_TARGET_ROLL.append(TARGET_ROLL)
            R_POS_TARGET_PITCH.append(TARGET_PITCH)
            R_POS_ACTUAL_ROLL.append(math.degrees(uav.attitude.roll))
            R_POS_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1 / FREQ_VEL) - (time.time() - t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為定位時，跳出迴圈並暫停函式
            if not MODE == 'POS':
                # 重設暫存器
                vel_gain_d_x_register = []
                vel_gain_d_y_register = []
                vel_int_error_x = 0
                vel_int_error_y = 0
                vel_previous_error_x = 0
                vel_previous_error_y = 0
                previous_time = time.time()

                # 歸零姿態數值
                TARGET_ROLL = 0
                TARGET_PITCH = 0

                # 回報控制器狀態
                STATE_VEL = False
                reset = True
                print('<飛行狀態> 速度控制器：暫停')

        # 未成功執行跳出重設時，執行重設
        if not reset:
            # 重設暫存器
            vel_gain_d_x_register = []
            vel_gain_d_y_register = []
            vel_int_error_x = 0
            vel_int_error_y = 0
            vel_previous_error_x = 0
            vel_previous_error_y = 0
            previous_time = time.time()

            # 歸零姿態數值
            TARGET_ROLL = 0
            TARGET_PITCH = 0

            # 回報控制器狀態
            STATE_VEL = False
            reset = True
            print('<飛行狀態> 速度控制器：暫停')

        # 記錄值
        t_loop = time.time()
        R_VEL_X_TAR.append(0)
        R_VEL_Y_TAR.append(0)
        R_POS_VEL_X_ACT.append(0)
        R_POS_VEL_Y_ACT.append(0)

        R_POS_VEL_X_ERROR.append(0)
        R_POS_VEL_Y_ERROR.append(0)
        R_POS_VEL_X_GP.append(0)
        R_POS_VEL_X_GI.append(0)
        R_POS_VEL_X_GD.append(0)
        R_POS_VEL_Y_GP.append(0)
        R_POS_VEL_Y_GI.append(0)
        R_POS_VEL_Y_GD.append(0)

        R_POS_TARGET_ROLL.append(TARGET_ROLL)
        R_POS_TARGET_PITCH.append(TARGET_PITCH)
        R_POS_ACTUAL_ROLL.append(math.degrees(uav.attitude.roll))
        R_POS_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))

        # 迴圈執行間隔
        t_wait = (1 / FREQ_VEL) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

        # 當飛行模式為定位時，跳入迴圈並啟動函式
        if MODE == 'POS' and STATE_THREAD:
            # 回報控制器狀態
            STATE_VEL = True
            reset = False
            print('<飛行狀態> 速度控制器：啟動')

        # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_VEL = False
    print('<關閉程序> 速度控制器：關閉')


# =========== 循跡控制器外迴圈 ===============================================
# 說明：
# 採用PID控制器型態，每一迴圈讀取一次影像計算線條型心，並將型心值做均值綠波
# 濾波完畢後再以該值計算出型心至畫面下方中心之夾角，並以此值為控制器的誤差值
# 以該值來做PID增益處理，輸出角速度來追尋角度，此控制器可處理平行、轉角、Ｔ路口轉向
# 為不可逆型循跡控制器，參數可調最大角速度輸出、切斷區間、濾波等級、面積門檻值
# 使用Global-Frame參照，以地面標線為參照座標，預測為PD控制器(前進速度控制分離)
# 偏航與滾轉分別隸屬兩個控制器操作，注意滾轉為二階控制，需使用PID調節(可參考距離單迴圈控制版本)
def heading_outer_controller():
    global TARGET_LINE_VEL, STATE_HDG_OUT
    # PID變數設定-外迴圈
    int_error_x = 0
    previous_error_x = 0
    previous_time = time.time()

    # 均值濾波設定
    gain_d_x_register = []

    # 等待影像感測器啟動
    if not STATE_SENSOR_CAM:
        print('<啟動程序> 航向外迴圈控制器：等待影像感測器啟動')
    while not STATE_SENSOR_CAM:
        time.sleep(0.1)

    # 回報航向外迴圈控制器啟用
    STATE_HDG_OUT = True
    reset = False
    print('<啟動程序> 航向外迴圈控制器：啟動')

    # 程式迴圈
    while STATE_THREAD:
        # 定位作動迴圈(同捆執行模式，於循跡模式作動)
        while MODE == 'LINE' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 讀取回授值
            actual_x = ACTUAL_LINE_POS

            # 回授值截止區修正
            if TARGET_LINE_POS - HDG_X_DZ < actual_x < TARGET_LINE_POS + HDG_X_DZ:
                actual_x = TARGET_LINE_POS

            # 計算誤差值
            error_x = TARGET_LINE_POS - actual_x

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            delta_error_x = error_x - previous_error_x
            int_error_x += (error_x * delta_time)
            derivative_x = delta_error_x / delta_time

            # X軸誤差增益處理
            gain_p_x = error_x * HDG_X_KP
            gain_i_x = int_error_x * HDG_X_KI
            gain_d_x = derivative_x * HDG_X_KD

            # 微分項均值濾波器
            gain_d_x_register.insert(0, gain_d_x)
            if len(gain_d_x_register) > HDG_X_GD_SAMPLE_NUM_AVE:
                gain_d_x_register.pop()
            gain_d_x = np.average(gain_d_x_register)

            vx_out = gain_p_x + gain_i_x + gain_d_x

            # 輸出值到全域變數
            TARGET_LINE_VEL = vx_out

            # 傳遞值給下次迴圈使用
            previous_error_x = error_x
            previous_time = current_time

            # 記錄數值
            R_HDG_LINE_ACTUAL_X.append(ACTUAL_LINE_POS)
            R_HDG_LINE_TARGET_VX.append(vx_out)
            R_HDG_LINE_X_ERROR.append(error_x)
            R_HDG_LINE_X_GP.append(gain_p_x)
            R_HDG_LINE_X_GI.append(gain_i_x)
            R_HDG_LINE_X_GD.append(gain_d_x)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1 / FREQ_HDG_OUT) - (time.time() - t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if not MODE == 'LINE':
                # 重設暫存器
                int_error_x = 0
                previous_error_x = 0
                previous_time = time.time()

                # 均值濾波設定
                gain_d_x_register = []

                # 歸零姿態數值
                TARGET_LINE_VEL = 0

                # 回報控制器狀態
                STATE_HDG_OUT = False
                reset = True
                print('<飛行狀態> 航向外迴圈控制器：暫停')

        # 無循跡作動區間
        # 未成功執行跳出重設時，執行重設
        if not reset:
            # 重設暫存器
            int_error_x = 0
            previous_error_x = 0
            previous_time = time.time()

            # 均值濾波設定
            gain_d_x_register = []

            # 歸零姿態數值
            TARGET_LINE_VEL = 0

            # 回報控制器狀態
            STATE_HDG_OUT = False
            reset = True
            print('<飛行狀態> 航向外迴圈控制器：暫停')

        # 記錄數值
        t_loop = time.time()
        # 外迴圈
        R_HDG_LINE_ACTUAL_X.append(0)
        R_HDG_LINE_TARGET_VX.append(0)
        R_HDG_LINE_X_ERROR.append(0)
        R_HDG_LINE_X_GP.append(0)
        R_HDG_LINE_X_GI.append(0)
        R_HDG_LINE_X_GD.append(0)

        # 迴圈執行間隔
        t_wait = (1 / FREQ_HDG_OUT) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

        # 當飛行模式為循跡時，跳入迴圈並啟動函式
        if MODE == 'LINE' and STATE_THREAD:
            STATE_HDG_OUT = True
            reset = False
            print('<飛行狀態> 航向外迴圈控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_HDG_OUT = False
    print('<關閉程序> 航向外迴圈控制器：關閉')


# =========== 循跡控制器內迴圈 ===============================================
# 說明：
# 採用PID控制器型態，每一迴圈讀取一次影像計算線條型心，並將型心值做均值綠波
# 濾波完畢後再以該值計算出型心至畫面下方中心之夾角，並以此值為控制器的誤差值
# 以該值來做PID增益處理，輸出角速度來追尋角度，此控制器可處理平行、轉角、Ｔ路口轉向
# 為不可逆型循跡控制器，參數可調最大角速度輸出、切斷區間、濾波等級、面積門檻值
# 使用Global-Frame參照，以地面標線為參照座標，預測為PD控制器(前進速度控制分離)
# 偏航與滾轉分別隸屬兩個控制器操作，注意滾轉為二階控制，需使用PID調節(可參考距離單迴圈控制版本)
def heading_inner_controller():
    global TARGET_ROLL, TARGET_YAW_RATE, STATE_HDG_IN
    # PID變數設定-內迴圈
    gain_d_vx_register = []
    gain_d_hdg_register = []
    int_error_vx = 0
    int_error_hdg = 0
    previous_error_vx = 0
    previous_error_hdg = 0
    previous_time = time.time()

    # 等待影像感測器啟動
    if not STATE_HDG_OUT:
        print('<啟動程序> 航向內迴圈控制器：等待航向外迴圈控制器啟動')
    while not STATE_HDG_OUT:
        time.sleep(0.1)

    # 回報定位控制器啟用
    STATE_HDG_IN = True
    reset = False
    print('<啟動程序> 航向內迴圈控制器：啟動')

    # 程式迴圈
    while STATE_THREAD:
        # 定位作動迴圈(同捆執行模式，於循跡模式作動)
        while MODE == 'LINE' and STATE_THREAD:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 誤差增益
            actual_vx = ACTUAL_LINE_VEL * HDG_VX_KH

            # 回授值截止區修正
            if TARGET_LINE_VEL - VEL_X_DZ < actual_vx < TARGET_LINE_VEL + VEL_X_DZ:
                actual_vx = TARGET_LINE_VEL

            # 計算誤差值
            error_vx = TARGET_LINE_VEL - actual_vx
            error_hdg = TARGET_HDG - ACTUAL_HDG

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            delta_error_vx = error_vx - previous_error_vx
            int_error_vx += (error_vx * delta_time)
            derivative_vx = delta_error_vx / delta_time

            delta_error_hdg = error_hdg - previous_error_hdg
            int_error_hdg += (error_hdg * delta_time)
            derivative_hdg = delta_error_hdg / delta_time

            # X軸誤差增益處理
            gain_p_vx = error_vx * HDG_VX_KP
            gain_i_vx = int_error_vx * HDG_VX_KI
            gain_d_vx = derivative_vx * HDG_VX_KD

            # 微分項均值濾波器
            gain_d_vx_register.insert(0, gain_d_vx)
            if len(gain_d_vx_register) > HDG_VX_GD_SAMPLE_NUM_AVE:
                gain_d_vx_register.pop()
            gain_d_vx = np.average(gain_d_vx_register)

            roll_out = gain_p_vx + gain_i_vx + gain_d_vx

            # Yaw軸誤差增益處理
            gain_p_hdg = error_hdg * HDG_YAW_KP
            gain_i_hdg = int_error_hdg * HDG_YAW_KI
            gain_d_hdg = derivative_hdg * HDG_YAW_KD

            # 微分項均值濾波器
            gain_d_hdg_register.insert(0, gain_d_hdg)
            if len(gain_d_hdg_register) > HDG_YAW_GD_SAMPLE_NUM_AVE:
                gain_d_hdg_register.pop()
            gain_d_hdg = np.average(gain_d_hdg_register)

            yaw_rate_out = gain_p_hdg + gain_i_hdg + gain_d_hdg

            # 角度限制(飽和區限幅與截止區處理)
            if roll_out > HDG_ROLL_LIMIT:
                roll_out = HDG_ROLL_LIMIT
            elif roll_out < -HDG_ROLL_LIMIT:
                roll_out = -HDG_ROLL_LIMIT
            elif -HDG_ROLL_CUTOFF < roll_out < HDG_ROLL_CUTOFF:
                roll_out = 0

            if yaw_rate_out > HDG_YAW_RATE_LIMIT:
                yaw_rate_out = HDG_YAW_RATE_LIMIT
            elif yaw_rate_out < -HDG_YAW_RATE_LIMIT:
                yaw_rate_out = -HDG_YAW_RATE_LIMIT
            elif -HDG_YAW_RATE_CUTOFF < yaw_rate_out < HDG_YAW_RATE_CUTOFF:
                yaw_rate_out = 0

            # 輸出至全域變數
            TARGET_ROLL = -roll_out
            TARGET_YAW_RATE = yaw_rate_out

            # 傳遞值給下次迴圈使用
            previous_error_vx = error_vx
            previous_error_hdg = error_hdg
            previous_time = current_time

            # 記錄數值
            # 內迴圈Roll
            R_HDG_LINE_VX_TARGET.append(TARGET_LINE_VEL)
            R_HDG_LINE_VX_ACTUAL.append(ACTUAL_LINE_VEL)
            R_HDG_LINE_VX_TARGET_ROLL.append(TARGET_ROLL)
            R_HDG_LINE_VX_ACTUAL_ROLL.append(math.degrees(uav.attitude.roll))
            R_HDG_LINE_VX_ERROR.append(error_vx)
            R_HDG_LINE_VX_GP.append(gain_p_vx)
            R_HDG_LINE_VX_GI.append(gain_i_vx)
            R_HDG_LINE_VX_GD.append(gain_d_vx)

            # 內迴圈Yaw
            R_HDG_YAW_ACTUAL.append(ACTUAL_HDG)
            R_HDG_YAW_TARGET_YAW_RATE.append(TARGET_YAW_RATE)
            R_HDG_YAW_ERROR.append(error_hdg)
            R_HDG_YAW_GP.append(gain_p_hdg)
            R_HDG_YAW_GI.append(gain_i_hdg)
            R_HDG_YAW_GD.append(gain_d_hdg)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1 / FREQ_HDG_IN) - (time.time() - t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if not MODE == 'LINE':
                # 重設暫存器
                gain_d_vx_register = []
                gain_d_hdg_register = []
                int_error_vx = 0
                int_error_hdg = 0
                previous_error_vx = 0
                previous_error_hdg = 0
                previous_time = time.time()

                # 歸零姿態數值
                TARGET_ROLL = 0
                TARGET_YAW_RATE = 0

                # 回報控制器狀態
                STATE_HDG_IN = False
                reset = True
                print('<飛行狀態> 航向內迴圈控制器：暫停')

        # 無循跡作動區間
        # 未成功執行跳出重設時，執行重設
        if not reset:
            # 重設暫存器
            gain_d_vx_register = []
            gain_d_hdg_register = []
            int_error_vx = 0
            int_error_hdg = 0
            previous_error_vx = 0
            previous_error_hdg = 0
            previous_time = time.time()

            # 歸零姿態數值
            TARGET_ROLL = 0
            TARGET_YAW_RATE = 0

            # 回報控制器狀態
            STATE_HDG_IN = False
            reset = True
            print('<飛行狀態> 航向內迴圈控制器：暫停')

        # 記錄數值
        t_loop = time.time()
        # 內迴圈Roll
        R_HDG_LINE_VX_TARGET.append(0)
        R_HDG_LINE_VX_ACTUAL.append(0)
        R_HDG_LINE_VX_TARGET_ROLL.append(0)
        R_HDG_LINE_VX_ACTUAL_ROLL.append(0)
        R_HDG_LINE_VX_ERROR.append(0)
        R_HDG_LINE_VX_GP.append(0)
        R_HDG_LINE_VX_GI.append(0)
        R_HDG_LINE_VX_GD.append(0)

        # 內迴圈Yaw
        R_HDG_YAW_ACTUAL.append(0)
        R_HDG_YAW_TARGET_YAW_RATE.append(0)
        R_HDG_YAW_ERROR.append(0)
        R_HDG_YAW_GP.append(0)
        R_HDG_YAW_GI.append(0)
        R_HDG_YAW_GD.append(0)

        # 迴圈執行間隔
        t_wait = (1 / FREQ_HDG_IN) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

        # 當飛行模式為循跡時，跳入迴圈並啟動函式
        if MODE == 'LINE' and STATE_THREAD:
            STATE_HDG_IN = True
            reset = False
            print('<飛行狀態> 航向內迴圈控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_HDG_IN = False
    print('<關閉程序> 航向內迴圈控制器：關閉')


# =========== 飛行速度控制器 ============================================
# 說明：
# 於循線期間控制往前飛行的速度
# 不一定會使用(視感測器響應情形而定)
def flight_speed_controller():
    global TARGET_PITCH, STATE_SPD
    # PID變數設定
    int_error = 0
    previous_error = 0
    previous_time = time.time()

    # 均值濾波器定
    gain_d_register = []

    # 等待飛行速度感測器啟動
    if not STATE_SENSOR_SPD:
        print('<啟動程序> 飛行速度控制器：等待飛行速度感測器啟動\n')
    while not STATE_SENSOR_SPD:
        time.sleep(0.1)

    # 回報飛行速度控制器啟用
    STATE_SPD = True
    reset = False
    print('<啟動程序> 飛行速度控制器：啟動，控制模式：%d' % SPD_MODE)

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 閉迴路控制
        while MODE == 'LINE' and STATE_THREAD and SPD_MODE:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 計算誤差值
            error = TARGET_SPD - ACTUAL_SPD

            # 誤差值微積分計算
            current_time = time.time()
            delta_time = current_time - previous_time

            delta_error = error - previous_error
            int_error += (error * delta_time)
            derivative = delta_error / delta_time

            # 誤差增益處理
            gain_p = error * SPD_KP
            gain_i = int_error * SPD_KI
            gain_d = derivative * SPD_KD

            # 微分項均值濾波器
            gain_d_register.insert(0, gain_d)
            if len(gain_d_register) > SPD_GD_SAMPLE_NUM_AVE:
                gain_d_register.pop()
            gain_d = np.average(gain_d_register)

            pitch_out = gain_p + gain_i + gain_d

            # 角度限制(包含截止區設定)
            if pitch_out > SPD_ANGLE_LIMIT:
                pitch_out = SPD_ANGLE_LIMIT
            elif pitch_out < -SPD_ANGLE_LIMIT:
                pitch_out = -SPD_ANGLE_LIMIT
            elif -SPD_ANGLE_CUTOFF < pitch_out < SPD_ANGLE_CUTOFF:
                pitch_out = 0

            # 輸出至全域變數
            TARGET_PITCH = -pitch_out

            # 傳遞值給下次迴圈使用
            previous_error = error
            previous_time = current_time

            # 記錄數值
            R_SPD_TARGET.append(TARGET_SPD)
            R_SPD_ACTUAL.append(ACTUAL_SPD)
            R_SPD_TARGET_PITCH.append(TARGET_PITCH)
            R_SPD_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))

            R_SPD_ERROR.append(error)
            R_SPD_GP.append(gain_p)
            R_SPD_GI.append(gain_i)
            R_SPD_GD.append(gain_d)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1 / FREQ_SPD) - (time.time() - t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if not MODE == 'LINE':
                # 重設暫存器
                gain_d_register = []
                int_error = 0
                previous_error = 0
                previous_time = time.time()

                # 切換跳出姿態數值(慣性抵銷用)
                TARGET_PITCH = SPD_OFF_ANGLE

                # 回報控制器狀態
                STATE_SPD = False
                reset = True
                print('<飛行狀態> 飛行速度控制器：暫停')

        # 開迴路控制
        while MODE == 'LINE' and STATE_THREAD and not SPD_MODE:
            # 作動輸出區間
            t_on = time.time()
            if SPD_OL_ACTIVE_TIME != 0:
                TARGET_PITCH = -SPD_OL_ANGLE
            while time.time() - t_on < SPD_OL_ACTIVE_TIME:
                t_loop = time.time()
                R_SPD_OL_ACTUAL_SPD.append(ACTUAL_SPD)
                R_SPD_OL_TARGET_PITCH.append(TARGET_PITCH)
                R_SPD_OL_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))
                t_wait = (1 / FREQ_SPD) - (time.time() - t_loop)
                if t_wait > 0:
                    time.sleep(t_wait)

            # 作動關閉區間
            t_off = time.time()
            if SPD_OL_OFF_TIME != 0:
                TARGET_PITCH = 0
            while time.time() - t_off < SPD_OL_OFF_TIME:
                t_loop = time.time()
                R_SPD_OL_ACTUAL_SPD.append(ACTUAL_SPD)
                R_SPD_OL_TARGET_PITCH.append(TARGET_PITCH)
                R_SPD_OL_ACTUAL_PITCH.append(math.degrees(uav.attitude.pitch))
                t_wait = (1 / FREQ_SPD) - (time.time() - t_loop)
                if t_wait > 0:
                    time.sleep(t_wait)

            # 當飛行模式不為循跡時，跳出迴圈並暫停函式
            if not MODE == 'LINE':
                # 切換跳出姿態數值(慣性抵銷用)
                TARGET_PITCH = SPD_OFF_ANGLE
                STATE_SPD = False
                reset = True
                print('<飛行狀態> 飛行速度控制器：暫停')

        # 無循跡作動區間
        t_loop = time.time()
        # 未成功執行跳出重設時，執行重設
        if not reset:
            # 重設暫存器
            gain_d_register = []
            int_error = 0
            previous_error = 0
            previous_time = time.time()

            # 切換跳出姿態數值
            TARGET_PITCH = 0

            # 回報控制器狀態
            STATE_SPD = False
            reset = True
            print('<飛行狀態> 飛行速度控制器：暫停')

        # 記錄數值
        if SPD_MODE:
            R_SPD_TARGET.append(0)
            R_SPD_ACTUAL.append(0)
            R_SPD_TARGET_PITCH.append(0)
            R_SPD_ACTUAL_PITCH.append(0)

            R_SPD_ERROR.append(0)
            R_SPD_GP.append(0)
            R_SPD_GI.append(0)
            R_SPD_GD.append(0)

        elif not SPD_MODE:
            R_SPD_OL_ACTUAL_SPD.append(0)
            R_SPD_OL_TARGET_PITCH.append(0)
            R_SPD_OL_ACTUAL_PITCH.append(0)

        # 迴圈執行間隔
        t_wait = (1 / FREQ_SPD) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

        # 當飛行模式為循跡時，跳入迴圈並啟動函式
        if MODE == 'LINE' and STATE_THREAD:
            print('<飛行狀態> 飛行速度控制器：啟動')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_SPD = False
    print('<關閉程序> 飛行速度控制器：關閉')


# ========== 影像擷取程式 ==============================================
# 負責截取機載影像畫面，儲存至全域值供控制器使用
# 因主線程衝突故將畫面顯示整合至GUI介面中方便瀏覽
# 另將飛行影像存成影片以供記錄(處理後之影像若需錄下，須經格式轉換處理，暫不增加)
def camera():
    global CAM_IMAGE, STATE_CAM, STATE_THREAD
    # 程式迴圈(同捆執行模式)
    while CAM_VIDEO.isOpened() and STATE_THREAD:
        # 確認影像擷取狀態
        state_grab = CAM_VIDEO.grab()

        # 擷取成功
        if state_grab:
            # 同捆執行起點
            THREAD_OCCUPY.acquire()

            # 擷取影像並儲存
            state_retrieve, image = CAM_VIDEO.retrieve()
            CAM_IMAGE = cv2.resize(image, (CAM_IMAGE_WIDTH, CAM_IMAGE_HEIGHT))
            VIDEO_SAVE.write(CAM_IMAGE)

            # 回報影像擷取程式啟動
            if not STATE_CAM:
                STATE_CAM = True
                print('<啟動程序> 影像擷取程式：啟動\n')

            # 顯示影像視窗
            if MODE == 'POS' or MODE == 'LINE':
                cv2.imshow("Flight Camera", CAM_IMAGE_LIVE)
            else:
                cv2.imshow("Flight Camera", CAM_IMAGE)

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 手動關閉影像顯示
            if cv2.waitKey(int(1000 / FREQ_CAM)) & 0xff == ord("q"):
                STATE_THREAD = False
                print('<關閉程序> 已手動關閉相機，請手動重啟程式')
                break

        # 擷取失敗
        else:
            print('<錯誤> 影像擷取程式：無法擷取影像')
            STATE_THREAD = False
            print('<嚴重錯誤> 請檢查相機並手動重啟程式!!!!')
            break

    # 當系統狀態為否時，跳出迴圈並關閉函式
    CAM_VIDEO.release()
    VIDEO_SAVE.release()
    cv2.destroyAllWindows()  # 關閉影像顯示
    STATE_CAM = False
    print('<關閉程序> 影像擷取程式：關閉')


# ========== 影像感測器 ===============================================
# 負責計算影像的位置、速度、角度值，結合位置與循線模式
def camera_sensor():
    global STATE_SENSOR_CAM, ACTUAL_POS, ACTUAL_VEL, ACTUAL_AREA
    global ACTUAL_HDG, ACTUAL_LINE_POS, ACTUAL_LINE_VEL, ACTUAL_LINE_AREA
    # 均值濾波器設定
    pos_x_register = []
    pos_y_register = []
    vel_x_register = []
    vel_y_register = []
    line_angle_register = []

    # 速度值計算初始設定
    pos_x_old = 0
    pos_y_old = 0
    vel_x = 0
    vel_y = 0
    previous_time = time.time()

    # 等待影像擷取程式啟動
    if not STATE_CAM:
        print('<啟動程序> 影像感測器：等待影像擷取程式啟動')
    while not STATE_CAM:
        time.sleep(0.1)

    # 回報影像感測器啟用
    STATE_SENSOR_CAM = True
    _first_loop = True
    print('<啟動程序> 影像感測器：啟動')

    # 程式迴圈
    while STATE_THREAD:
        # 定位模式計算迴圈(同捆執行模式)
        _first_loop = True
        while MODE == 'POS' and STATE_THREAD:
            # 確認已經完成重設
            if _first_loop:
                pos_x_register = []
                pos_y_register = []
                vel_x_register = []
                vel_y_register = []
                line_angle_register = []
                _first_loop = False

            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 取得色塊XY座標
            x_raw, y_raw, block_area = image_moment(POS_COLOR)

            # 位置出界狀態值保留
            if x_raw == 0 and math.fabs(pos_x_old) > CAM_IMAGE_WIDTH*0.1:
                x_raw = pos_x_old
                outbound_x = True
            else:
                outbound_x = False

            if y_raw == 0 and math.fabs(pos_y_old) > CAM_IMAGE_HEIGHT*0.1:
                y_raw = pos_y_old
                outbound_y = True
            else:
                outbound_y = False

            # X座標均值濾波器
            pos_x_register.insert(0, x_raw)
            if len(pos_x_register) > IMG_SAMPLE_NUM_AVE:
                pos_x_register.pop()
            pos_x = np.average(pos_x_register)

            # Y座標均值濾波器
            pos_y_register.insert(0, y_raw)
            if len(pos_y_register) > IMG_SAMPLE_NUM_AVE:
                pos_y_register.pop()
            pos_y = np.average(pos_y_register)

            # 計算型心移動速度
            current_time = time.time()
            delta_time = current_time - previous_time

            pos_delta_x = pos_x - pos_x_old
            pos_delta_y = pos_y - pos_y_old
            vel_x_raw = pos_delta_x / delta_time
            vel_y_raw = pos_delta_y / delta_time

            # 速度出界狀態值保留
            if outbound_x:
                vel_x_raw = vel_x
            if outbound_y:
                vel_y_raw = vel_y

            # X速度均值濾波器
            vel_x_register.insert(0, vel_x_raw)
            if len(vel_x_register) > IMG_SAMPLE_NUM_AVE:
                vel_x_register.pop()
            vel_x = np.average(vel_x_register)

            # Y速度均值濾波器
            vel_y_register.insert(0, vel_y_raw)
            if len(vel_y_register) > IMG_SAMPLE_NUM_AVE:
                vel_y_register.pop()
            vel_y = np.average(vel_y_register)

            # 輸出數值至全域變數
            ACTUAL_POS = [pos_x, pos_y]
            ACTUAL_VEL = [vel_x, vel_y]
            ACTUAL_AREA = block_area

            # 傳遞值給下次迴圈使用
            pos_x_old = pos_x
            pos_y_old = pos_y
            previous_time = current_time

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1 / FREQ_SENSOR_CAM) - (time.time() - t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為定位時，跳出迴圈
            if MODE == 'LINE':
                # 歸零數值
                ACTUAL_POS = [0, 0]
                ACTUAL_VEL = [0, 0]
                ACTUAL_AREA = 0
                _first_loop = True
                print('<飛行狀態> 影像感測器：切換至角度輸出')

        # 循跡模式計算迴圈(同捆執行模式)
        _first_loop = True
        while MODE == 'LINE' and STATE_THREAD:
            # 確認已經完成重設
            if _first_loop:
                pos_x_register = []
                pos_y_register = []
                vel_x_register = []
                vel_y_register = []
                line_angle_register = []
                _first_loop = False

            # 同捆執行起點
            THREAD_OCCUPY.acquire()
            t_loop = time.time()

            # 取得線條型心位置
            pos_x_raw, pos_y_raw, area_sum, line_angle_raw = line_moment()

            # 型心位置均值濾波器
            # X座標均值濾波器
            pos_x_register.insert(0, pos_x_raw)
            if len(pos_x_register) > ANG_SAMPLE_NUM_AVE:
                pos_x_register.pop()
            pos_x = np.average(pos_x_register)

            # 計算型心移動速度
            current_time = time.time()
            delta_time = current_time - previous_time

            pos_delta_x = pos_x - pos_x_old
            vel_x_raw = pos_delta_x / delta_time

            # X速度均值濾波器
            vel_x_register.insert(0, vel_x_raw)
            if len(vel_x_register) > IMG_SAMPLE_NUM_AVE:
                vel_x_register.pop()
            vel_x = np.average(vel_x_register)

            # 線角均值濾波器
            line_angle_register.insert(0, line_angle_raw)
            if len(line_angle_register) > ANG_SAMPLE_NUM_AVE:
                line_angle_register.pop()
            line_angle = np.average(line_angle_register)

            # 輸出值到全域變數
            ACTUAL_HDG = line_angle
            ACTUAL_LINE_POS = pos_x
            ACTUAL_LINE_VEL = vel_x
            ACTUAL_LINE_AREA = area_sum

            # 同捆執行終點
            THREAD_OCCUPY.release()

            # 迴圈執行間隔
            t_wait = (1 / FREQ_SENSOR_CAM) - (time.time() - t_loop)
            if t_wait > 0:
                time.sleep(t_wait)

            # 當飛行模式不為循跡時，跳出迴圈
            if MODE == 'POS':
                # 歸零數值
                ACTUAL_HDG = 0
                ACTUAL_LINE_POS = 0
                ACTUAL_LINE_VEL = 0
                _first_loop = True
                print('<飛行狀態> 影像感測器：切換至位置-速度輸出')

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_CAM = False
    print('<關閉程序> 影像感測器：關閉')


# =========== 高度感測器整合程式 =========================================
# 說明：
# 負責整合超音波高度計、飛控氣壓計(可能為IMU來源)，做訊號整理給控制器使用
# 各訊號先經均值濾波後，再用卡爾曼濾波處理，最後再針對所有來源做加權平均
# 本函式處理氣壓計之濾波、高度感測器整合濾波
def altitude_sensor():
    global ACTUAL_ALT, ACTUAL_ALT_IMU, STATE_SENSOR_ALT
    # 平均降噪設定
    height_register_imu = []

    # 卡爾曼濾波器初始化(IMU高度計用)
    x_o = 0                     # 前一次狀態
    p_o = 0                     # 前一次斜方差
    z_o = 0                     # 前一次量測值
    q = math.exp(-IMU_KF_Q)     # 斜方差噪音值
    r = 2.92 * math.exp(-3)     # 斜方差量測噪音值(定值不更改)

    # 執行IMU高度計校正程序
    imu_calibrate()

    # 等待超音波高度計啟動
    if not STATE_SENSOR_SONAR:
        print('<啟動程序> 高度感測器：等待超音波高度計啟動\n')
    while not STATE_SENSOR_SONAR:
        time.sleep(0.1)

    # 回報高度感測器啟動
    STATE_SENSOR_ALT = True
    print('<啟動程序> 高度感測器：啟動\n')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 同捆執行起點
        THREAD_OCCUPY.acquire()
        t_loop = time.time()

        # 讀取超音波高度計資料
        height_sonar = ACTUAL_ALT_SONAR

        # IMU高度計數據處理
        # 卡爾曼濾波-預測
        x_p = x_o
        p_p = p_o + q

        # IMU讀值與補正
        height_raw_imu = uav.location.global_relative_frame.alt + IMU_OFFSET

        # IMU均值濾波器
        height_register_imu.insert(0, height_raw_imu)
        if len(height_register_imu) > IMU_SAMPLE_NUM_AVE:
            height_register_imu.pop()
        height_imu_ave = np.average(height_register_imu)

        # 卡爾曼濾波器輸入限制(防止上衝突波)
        if height_imu_ave <= IMU_KF_INPUT_MAX:
            z = height_imu_ave  # 若低於限制則使用新值
        else:
            z = z_o  # 若超出限制則使用舊值

        # 卡爾曼濾波-更新
        k = p_p / (p_p + r)  # 卡爾曼增益
        x = x_p + k * (z - x_p)  # 狀態(濾波輸出值)
        p = (1 - k) * p_p  # 斜方差

        # 更新高度值至區域變數
        height_imu = x

        # 卡爾曼濾波：傳遞值給下次迴圈使用
        x_o = x
        p_o = p
        z_o = z

        # 更新高度值至全域變數
        ACTUAL_ALT_IMU = height_imu

        # 加權平均(可調整加權值)
        alt_num = (WEIGHTED_SONAR * height_sonar + WEIGHTED_IMU * height_imu)
        alt_den = (WEIGHTED_SONAR + WEIGHTED_IMU)
        alt_ave = alt_num / alt_den

        # 更新高度值至全域變數
        ACTUAL_ALT = alt_ave

        # 記錄數值
        R_ACTUAL_ALT_SONAR.append(height_sonar)         # 記錄超音波高度計值
        R_ACTUAL_ALT_IMU_RAW.append(uav.location.global_relative_frame.alt)     # 記錄氣壓高度計原始值
        R_ACTUAL_ALT_IMU_OFFSET.append(height_raw_imu)  # 記錄補正值
        R_ACTUAL_ALT_IMU_AVE.append(height_imu_ave)     # 記錄均值濾波值
        R_ACTUAL_ALT_IMU_KF.append(height_imu)          # 記錄KF值
        R_ACTUAL_ALT.append(ACTUAL_ALT)                 # 記錄總高度值

        # 同捆執行終點
        THREAD_OCCUPY.release()

        # 等待下次迴圈
        t_wait = (1 / FREQ_SENSOR_ALT) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_ALT = False
    print('<關閉程序> 高度感測器：關閉\n')


# ========== 超音波高度計 ==============================================
# 說明：
# 使用HC-SR04為超音波感測模組，發出一脈波訊號使模組執行量測後傳回經過時間長的脈波
# 實際使用上雜訊相對多(可能受飛行時噪音影響)，故施加多重濾波條件：
# 超時防止(防止過長的無效等待)、角度補正、突波保護、均值濾波、卡爾曼濾波
# 經處理後的值堪用度大幅提高，對吸音面也能有效處理，使高度控制器能穩定運作
# 於過多線程執行中因線程排序問題需做應對處置
def altitude_sensor_sonar():
    global ACTUAL_ALT_SONAR, STATE_SENSOR_SONAR, SONAR_ERR_COUNT
    # 初始化腳位
    GPIO.setmode(GPIO.BCM)                  # 設定為BCM模式
    GPIO.setwarnings(False)                 # 忽略警告
    GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)    # TRIG腳位輸出設定
    GPIO.setup(SONAR_ECHO_PIN, GPIO.IN)     # ECHO腳位輸入設定
    GPIO.output(SONAR_TRIG_PIN, False)      # 預設輸出為0

    # 脈波變化時間值變數
    pulse_start = time.time()               # 脈波起始時間
    pulse_end = time.time()                 # 脈波結束時間

    # 卡爾曼濾波器初始化
    x_o = 0                                 # 前一次狀態
    p_o = 0                                 # 前一次斜方差
    z_o = 0                                 # 前一次量測值
    q = math.exp(-SONAR_KF_Q)               # 斜方差噪音值
    r = 2.92 * math.exp(-3)                 # 斜方差量測噪音值(定值不更改)

    # 平均降噪設定
    height_register = []
    height_ave = 0

    # 執行超音波高度計校正程序
    sonar_calibrate()

    # 回報超音波高度計啟動
    STATE_SENSOR_SONAR = True
    print('<啟動程序> 超音波高度計：啟動\n')

    # 程式迴圈
    while STATE_THREAD:
        # 線程優先執行起點
        THREAD_OCCUPY.acquire()
        t_loop = time.time()

        # 重設超時旗標
        timeout = True

        # 卡爾曼濾波-預測
        x_p = x_o
        p_p = p_o + q

        # 腳位初始化(因有報錯記錄，於每次前重新初始化)
        GPIO.setmode(GPIO.BCM)                  # 設定為BCM模式
        GPIO.setwarnings(False)                 # 忽略警告
        GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)    # TRIG腳位輸出設定
        GPIO.setup(SONAR_ECHO_PIN, GPIO.IN)     # ECHO腳位輸入設定
        GPIO.output(SONAR_TRIG_PIN, False)      # 預設輸出為0

        # 輸出啟用脈波
        GPIO.output(SONAR_TRIG_PIN, True)
        time.sleep(0.00001)                     # 發出10us脈波
        GPIO.output(SONAR_TRIG_PIN, False)

        # 記錄超時基準時間
        timeout_start = time.time()

        # 接收計時脈波
        # 於低電位迴圈等待，跳出時記錄脈波起始時間
        while not GPIO.input(SONAR_ECHO_PIN):
            pulse_start = time.time()           # 記錄脈波起始時間
            timeout = False                     # 成功記錄脈波時間，切換旗標
            if pulse_start - timeout_start > SONAR_TIMEOUT:
                timeout = True                  # 等待接收時發生超時，跳出迴圈
                break

        # 於高電位迴圈等待，跳出時記錄脈波結束時間(若發生超時則不執行本迴圈)
        while GPIO.input(SONAR_ECHO_PIN) and not timeout:
            pulse_end = time.time()             # 記錄脈波結束時間
            if pulse_end - pulse_start > SONAR_TIMEOUT:
                timeout = True                  # 接收時長發生超時，跳出迴圈
                break

        # 若發生超時則顯示於訊息(可停用)
        if timeout:
            SONAR_ERR_COUNT += 1        # 超時次數計(統計用)
            # print('<錯誤> 超音波高度計：量測超時')

        if not timeout:
            # 計算高度值
            pulse_duration = pulse_end - pulse_start
            height_raw = (pulse_duration * (331.3 + 0.606 * SONAR_AIR_TEMP)) / 2

            # 高度角度補正
            roll_angle = uav.attitude.roll
            pitch_angle = uav.attitude.pitch
            height_offset = (height_raw * math.cos(roll_angle) * math.cos(pitch_angle)) + SONAR_OFFSET
            height_offset_r = height_offset  # 記錄用值

            # 上下降突波防止
            if ACTUAL_ALT > SONAR_FILTER_ALT:
                if math.fabs(height_offset - z_o) > SONAR_MAX_CHANGE:
                    height_offset = height_ave

            # 均值濾波器(使線條滑順)
            height_register.insert(0, height_offset)
            if len(height_register) > SONAR_SAMPLE_NUM_AVE:
                height_register.pop()
            height_ave = np.average(height_register)  # 依據取樣次數計算即時平均值(會產生響應延遲)

            # 卡爾曼濾波器輸入限制(防止突波)
            if height_ave <= SONAR_KF_INPUT_MAX:
                z = height_ave  # 若低於限制則使用新值
            else:
                z = z_o  # 若超出限制則使用舊值

            # 卡爾曼濾波-更新
            k = p_p / (p_p + r)  # 卡爾曼增益
            x = x_p + k * (z - x_p)  # 狀態(濾波輸出值)
            p = (1 - k) * p_p  # 斜方差

            # 更新高度值至全域變數
            ACTUAL_ALT_SONAR = x

            # 傳遞值給下次迴圈使用
            x_o = x
            p_o = p
            z_o = z

            # 記錄數值
            R_ACTUAL_ALT_SONAR_RAW.append(height_raw)  # 記錄原始值
            R_ACTUAL_ALT_SONAR_OFFSET.append(height_offset_r)  # 記錄補正值(突波防止前)
            R_ACTUAL_ALT_SONAR_AVE.append(height_ave)  # 記錄平均值
            R_ACTUAL_ALT_SONAR_KF.append(ACTUAL_ALT_SONAR)  # 記錄KF值

        # 結束優先執行
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        t_wait = (1 / FREQ_SENSOR_SONAR) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_SONAR = False
    print('<關閉程序> 超音波高度計：關閉；量測超時次數共 %d 次\n' % SONAR_ERR_COUNT)


# =========== 飛行速度感測器 ============================================
# 說明：
# 計算循跡時的飛行速度，利用IMU的速度值計算合速度
def speed_sensor():
    global ACTUAL_SPD, STATE_SENSOR_SPD
    # 均值濾波器設定
    speed_register = []

    # 等待無人機連線
    if not STATE_UAV:
        print('<啟動程序> 飛行速度感測器：等待無人機連線')
    while not STATE_UAV:
        time.sleep(0.1)

    # 回報飛行速度控制器啟用
    STATE_SENSOR_SPD = True
    print('<啟動程序> 飛行速度感測器：啟動')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 線程優先執行起點
        THREAD_OCCUPY.acquire()
        t_loop = time.time()

        # 讀取速度值
        vx = uav.velocity[0]
        vy = uav.velocity[1]

        # 計算和速度
        if MODE == 'POS':
            speed_raw = 0
        else:
            speed_raw = math.sqrt(math.pow(vx, 2) + math.pow(vy, 2))

        # 均值濾波處理
        speed_register.insert(0, speed_raw)
        if len(speed_register) > SPD_SAMPLE_NUM_AVE:
            speed_register.pop()
        speed = np.average(speed_register)

        # 方向判斷
        hdg = uav.heading
        if vx > 0 and (hdg < 90 or hdg > 270):
            speed = speed
        elif vx > 0 and (90 < hdg < 270):
            speed = -speed
        elif vx < 0 and (hdg < 90 or hdg > 270):
            speed = -speed
        elif vx < 0 and (90 < hdg < 270):
            speed = speed
        elif vx == 0:
            if vy > 0 and hdg == 90:
                speed = speed
            elif vy > 0 and hdg == 270:
                speed = -speed
            elif vy < 0 and hdg == 270:
                speed = speed
            elif vy < 0 and hdg == 90:
                speed = -speed

        # 輸出至全域變數
        ACTUAL_SPD = speed

        # 結束優先執行
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        t_wait = (1 / FREQ_SENSOR_SPD) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

    # 當系統狀態為否時，跳出迴圈並關閉函式
    STATE_SENSOR_SPD = False
    print('<關閉程序> 飛行速度感測器：關閉')


# ========== 命令傳輸程式 ==============================================
# 根據Ardupilot與MavLink協定規範，僅能使用set_attitude_target命令
# 另DroneKit套件提供函式封裝上也是依據上述規範編碼，故所使用的命令包括：
# 姿態四元數、各軸角速度、推力(昇降率)、操作模式(角度、角速度)
# 姿態四元數將控制器的角度值傳至函式做計算後返還、角速度採徑度制
# 根據測試及參考習慣操作，故選用姿態控制但於Yaw軸採角速度控制，符合手控習慣
# 流程上為將各控制器所輸出的命令值做統一編碼封裝後傳送
def command_transfer():
    global STATE_CMD

    # 回報命令傳輸程式啟動
    STATE_CMD = True
    print('<啟動程序> 命令傳輸程式：啟動\n')

    # 程式迴圈(同捆執行模式)
    while STATE_THREAD:
        # 同捆執行起點
        THREAD_OCCUPY.acquire()
        t_loop = time.time()

        # 命令編碼，使用偏航角速度控制(餘採姿態控制)
        msg = uav.message_factory.set_attitude_target_encode(
            0,                  # 開機時間
            0,                  # 目標系統
            0,                  # 目標裝置
            0b00000011,         # 命令遮罩(1忽略，0啟用)
            to_quaternion(TARGET_ROLL, TARGET_PITCH, TARGET_YAW),  # 姿態四元數
            0,                  # Roll滾轉角速度(radian/s)
            0,                  # Pitch滾轉角速度(radian/s)
            math.radians(TARGET_YAW_RATE),          # Yaw滾轉角速度(radian/s)
            TARGET_THRUST)      # 推力升降比(0~1，中點0.5)

        # 傳輸命令
        uav.send_mavlink(msg)

        # 同捆執行終點
        THREAD_OCCUPY.release()

        # 迴圈執行間隔
        t_wait = (1 / FREQ_CMD) - (time.time() - t_loop)
        if t_wait > 0:
            time.sleep(t_wait)

    # 當多線程狀態為否時，跳出迴圈並關閉函式
    STATE_CMD = False
    print('<關閉程序> 命令傳輸程式：關閉')


# ====================================================================
# ========== 子函式區塊 ================================================
# ====================================================================
# 本區放置各主函式下會使用到的功能函式，分有校正類、計算類、裝備控制類
# 將常使用的重複功能函式化以節省程式大小(僅限無傳遞值類的函式)


# ========== 超音波高度計校正程序 ========================================
# 於啟動前自動校正高度，為避免單次量測不準確，故以多次取樣後平均為基準來補正
def sonar_calibrate():
    """超音波校正，設定高度偏移量"""
    global SONAR_OFFSET
    # 平均降噪設定
    height_register = []

    # 脈波變化時間值變數
    pulse_start = time.time()           # 脈波起始時間
    pulse_end = time.time()             # 脈波結束時間

    # 量測迴圈(線程優先執行)
    THREAD_OCCUPY.acquire()
    for i in range(0, SONAR_SAMPLE_NUM_CAL):
        # 設定超時旗標
        timeout = True

        # 量測迴圈(超時防止)
        while timeout:
            # 腳位初始化(因有報錯記錄，於每次前重新初始化)
            GPIO.setmode(GPIO.BCM)                  # 設定為BCM模式
            GPIO.setwarnings(False)                 # 忽略警告
            GPIO.setup(SONAR_TRIG_PIN, GPIO.OUT)  # TRIG腳位輸出設定
            GPIO.setup(SONAR_ECHO_PIN, GPIO.IN)     # ECHO腳位輸入設定
            GPIO.output(SONAR_TRIG_PIN, False)

            # 輸出啟用脈波
            GPIO.output(SONAR_TRIG_PIN, True)
            time.sleep(0.000025)                     # 發出10us脈波
            GPIO.output(SONAR_TRIG_PIN, False)
            timeout_start = time.time()             # 記錄超時基準時間

            # 接收計時脈波
            # 於低電位迴圈等待，跳出時記錄脈波起始時間
            while not GPIO.input(SONAR_ECHO_PIN):
                pulse_start = time.time()           # 記錄脈波起始時間
                timeout = False                     # 成功記錄脈波時間，切換旗標
                if pulse_start - timeout_start > SONAR_TIMEOUT:
                    timeout = True                  # 等待接收時發生超時，跳出迴圈
                    break

            # 於高電位迴圈等待，跳出時記錄脈波結束時間(若發生超時則不執行本迴圈)
            while GPIO.input(SONAR_ECHO_PIN) and not timeout:
                pulse_end = time.time()             # 記錄脈波結束時間
                timeout = False                     # 成功記錄脈波時間，切換旗標
                if pulse_end - pulse_start > SONAR_TIMEOUT:
                    timeout = True                  # 接收時長發生超時，跳出迴圈
                    break

            # 若發生超時則顯示於訊息
            if timeout:
                # print('<錯誤> 超音波感測器：量測超時')
                pass

        # 計算高度值
        pulse_duration = pulse_end - pulse_start
        height = (pulse_duration * (331.3 + 0.606 * SONAR_AIR_TEMP)) / 2

        # 均值濾波器
        height_register.append(height)

        # 等待10ms執行下次量測
        time.sleep(0.01)

    # 結束優先執行
    THREAD_OCCUPY.release()

    # 更新高度校正偏移量
    SONAR_OFFSET = -np.average(height_register)
    print('<啟動程序> 超音波高度計校正完畢，偏移 %.3f m\n' % SONAR_OFFSET)


# ========== IMU高度計校正程序 =========================================
# 因IMU高度計會有波動，需做補正
def imu_calibrate():
    """IMU高度計校正，設定高度偏移量"""
    global IMU_OFFSET
    # 平均降噪設定
    height_register = []

    # 校正氣壓計
    uav.send_calibrate_barometer()
    uav.send_calibrate_vehicle_level()
    time.sleep(1)

    # 量測迴圈(線程優先執行)
    THREAD_OCCUPY.acquire()
    for i in range(0, IMU_SAMPLE_NUM_CAL):
        # 讀取氣壓計高度值
        height = uav.location.global_relative_frame.alt

        # 均值濾波器
        height_register.append(height)

        # 等待200ms執行下次量測
        time.sleep(0.2)

    # 結束優先執行
    THREAD_OCCUPY.release()

    # 更新高度校正偏移量
    IMU_OFFSET = -np.average(height_register)
    print('<啟動程序> 氣壓高度計校正完畢，偏移 %.3f m\n' % IMU_OFFSET)


# ========== 色塊型心計算程序 ==========================================
# 為計算當禎影像中，目標顏色的型心位置、色塊面積大小，主要供位置控制器、飛行條件判斷使用
# 為防止雜訊可設置最低面積值，因此若色塊大小在此之下將無法被偵測到
# 處理流程為將影像先依目標顏色做過濾，再依此做型心、面積運算
def image_moment(color):
    """計算指定顏色型心位置、色塊大小，支援紅、綠、藍、黑色"""
    global CAM_IMAGE_LIVE
    # 色域轉換BGR->HSV
    image_hsv = cv2.cvtColor(CAM_IMAGE, cv2.COLOR_BGR2HSV)
    # 閉運算
    image_close = cv2.morphologyEx(image_hsv, cv2.MORPH_CLOSE,
                                   np.ones(IMG_MORPHOLOGY_EX_KERNEL, np.uint8))

    # 依照目標色抓該顏色輪廓
    if color == 'red':          # 紅色
        image_mask = cv2.inRange(image_close, IMG_RED_L, IMG_RED_H)
    elif color == 'green':      # 綠色
        image_mask = cv2.inRange(image_close, IMG_GREEN_L, IMG_GREEN_H)
    elif color == 'blue':       # 藍色
        image_mask = cv2.inRange(image_close, IMG_BLUE_L, IMG_BLUE_H)
    elif color == 'black':      # 黑色
        image_mask = cv2.inRange(image_close, IMG_BLACK_L, IMG_BLACK_H)
    else:       # 若無指定顏色則回傳0
        return 0, 0, 0

    # 取得輪廓點與索引(OpenCV Ver.2/4)
    contours, hierarchy = cv2.findContours(image_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # 設定型心暫存列表
    moment_x_register = []
    moment_y_register = []
    area_register = []

    # 計算型心與面積
    for c in contours:
        # 計算色塊面積
        block_area = cv2.contourArea(c)

        # 面積符合使用門檻方取用(降噪處理)
        if block_area >= IMG_AREA_LIMIT:
            # 取得型心位置(左上為原點，加法增量)
            moment = cv2.moments(c)
            if moment["m00"] != 0:
                moment_x_raw = int(moment["m10"] / moment["m00"])
                moment_y_raw = int(moment["m01"] / moment["m00"])
            else:
                moment_x_raw = 0
                moment_y_raw = 0

            # 計算色塊位置(畫面中心基準，座標位移採增量表示)
            block_x_offset = moment_x_raw - BODY_REF_X + BODY_OFFSET_X
            block_y_offset = BODY_REF_Y - moment_y_raw + BODY_OFFSET_Y

            moment_x_register.append(block_x_offset)
            moment_y_register.append(block_y_offset)
            area_register.append(block_area)

    # 設定型心運算暫存列表
    moment_area_x = 0
    moment_area_y = 0
    area_sum = 0

    # 計算面積一次矩
    for i in range(len(moment_x_register)):
        moment_area_x += moment_x_register[i] * area_register[i]
        moment_area_y += moment_y_register[i] * area_register[i]
        area_sum += area_register[i]

    # 求得組合型心位置
    if len(moment_x_register) != 0:
        block_x = moment_area_x / area_sum
        block_y = moment_area_y / area_sum
    else:
        block_x = 0
        block_y = 0

    # 輸出影像(GUI顯示用)
    if MODE == 'POS':
        CAM_IMAGE_LIVE = image_mask

    return block_x, block_y, area_sum


# ========== 線條型心計算 =============================================
# 將影像做二值化後再做型心與面積計算，除影響去用過程與畫面中心設定點不同外，其餘與前函式相同
# 但因中間有程式碼不同，為避免增加樹狀複雜度，故不再做副程式呼叫
# 本函式供航向控制器使用，回傳型心位置值、線角度值，面積閥值亦獨立設定(角度執行待驗證)
def line_moment():
    """計算線條型心、線條角度"""
    global CAM_IMAGE_LIVE
    # HSV色域轉換版
    # 色域轉換BGR->HSV
    image_hsv = cv2.cvtColor(CAM_IMAGE, cv2.COLOR_BGR2HSV)
    # 閉運算
    image_close = cv2.morphologyEx(image_hsv, cv2.MORPH_CLOSE,
                                   np.ones(IMG_MORPHOLOGY_EX_KERNEL, np.uint8))
    # 抓黑色輪廓
    img_th = cv2.inRange(image_close, IMG_BLACK_L, IMG_BLACK_H)

    # 灰階轉換版本
    # 色域轉換BGR->GRAY
    # img_gray = cv2.cvtColor(CAM_IMAGE, cv2.COLOR_BGR2GRAY)
    # 高斯模糊
    # img_blur = cv2.GaussianBlur(img_gray, ANG_GB_KERNEL_SIZE, ANG_GB_SIGMA)
    # 二值化
    # ret, img_th = cv2.threshold(img_blur, ANG_THRESHOLD,
    #                             ANG_THRESHOLD_NEW, cv2.THRESH_BINARY_INV)

    # 取得輪廓點與索引(OpenCV Ver.2/4)
    contours, hierarchy = cv2.findContours(img_th, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # 設定型心暫存列表
    moment_x_register = []
    moment_y_register = []
    area_register = []
    line_angle_register = []

    # 計算型心與面積
    for c in contours:
        # 計算色塊面積
        block_area = cv2.contourArea(c)

        # 面積符合使用門檻方取用(降噪處理)
        if block_area >= ANG_LINE_TRACK_AREA_L:
            # 取得型心位置(左上為原點，加法增量)
            moment = cv2.moments(c)
            if moment["m00"] != 0:
                moment_x_raw = int(moment["m10"] / moment["m00"])
                moment_y_raw = int(moment["m01"] / moment["m00"])
            else:
                moment_x_raw = 0
                moment_y_raw = 0

            # 矩形框選
            (x, y, w, h) = cv2.boundingRect(c)

            # 計算角度
            if block_area > ANG_LINE_TRACK_AREA_H:      # 面積大於限制時不輸出角度
                line_angle = 0
            elif h == 0:            # 無讀取到時不輸出(安全裝置)
                line_angle = 0
            else:
                line_angle = math.degrees(math.atan(w / h))
                if ANG_T_TURN == 'left':        # 左T
                    line_angle *= -1
                elif ANG_T_TURN == 'right':     # 右T
                    line_angle *= 1
                elif ANG_T_TURN == 'straight':  # T後直線
                    line_angle = 0
                elif img_th[y, x] == 255 and ANG_T_TURN == '':      # 一般情形
                    line_angle *= -1

            # 計算色塊位置(相對於畫面下方中心)
            block_x_offset = moment_x_raw - BODY_REF_X + BODY_OFFSET_X
            block_y_offset = 2 * BODY_REF_Y - moment_y_raw + BODY_OFFSET_Y

            moment_x_register.append(block_x_offset)
            moment_y_register.append(block_y_offset)
            area_register.append(block_area)
            line_angle_register.append(line_angle)

    # 設定型心運算暫存列表
    moment_area_x = 0
    moment_area_y = 0
    area_sum = 0

    # 求得組合型心位置
    if len(moment_x_register) != 0:
        # 計算面積一次矩
        for i in range(len(moment_x_register)):
            moment_area_x += moment_x_register[i] * area_register[i]
            moment_area_y += moment_y_register[i] * area_register[i]
            area_sum += area_register[i]
        block_x = moment_area_x / area_sum
        block_y = moment_area_y / area_sum
    else:
        block_x = 0
        block_y = 0

    # 計算平均線角度
    if len(line_angle_register) != 0:
        line_angle_ave = -np.average(line_angle_register)
    else:
        line_angle_ave = 0

    # 輸出影像(GUI顯示用)
    if MODE == 'LINE':
        CAM_IMAGE_LIVE = img_th

    return block_x, block_y, area_sum, line_angle_ave


# ========== T字路口判別 ===============================================
def t_intersection_detect():
    """偵測T字路口：讀取照片後依照邊界觸碰數來做判斷"""
    # 色域轉換BGR->GRAY
    img_gray = cv2.cvtColor(CAM_IMAGE, cv2.COLOR_BGR2GRAY)
    # 高斯模糊
    img_blur = cv2.GaussianBlur(img_gray, ANG_GB_KERNEL_SIZE, ANG_GB_SIGMA)
    # 二值化
    ret, img_th = cv2.threshold(img_blur, ANG_THRESHOLD,
                                ANG_THRESHOLD_NEW, cv2.THRESH_BINARY_INV)

    # 邊界判斷(有出現白色255)
    # 上下邊偵測
    up_count = 0
    down_count = 0
    for x in range(len(img_th[0])):
        if img_th[0, x] == 255:
            up_count += 1
        if img_th[-1, x] == 255:
            down_count += 1

    # 左右邊偵測
    left_count = 0
    right_count = 0
    for y in range(len(img_th)):
        if img_th[y, 0] == 255:
            left_count += 1
        if img_th[y, -1] == 255:
            right_count += 1

    # 邊界接線判斷
    if ANG_T_INTERSECT_THRESHOLD_UP >= up_count >= ANG_T_INTERSECT_THRESHOLD_DN:
        up_flag = True
    else:
        up_flag = False

    if ANG_T_INTERSECT_THRESHOLD_UP >= down_count >= ANG_T_INTERSECT_THRESHOLD_DN:
        down_flag = True
    else:
        down_flag = False

    if ANG_T_INTERSECT_THRESHOLD_UP >= left_count >= ANG_T_INTERSECT_THRESHOLD_DN:
        left_flag = True
    else:
        left_flag = False

    if ANG_T_INTERSECT_THRESHOLD_UP >= right_count >= ANG_T_INTERSECT_THRESHOLD_DN:
        right_flag = True
    else:
        right_flag = False

    # 角點註記
    if img_th[0, 0] == 255:
        up_left = True
    else:
        up_left = False

    if img_th[0, -1] == 255:
        up_right = True
    else:
        up_right = False

    if img_th[-1, 0] == 255:
        down_left = True
    else:
        down_left = False

    if img_th[-1, -1] == 255:
        down_right = True
    else:
        down_right = False

    # 路口朝向判斷(有潛在的多重衝突存在)
    if up_flag and down_flag and left_flag:
        return 'left'
    if up_flag and down_flag and right_flag:
        return 'right'
    elif left_flag and right_flag and up_flag:
        return 'straight'
    else:
        return 'NaN'


# ========== 轉換尤拉角到四元數 =========================================
# 因命令格式針對姿態需使用四元數格式，因此需先把尤拉角作轉換，此格式可避免萬向節鎖產生
# 輸入格式為度(degree)，輸出為四元數(q,x,y,z)
def to_quaternion(roll=0.0, pitch=0.0, yaw=0.0):
    """輸入尤拉角轉換成四元數"""
    t0 = math.cos(math.radians(yaw * 0.5))
    t1 = math.sin(math.radians(yaw * 0.5))
    t2 = math.cos(math.radians(roll * 0.5))
    t3 = math.sin(math.radians(roll * 0.5))
    t4 = math.cos(math.radians(pitch * 0.5))
    t5 = math.sin(math.radians(pitch * 0.5))

    w = t0 * t2 * t4 + t1 * t3 * t5
    x = t0 * t3 * t4 - t1 * t2 * t5
    y = t0 * t2 * t5 + t1 * t3 * t4
    z = t1 * t2 * t4 - t0 * t3 * t5

    return [w, x, y, z]


# ========= 沙包盒初始化 ===============================================
# 於起飛時送出脈波訊號固定艙門，防止貨物掉出，並維持較佳的空氣動力學外型
def box_initialize():
    """初始化沙包盒艙門，通電至關門位"""
    global STATE_BOX, BOX
    # 初始化腳位
    GPIO.setmode(GPIO.BCM)                      # 設定為BCM模式
    GPIO.setwarnings(False)                     # 忽略警告
    GPIO.setup(BOX_PWM_PIN, GPIO.OUT)           # PWM腳位輸出設定
    BOX = GPIO.PWM(BOX_PWM_PIN, BOX_PWM_FREQ)   # 建立PWM控制物件

    # 啟動PWM控制，使門維持關閉
    dc = BOX_PWM_FREQ * (BOX_PWM_CLOSE / 1000000) * 100
    BOX.start(dc)

    # 回報艙門狀態
    STATE_BOX = True
    print('<飛行狀態> 沙包盒艙門：關閉\n')


# ========= 沙包盒關門 =================================================
# 於起飛時送出脈波訊號固定艙門，防止貨物掉出，並維持較佳的空氣動力學外型
def box_close():
    """沙包盒關艙門命令"""
    global STATE_BOX
    # 啟動PWM控制，使門維持關閉
    dc = BOX_PWM_FREQ * (BOX_PWM_CLOSE / 1000000) * 100
    BOX.ChangeDutyCycle(dc)

    # 回報艙門狀態
    STATE_BOX = True
    print('<飛行狀態> 沙包盒艙門：關閉\n')


# ======== 沙包盒開門 =================================================
# 輸出訊號使艙門開啟，使貨物卸出
def box_open():
    """沙包盒開艙門命令"""
    global STATE_BOX
    # 啟動PWM控制，使門打開
    dc = BOX_PWM_FREQ * (BOX_PWM_OPEN / 1000000) * 100
    BOX.ChangeDutyCycle(dc)

    # 回報艙門狀態
    STATE_BOX = False
    print('<飛行狀態> 沙包盒艙門：開啟\n')


# ====================================================================
# ========== 數值紀錄區塊 ==============================================
# ====================================================================


# ========== Excel 輸出 ===============================================
def data_sheet():
    # 建立活頁簿
    db = Workbook()

    # 建立分頁
    ds1 = db.create_sheet("ALT Controller")
    ds2 = db.create_sheet("POS Controller")
    ds2b = db.create_sheet("POS-VEL Controller")
    ds3 = db.create_sheet("HDG Controller")
    ds4 = db.create_sheet("ALT Sensor")

    # 高度控制器(先列在行)
    ds1.cell(1, 1, '高度控制器')
    ds1.cell(2, 1, '實際高度')
    for x in range(len(R_ACTUAL_ALT_CTRL)):
        ds1.cell(x + 3, 1, R_ACTUAL_ALT_CTRL[x])

    ds1.cell(2, 2, '目標高度')
    for x in range(len(R_TARGET_ALT)):
        ds1.cell(x + 3, 2, R_TARGET_ALT[x])

    ds1.cell(2, 3, '爬升率')
    for x in range(len(R_TARGET_THRUST)):
        ds1.cell(x + 3, 3, R_TARGET_THRUST[x])

    # 高度控制器PID響應
    ds1.cell(1, 5, '高度控制器PID響應')
    ds1.cell(2, 5, 'Error')
    for x in range(len(R_ALT_ERROR)):
        ds1.cell(x + 3, 5, R_ALT_ERROR[x])

    ds1.cell(2, 6, 'Gain P')
    for x in range(len(R_ALT_GP)):
        ds1.cell(x + 3, 6, R_ALT_GP[x])

    ds1.cell(2, 7, 'Gain I')
    for x in range(len(R_ALT_GI)):
        ds1.cell(x + 3, 7, R_ALT_GI[x])

    ds1.cell(2, 8, 'Gain D')
    for x in range(len(R_ALT_GD)):
        ds1.cell(x + 3, 8, R_ALT_GD[x])

    # 位置控制器-第二頁
    # 位置控制器X響應圖
    ds2.cell(1, 1, '位置控制器X響應圖')
    ds2.cell(2, 1, 'X位置')
    for x in range(len(R_POS_X)):
        ds2.cell(x + 3, 1, R_POS_X[x])

    ds2.cell(2, 2, 'X速度輸出')
    for x in range(len(R_POS_X_TARGET_VEL)):
        ds2.cell(x + 3, 2, R_POS_X_TARGET_VEL[x])

    ds2.cell(2, 3, '面積大小')
    for x in range(len(R_POS_AREA)):
        ds2.cell(x + 3, 3, R_POS_AREA[x])

    # 位置控制器Y響應圖
    ds2.cell(1, 5, '位置控制器Y響應圖')
    ds2.cell(2, 5, 'Y位置')
    for x in range(len(R_POS_Y)):
        ds2.cell(x + 3, 5, R_POS_Y[x])

    ds2.cell(2, 6, 'Y速度輸出')
    for x in range(len(R_POS_Y_TARGET_VEL)):
        ds2.cell(x + 3, 6, R_POS_Y_TARGET_VEL[x])

    ds2.cell(2, 7, '面積大小')
    for x in range(len(R_POS_AREA)):
        ds2.cell(x + 3, 7, R_POS_AREA[x])

    # 軌跡記錄
    ds2.cell(1, 9, '位置控制器Y響應圖')
    ds2.cell(2, 9, 'X位置')
    for x in range(len(R_POS_X)):
        ds2.cell(x + 3, 9, R_POS_X[x])

    ds2.cell(2, 10, 'Y位置')
    for x in range(len(R_POS_Y)):
        ds2.cell(x + 3, 10, R_POS_Y[x])

    # 位置控制器X-PID響應圖
    ds2.cell(1, 12, '位置控制器X-PID響應圖')
    ds2.cell(2, 12, 'Error')
    for x in range(len(R_POS_X_ERROR)):
        ds2.cell(x + 3, 12, R_POS_X_ERROR[x])

    ds2.cell(2, 13, 'Gain P')
    for x in range(len(R_POS_X_GP)):
        ds2.cell(x + 3, 13, R_POS_X_GP[x])

    ds2.cell(2, 14, 'Gain I')
    for x in range(len(R_POS_X_GI)):
        ds2.cell(x + 3, 14, R_POS_X_GI[x])

    ds2.cell(2, 15, 'Gain D')
    for x in range(len(R_POS_X_GD)):
        ds2.cell(x + 3, 15, R_POS_X_GD[x])

    # 位置控制器Y-PID響應圖
    ds2.cell(1, 17, '位置控制器Y-PID響應圖')
    ds2.cell(2, 17, 'Error')
    for x in range(len(R_POS_Y_ERROR)):
        ds2.cell(x + 3, 17, R_POS_Y_ERROR[x])

    ds2.cell(2, 18, 'Gain P')
    for x in range(len(R_POS_Y_GP)):
        ds2.cell(x + 3, 18, R_POS_Y_GP[x])

    ds2.cell(2, 19, 'Gain I')
    for x in range(len(R_POS_Y_GI)):
        ds2.cell(x + 3, 19, R_POS_Y_GI[x])

    ds2.cell(2, 20, 'Gain D')
    for x in range(len(R_POS_Y_GD)):
        ds2.cell(x + 3, 20, R_POS_Y_GD[x])

    # 速度控制器-第三頁
    # 速度控制器X響應圖
    ds2b.cell(1, 1, '速度控制器X響應圖')
    ds2b.cell(2, 1, 'X目標速度')
    for x in range(len(R_VEL_X_TAR)):
        ds2b.cell(x + 3, 1, R_VEL_X_TAR[x])

    ds2b.cell(2, 2, 'X實際速度')
    for x in range(len(R_POS_VEL_X_ACT)):
        ds2b.cell(x + 3, 2, R_POS_VEL_X_ACT[x])

    ds2b.cell(2, 3, '目標Roll')
    for x in range(len(R_POS_TARGET_ROLL)):
        ds2b.cell(x + 3, 3, R_POS_TARGET_ROLL[x])

    ds2b.cell(2, 4, '實際Roll')
    for x in range(len(R_POS_ACTUAL_ROLL)):
        ds2b.cell(x + 3, 4, R_POS_ACTUAL_ROLL[x])

    # 速度控制器Y響應圖
    ds2b.cell(1, 6, '速度控制器Y響應圖')
    ds2b.cell(2, 6, 'Y目標速度')
    for x in range(len(R_VEL_Y_TAR)):
        ds2b.cell(x + 3, 6, R_VEL_Y_TAR[x])

    ds2b.cell(2, 7, 'Y實際速度')
    for x in range(len(R_POS_VEL_Y_ACT)):
        ds2b.cell(x + 3, 7, R_POS_VEL_Y_ACT[x])

    ds2b.cell(2, 8, '目標Pitch')
    for x in range(len(R_POS_TARGET_PITCH)):
        ds2b.cell(x + 3, 8, R_POS_TARGET_PITCH[x])

    ds2b.cell(2, 9, '實際Pitch')
    for x in range(len(R_POS_ACTUAL_PITCH)):
        ds2b.cell(x + 3, 9, R_POS_ACTUAL_PITCH[x])

    # 速度控制器X-PID響應圖
    ds2b.cell(1, 11, '速度控制器X-PID響應圖')
    ds2b.cell(2, 11, 'Error')
    for x in range(len(R_POS_VEL_X_ERROR)):
        ds2b.cell(x + 3, 11, R_POS_VEL_X_ERROR[x])

    ds2b.cell(2, 12, 'Gain P')
    for x in range(len(R_POS_VEL_X_GP)):
        ds2b.cell(x + 3, 12, R_POS_VEL_X_GP[x])

    ds2b.cell(2, 13, 'Gain I')
    for x in range(len(R_POS_VEL_X_GI)):
        ds2b.cell(x + 3, 13, R_POS_VEL_X_GI[x])

    ds2b.cell(2, 14, 'Gain D')
    for x in range(len(R_POS_VEL_X_GD)):
        ds2b.cell(x + 3, 14, R_POS_VEL_X_GD[x])

    # 速度控制器Y-PID響應圖
    ds2b.cell(1, 16, '速度控制器Y-PID響應圖')
    ds2b.cell(2, 16, 'Error')
    for x in range(len(R_POS_VEL_Y_ERROR)):
        ds2b.cell(x + 3, 16, R_POS_VEL_Y_ERROR[x])

    ds2b.cell(2, 17, 'Gain P')
    for x in range(len(R_POS_VEL_Y_GP)):
        ds2b.cell(x + 3, 17, R_POS_VEL_Y_GP[x])

    ds2b.cell(2, 18, 'Gain I')
    for x in range(len(R_POS_VEL_Y_GI)):
        ds2b.cell(x + 3, 18, R_POS_VEL_Y_GI[x])

    ds2b.cell(2, 19, 'Gain D')
    for x in range(len(R_POS_VEL_Y_GD)):
        ds2b.cell(x + 3, 19, R_POS_VEL_Y_GD[x])

    # 第四頁-循線控制器

    # ->航向控制器
    # 航向-X位置控制器數值處理圖
    ds3.cell(1, 1, '航向-X位置控制器數值處理圖')
    ds3.cell(2, 1, '與線中心距')
    for x in range(len(R_HDG_LINE_ACTUAL_X)):
        ds3.cell(x + 3, 1, R_HDG_LINE_ACTUAL_X[x])

    ds3.cell(2, 2, '輸出左右移速度')
    for x in range(len(R_HDG_LINE_TARGET_VX)):
        ds3.cell(x + 3, 2, R_HDG_LINE_TARGET_VX[x])

    # 航向-X位置控制器PID數值處理圖
    ds3.cell(1, 4, '航向-X位置控制器PID數值處理圖')
    ds3.cell(2, 4, 'Error')
    for x in range(len(R_HDG_LINE_X_ERROR)):
        ds3.cell(x + 3, 4, R_HDG_LINE_X_ERROR[x])

    ds3.cell(2, 5, 'Gain P')
    for x in range(len(R_HDG_LINE_X_GP)):
        ds3.cell(x + 3, 5, R_HDG_LINE_X_GP[x])

    ds3.cell(2, 6, 'Gain I')
    for x in range(len(R_HDG_LINE_X_GI)):
        ds3.cell(x + 3, 6, R_HDG_LINE_X_GI[x])

    ds3.cell(2, 7, 'Gain D')
    for x in range(len(R_HDG_LINE_X_GD)):
        ds3.cell(x + 3, 7, R_HDG_LINE_X_GD[x])

    # 航向-X速度控制器數值處理圖
    ds3.cell(1, 9, '航向-X速度控制器數值處理圖')
    ds3.cell(2, 9, '目標X速度')
    for x in range(len(R_HDG_LINE_VX_TARGET)):
        ds3.cell(x + 3, 9, R_HDG_LINE_VX_TARGET[x])

    ds3.cell(2, 10, '實際X速度')
    for x in range(len(R_HDG_LINE_VX_ACTUAL)):
        ds3.cell(x + 3, 10, R_HDG_LINE_VX_ACTUAL[x])

    ds3.cell(2, 11, '輸出滾轉角')
    for x in range(len(R_HDG_LINE_VX_TARGET_ROLL)):
        ds3.cell(x + 3, 11, R_HDG_LINE_VX_TARGET_ROLL[x])

    ds3.cell(2, 12, '實際滾轉角')
    for x in range(len(R_HDG_LINE_VX_ACTUAL_ROLL)):
        ds3.cell(x + 3, 12, R_HDG_LINE_VX_ACTUAL_ROLL[x])

    # 航向-X速度控制器PID數值處理圖
    ds3.cell(1, 14, '航向-X速度控制器PID數值處理圖')
    ds3.cell(2, 14, 'Error')
    for x in range(len(R_HDG_LINE_VX_ERROR)):
        ds3.cell(x + 3, 14, R_HDG_LINE_VX_ERROR[x])

    ds3.cell(2, 15, 'Gain P')
    for x in range(len(R_HDG_LINE_VX_GP)):
        ds3.cell(x + 3, 15, R_HDG_LINE_VX_GP[x])

    ds3.cell(2, 16, 'Gain I')
    for x in range(len(R_HDG_LINE_VX_GI)):
        ds3.cell(x + 3, 16, R_HDG_LINE_VX_GI[x])

    ds3.cell(2, 17, 'Gain D')
    for x in range(len(R_HDG_LINE_VX_GD)):
        ds3.cell(x + 3, 17, R_HDG_LINE_VX_GD[x])

    # 航向-偏航控制器數值處理圖
    ds3.cell(1, 19, '航向-偏航控制器數值處理圖')
    ds3.cell(2, 19, '實際角度差')
    for x in range(len(R_HDG_YAW_ACTUAL)):
        ds3.cell(x + 3, 19, R_HDG_YAW_ACTUAL[x])

    ds3.cell(2, 20, '輸出偏航角速率')
    for x in range(len(R_HDG_YAW_TARGET_YAW_RATE)):
        ds3.cell(x + 3, 20, R_HDG_YAW_TARGET_YAW_RATE[x])

    # 航向控制器PID數值處理圖
    ds3.cell(1, 22, '航向-偏航控制器PID數值處理圖')
    ds3.cell(2, 22, 'Error')
    for x in range(len(R_HDG_YAW_ERROR)):
        ds3.cell(x + 3, 22, R_HDG_YAW_ERROR[x])

    ds3.cell(2, 23, 'Gain P')
    for x in range(len(R_HDG_YAW_GP)):
        ds3.cell(x + 3, 23, R_HDG_YAW_GP[x])

    ds3.cell(2, 24, 'Gain I')
    for x in range(len(R_HDG_YAW_GI)):
        ds3.cell(x + 3, 24, R_HDG_YAW_GI[x])

    ds3.cell(2, 25, 'Gain D')
    for x in range(len(R_HDG_YAW_GD)):
        ds3.cell(x + 3, 25, R_HDG_YAW_GP[x])

    # 飛行速度控制器
    ds3.cell(1, 27, '飛行速度控制器數值處理圖')
    ds3.cell(2, 27, '目標速度')
    for x in range(len(R_SPD_TARGET)):
        ds3.cell(x + 3, 27, R_SPD_TARGET[x])

    ds3.cell(2, 28, '實際速度')
    for x in range(len(R_SPD_ACTUAL)):
        ds3.cell(x + 3, 28, R_SPD_ACTUAL[x])

    ds3.cell(2, 29, '目標傾角')
    for x in range(len(R_SPD_TARGET_PITCH)):
        ds3.cell(x + 3, 29, R_SPD_TARGET_PITCH[x])

    ds3.cell(2, 30, '實際傾角')
    for x in range(len(R_SPD_ACTUAL_PITCH)):
        ds3.cell(x + 3, 30, R_SPD_ACTUAL_PITCH[x])

    # 飛行速度控制器PID數值處理圖
    ds3.cell(1, 32, '飛行速度控制器PID數值處理圖')
    ds3.cell(2, 32, 'Error')
    for x in range(len(R_SPD_ERROR)):
        ds3.cell(x + 3, 32, R_SPD_ERROR[x])

    ds3.cell(2, 33, 'Gain P')
    for x in range(len(R_SPD_GP)):
        ds3.cell(x + 3, 33, R_SPD_GP[x])

    ds3.cell(2, 34, 'Gain I')
    for x in range(len(R_SPD_GI)):
        ds3.cell(x + 3, 34, R_SPD_GI[x])

    ds3.cell(2, 35, 'Gain D')
    for x in range(len(R_SPD_GP)):
        ds3.cell(x + 3, 35, R_SPD_GP[x])

    # 飛行速度控制器
    ds3.cell(1, 37, '飛行速度控制器數值處理圖-開迴路')
    ds3.cell(2, 37, '目標俯仰角')
    for x in range(len(R_SPD_OL_TARGET_PITCH)):
        ds3.cell(x + 3, 37, R_SPD_OL_TARGET_PITCH[x])

    ds3.cell(2, 38, '實際俯仰角')
    for x in range(len(R_SPD_OL_ACTUAL_PITCH)):
        ds3.cell(x + 3, 38, R_SPD_OL_ACTUAL_PITCH[x])

    ds3.cell(2, 39, '實際速度')
    for x in range(len(R_SPD_OL_ACTUAL_SPD)):
        ds3.cell(x + 3, 39, R_SPD_OL_ACTUAL_SPD[x])

    # 第五頁-感測器
    # 高度感測器數值處理圖
    ds4.cell(1, 1, '高度感測器數值處理圖')
    ds4.cell(2, 1, '實際高')
    for x in range(len(R_ACTUAL_ALT)):
        ds4.cell(x + 3, 1, R_ACTUAL_ALT[x])

    ds4.cell(2, 2, '超音波')
    for x in range(len(R_ACTUAL_ALT_SONAR)):
        ds4.cell(x + 3, 2, R_ACTUAL_ALT_SONAR[x])

    ds4.cell(2, 3, 'IMU')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds4.cell(x + 3, 3, R_ACTUAL_ALT_IMU_KF[x])

    # 超音波高度計數值處理圖
    ds4.cell(1, 6, '超音波高度計數值處理圖')
    ds4.cell(2, 6, '原始值')
    for x in range(len(R_ACTUAL_ALT_SONAR_RAW)):
        ds4.cell(x + 3, 6, R_ACTUAL_ALT_SONAR_RAW[x])

    ds4.cell(2, 7, '補正後')
    for x in range(len(R_ACTUAL_ALT_SONAR_OFFSET)):
        ds4.cell(x + 3, 7, R_ACTUAL_ALT_SONAR_OFFSET[x])

    ds4.cell(2, 8, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_AVE)):
        ds4.cell(x + 3, 8, R_ACTUAL_ALT_SONAR_AVE[x])

    ds4.cell(2, 9, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_SONAR_KF)):
        ds4.cell(x + 3, 9, R_ACTUAL_ALT_SONAR_KF[x])

    # IMU高度計數值處理圖
    ds4.cell(1, 11, 'IMU高度計數值處理圖')
    ds4.cell(2, 11, '原始值')
    for x in range(len(R_ACTUAL_ALT_IMU_RAW)):
        ds4.cell(x + 3, 11, R_ACTUAL_ALT_IMU_RAW[x])

    ds4.cell(2, 12, '補正後')
    for x in range(len(R_ACTUAL_ALT_IMU_OFFSET)):
        ds4.cell(x + 3, 12, R_ACTUAL_ALT_IMU_OFFSET[x])

    ds4.cell(2, 13, '均值濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_AVE)):
        ds4.cell(x + 3, 13, R_ACTUAL_ALT_IMU_AVE[x])

    ds4.cell(2, 14, 'KF濾波')
    for x in range(len(R_ACTUAL_ALT_IMU_KF)):
        ds4.cell(x + 3, 14, R_ACTUAL_ALT_IMU_KF[x])

    # 依照時間存檔
    t_save = time.localtime()
    db.save('Flight Data %d_%d_%d_%d_%d_%d.xlsx' %
            (t_save[0], t_save[1], t_save[2], t_save[3], t_save[4], t_save[5]))


# ========== 高度控制器 ================================================
def alt_ctrl_data():
    # ======高度控制器================
    # 高度控制器響應圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_CTRL) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_CTRL, label="Actual Altitude")
    plt.plot(num_alt, R_TARGET_ALT, label="Target Altitude")
    plt.plot(num_alt, R_TARGET_THRUST, label="Climb Rate")

    plt.xlabel('Data number')
    plt.ylabel('Height(m), Velocity(m/s)')
    plt.title('Altitude Controller Data')
    plt.legend()
    plt.savefig('Altitude Controller Data.png')
    plt.show()

    # 高度控制器PID輸出圖
    num_alt = []
    for i in range(1, len(R_ALT_ERROR) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ALT_ERROR, label="Error")
    plt.plot(num_alt, R_ALT_GP, label="Gain P")
    plt.plot(num_alt, R_ALT_GI, label="Gain I")
    plt.plot(num_alt, R_ALT_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Height(m), Velocity(m/s)')
    plt.title('Altitude Controller PID Output Data')
    plt.legend()
    plt.savefig('Altitude Controller PID Output Data.png')
    plt.show()


# ========== 位置控制器 ================================================
def pos_ctrl_data():
    # 位置控制器X響應圖
    num_pos = []
    for i in range(1, len(R_POS_X) + 1):
        num_pos.append(i)

    plt.plot(num_pos, R_POS_X, label="X Position")
    plt.plot(num_pos, R_POS_X_TARGET_VEL, label="X Velocity Output")
    plt.plot(num_pos, R_POS_AREA, label="Block Area")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller Data - X axis')
    plt.legend()
    plt.savefig('Position Controller Data - X axis.png')
    plt.show()

    # 位置控制器Y響應圖
    plt.plot(num_pos, R_POS_Y, label="Y Position")
    plt.plot(num_pos, R_POS_Y_TARGET_VEL, label="Y Velocity Output")
    plt.plot(num_pos, R_POS_AREA, label="Block Area")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller Data - Y axis')
    plt.legend()
    plt.savefig('Position Controller Data - Y axis.png')
    plt.show()

    # 位置控制器X-PID響應圖
    plt.plot(num_pos, R_POS_X_ERROR, label="Error")
    plt.plot(num_pos, R_POS_X_GP, label="Gain P")
    plt.plot(num_pos, R_POS_X_GI, label="Gain I")
    plt.plot(num_pos, R_POS_X_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller PID Output Data - X axis')
    plt.legend()
    plt.savefig('Position Controller PID Output Data - X axis.png')
    plt.show()

    # 位置控制器Y-PID響應圖
    plt.plot(num_pos, R_POS_Y_ERROR, label="Error")
    plt.plot(num_pos, R_POS_Y_GP, label="Gain P")
    plt.plot(num_pos, R_POS_Y_GI, label="Gain I")
    plt.plot(num_pos, R_POS_Y_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Position(px), Velocity(px/s)')
    plt.title('Position Controller PID Output Data - Y axis')
    plt.legend()
    plt.savefig('Position Controller PID Output Data - Y axis.png')
    plt.show()

    # 位置控制器軌跡響應圖
    plt.plot(R_POS_X, R_POS_Y, label="Block Trajectory")

    plt.xlabel('X-Position(px)')
    plt.ylabel('Y-Position(px)')
    plt.title('Position Controller Trajectory Data')
    plt.legend()
    plt.savefig('Position Controller Trajectory Data.png')
    plt.show()


# ========== 速度控制器 ================================================
def vel_ctrl_data():
    # 位置控制器X響應圖
    num_vel = []
    for i in range(1, len(R_VEL_X_TAR) + 1):
        num_vel.append(i)

    # 速度控制器X響應圖
    plt.plot(num_vel, R_VEL_X_TAR, label="Target X Velocity")
    plt.plot(num_vel, R_POS_VEL_X_ACT, label="Actual X Velocity")
    plt.plot(num_vel, R_POS_TARGET_ROLL, label="Target Roll Angle")
    plt.plot(num_vel, R_POS_ACTUAL_ROLL, label="Actual Roll Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller Data - X axis')
    plt.legend()
    plt.savefig('Velocity Controller Data - X axis.png')
    plt.show()

    # 速度控制器Y響應圖
    plt.plot(num_vel, R_VEL_Y_TAR, label="Target Y Velocity")
    plt.plot(num_vel, R_POS_VEL_Y_ACT, label="Actual Y Velocity")
    plt.plot(num_vel, R_POS_TARGET_PITCH, label="Target Pitch Angle")
    plt.plot(num_vel, R_POS_ACTUAL_PITCH, label="Actual Pitch Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller Data - Y axis')
    plt.legend()
    plt.savefig('Velocity Controller Data - Y axis.png')
    plt.show()

    # 速度控制器X-PID響應圖
    plt.plot(num_vel, R_POS_VEL_X_ERROR, label="Error")
    plt.plot(num_vel, R_POS_VEL_X_GP, label="Gain P")
    plt.plot(num_vel, R_POS_VEL_X_GI, label="Gain I")
    plt.plot(num_vel, R_POS_VEL_X_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller PID Output Data - X axis')
    plt.legend()
    plt.savefig('Velocity Controller PID Output Data - X axis.png')
    plt.show()

    # 速度控制器Y-PID響應圖
    plt.plot(num_vel, R_POS_VEL_Y_ERROR, label="Error")
    plt.plot(num_vel, R_POS_VEL_Y_GP, label="Gain P")
    plt.plot(num_vel, R_POS_VEL_Y_GI, label="Gain I")
    plt.plot(num_vel, R_POS_VEL_Y_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(px/s), Angle(degree)')
    plt.title('Velocity Controller PID Output Data - Y axis')
    plt.legend()
    plt.savefig('Velocity Controller PID Output Data - Y axis.png')
    plt.show()


# ========== 高度感測器 ================================================
def alt_sen_data():
    # 高度感測器數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT, label="Actual Height")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR, label="Sonar Data")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_KF, label="IMU Data")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Sensor Data')
    plt.legend()
    plt.savefig('Altitude Sensor Data.png')
    plt.show()

    # 超音波高度計數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_SONAR_KF) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_RAW, label="Raw Data")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_OFFSET, label="Angle & Height Offset")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_AVE, label="Average Filter")
    plt.plot(num_alt, R_ACTUAL_ALT_SONAR_KF, label="Kalman Filter")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Data - Sonar')
    plt.legend()
    plt.savefig('Altitude Data - Sonar.png')
    plt.show()

    # IMU高度計數值處理圖
    num_alt = []
    for i in range(1, len(R_ACTUAL_ALT_IMU_RAW) + 1):
        num_alt.append(i)

    plt.plot(num_alt, R_ACTUAL_ALT_IMU_RAW, label="Raw Data")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_OFFSET, label="Height Offset")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_AVE, label="Average Filter")
    plt.plot(num_alt, R_ACTUAL_ALT_IMU_KF, label="Kalman Filter")

    plt.xlabel('Data number')
    plt.ylabel('Height(m)')
    plt.title('Altitude Data - IMU')
    plt.legend()
    plt.savefig('Altitude Data - IMU.png')
    plt.show()


# ========== 航向控制器 ================================================
def hdg_ctrl_data():
    # 航向控制器外迴圈響應圖
    num_hdg = []
    for i in range(1, len(R_HDG_LINE_ACTUAL_X) + 1):
        num_hdg.append(i)

    # 航向-線位置控制器響應圖
    plt.plot(num_hdg, R_HDG_LINE_ACTUAL_X, label="Line Distance")
    plt.plot(num_hdg, R_HDG_LINE_TARGET_VX, label="Target X Velocity")

    plt.xlabel('Data number')
    plt.ylabel('Distance(m), Velocity(m/s)')
    plt.title('Heading Controller Distance Data')
    plt.legend()
    plt.savefig('Heading Controller Distance Data.png')
    plt.show()

    # 航向-線位置控制器PID響應圖
    plt.plot(num_hdg, R_HDG_LINE_X_ERROR, label="Error")
    plt.plot(num_hdg, R_HDG_LINE_X_GP, label="Gain P")
    plt.plot(num_hdg, R_HDG_LINE_X_GI, label="Gain I")
    plt.plot(num_hdg, R_HDG_LINE_X_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Distance(m), Velocity(m/s)')
    plt.title('Heading Controller Line Distance PID Output Data')
    plt.legend()
    plt.savefig('Heading Controller Line Distance PID Output Data.png')
    plt.show()

    # 航向控制器內迴圈響應圖
    num_hdg = []
    for i in range(1, len(R_HDG_LINE_VX_TARGET) + 1):
        num_hdg.append(i)

    # 航向-線速度控制器響應圖
    plt.plot(num_hdg, R_HDG_LINE_VX_TARGET, label="Target X Velocity")
    plt.plot(num_hdg, R_HDG_LINE_VX_ACTUAL, label="ACTUAL X Velocity")
    plt.plot(num_hdg, R_HDG_LINE_VX_TARGET_ROLL, label="Target Roll")
    plt.plot(num_hdg, R_HDG_LINE_VX_ACTUAL_ROLL, label="Actual Roll")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s), Angle(degree)')
    plt.title('Heading Controller X Velocity Data')
    plt.legend()
    plt.savefig('Heading Controller X Velocity Data.png')
    plt.show()

    # 航向-線速度控制器PID響應圖
    plt.plot(num_hdg, R_HDG_LINE_VX_ERROR, label="Error")
    plt.plot(num_hdg, R_HDG_LINE_VX_GP, label="Gain P")
    plt.plot(num_hdg, R_HDG_LINE_VX_GI, label="Gain I")
    plt.plot(num_hdg, R_HDG_LINE_VX_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s), Angle(degree)')
    plt.title('Heading Controller X Velocity PID Output Data')
    plt.legend()
    plt.savefig('Heading Controller X Velocity PID Output Data.png')
    plt.show()

    # 航向-偏航控制器響應圖
    plt.plot(num_hdg, R_HDG_YAW_ACTUAL, label="Actual Heading")
    plt.plot(num_hdg, R_HDG_YAW_TARGET_YAW_RATE, label="Target Yaw Rate")

    plt.xlabel('Data number')
    plt.ylabel('Heading(degree), Angular Rate(degree/s)')
    plt.title('Heading Controller Yaw Data')
    plt.legend()
    plt.savefig('Heading Controller Yaw Data.png')
    plt.show()

    # 航向-偏航控制器PID響應圖
    plt.plot(num_hdg, R_HDG_YAW_ERROR, label="Error")
    plt.plot(num_hdg, R_HDG_YAW_GP, label="Gain P")
    plt.plot(num_hdg, R_HDG_YAW_GI, label="Gain I")
    plt.plot(num_hdg, R_HDG_YAW_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Heading(degree), Angular Rate(degree/s)')
    plt.title('Heading Controller Yaw PID Output Data')
    plt.legend()
    plt.savefig('Heading Controller Yaw PID Output Data.png')
    plt.show()


# ========== 飛行速度控制器 =============================================
def spd_ctrl_cl_data():
    # 飛行速度控制器響應圖-閉迴路
    num_spd = []
    for i in range(1, len(R_SPD_TARGET) + 1):
        num_spd.append(i)

    # 速度控制器X響應圖
    plt.plot(num_spd, R_SPD_TARGET, label="Target Speed")
    plt.plot(num_spd, R_SPD_ACTUAL, label="Actual Speed")
    plt.plot(num_spd, R_SPD_TARGET_PITCH, label="Target Pitch Angle")
    plt.plot(num_spd, R_SPD_ACTUAL_PITCH, label="Actual Pitch Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s), Angle(degree)')
    plt.title('Flight Speed Controller Data')
    plt.legend()
    plt.savefig('Flight Speed Controller Data.png')
    plt.show()

    # 速度控制器X-PID響應圖
    plt.plot(num_spd, R_SPD_ERROR, label="Error")
    plt.plot(num_spd, R_SPD_GP, label="Gain P")
    plt.plot(num_spd, R_SPD_GI, label="Gain I")
    plt.plot(num_spd, R_SPD_GD, label="Gain D")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s)')
    plt.title('Flight Speed Controller PID Output Data')
    plt.legend()
    plt.savefig('Flight Speed Controller PID Output Data.png')
    plt.show()


# ========== 飛行速度控制器 =============================================
def spd_ctrl_ol_data():
    # 飛行速度控制器響應圖-開迴路
    num_spd = []
    for i in range(1, len(R_SPD_OL_ACTUAL_SPD) + 1):
        num_spd.append(i)

    # 速度控制器X響應圖
    plt.plot(num_spd, R_SPD_OL_ACTUAL_SPD, label="Actual Speed")
    plt.plot(num_spd, R_SPD_OL_TARGET_PITCH, label="Target Pitch Angle")
    plt.plot(num_spd, R_SPD_OL_ACTUAL_PITCH, label="Actual Pitch Angle")

    plt.xlabel('Data number')
    plt.ylabel('Velocity(m/s), Angle(degree)')
    plt.title('Flight Speed Controller OL Data')
    plt.legend()
    plt.savefig('Flight Speed Controller OL Data.png')
    plt.show()


# ====================================================================
# ========== 飛行程序指令(條件與命令) ====================================
# ====================================================================
# 任務中會使用的區段飛行命令、模式切換命令、觸發條件


# ========== 多線程啟動函式 ============================================
# 供主程式啟動多線程用，若連線中斷時，自動重新連線
def thread_activate():
    global STATE_THREAD, STATE_UAV, uav
    # 若未連線，重新執行連線
    if not STATE_UAV:
        print('<啟動程序> 與UAV連線中，請稍候...')
        # 樹莓派連線指令
        uav = connect(UAV_CONNECT_PORT, baud=UAV_CONNECT_BAUD)
        STATE_UAV = True
        print('<啟動程序> UAV已連線')

    # 定義多線程
    cam = threading.Thread(target=camera)
    cam_sen = threading.Thread(target=camera_sensor)
    sonar_sen = threading.Thread(target=altitude_sensor_sonar)
    alt_sen = threading.Thread(target=altitude_sensor)
    spd_sen = threading.Thread(target=speed_sensor)
    alt_ctrl = threading.Thread(target=altitude_controller)
    pos_ctrl = threading.Thread(target=position_controller)
    pos_vel_ctrl = threading.Thread(target=position_velocity_controller)
    hdg_out_ctrl = threading.Thread(target=heading_outer_controller)
    hdg_in_ctrl = threading.Thread(target=heading_inner_controller)
    spd_ctrl = threading.Thread(target=flight_speed_controller)
    cmd = threading.Thread(target=command_transfer)

    # 啟動多線程旗標
    STATE_THREAD = True

    # 執行多線程(依啟動順序放置)
    cam.start()
    cam_sen.start()
    sonar_sen.start()
    alt_sen.start()
    spd_sen.start()
    alt_ctrl.start()
    pos_ctrl.start()
    pos_vel_ctrl.start()
    hdg_out_ctrl.start()
    hdg_in_ctrl.start()
    spd_ctrl.start()
    cmd.start()

    # 回報多線程啟動
    time.sleep(5)
    print('<啟動程序> 多線程程序：啟動')


# ========== 啟動前檢查 ================================================
# 本程式為飛行前的安全檢查項目，除確認控制器是否上線外，也提醒控制員檢查安全程序
# 項目有：安全裝置移除(槳片銷)、起飛區淨空，完成後即交付起飛控制
def pre_check():
    """安全檢查程序"""
    # 執行起飛前檢查
    print('<啟動程序> 啟動前檢查程序：啟動')

    # 階段一：控制器檢查
    # 等待高度控制器啟動
    if not STATE_ALT:
        print('<啟動程序> 主程序：等待高度控制器啟動')
    while not STATE_ALT:
        time.sleep(0.1)

    # 等待影像感測啟動
    if not STATE_SENSOR_CAM:
        print('<啟動程序> 主程序：等待影像感測器啟動')
    while not STATE_SENSOR_CAM:
        time.sleep(0.1)

    # 等待命令傳輸啟動
    if not STATE_CMD:
        print('<啟動程序> 主程序：等待命令傳輸啟動')
    while not STATE_CMD:
        time.sleep(0.1)

    # 初始化儲物倉
    box_initialize()

    # 階段三：安全檢查程序
    # 確認安全裝置解除
    safe_pin = 'N'
    while safe_pin != 'Y':
        safe_pin = input('<啟動程序> 請確認安全裝置已移除(Y/N) => ')
    print('<啟動程序> 安全裝置已移除')
    time.sleep(0.5)

    # 確認起飛區已淨空
    lunch_clear = 'N'
    while lunch_clear != 'Y':
        lunch_clear = input('<啟動程序> 請確認起飛區已淨空(Y/N) => ')
    print('<啟動程序> 起飛區已淨空')
    time.sleep(0.5)

    # 回報起飛前檢點完成
    print('<啟動程序> 啟動前檢查程序：完成')
    time.sleep(0.5)

    # 任務開始確認
    go_or_nogo = 'N'
    while go_or_nogo != 'Y':
        go_or_nogo = input('<啟動程序> 請求任務開始(Y/N) => ')

    # 回報任務執行開始
    print('<啟動程序> TDK飛行任務，開始!')
    time.sleep(1)


# ========== 起飛程序 =================================================
# 執行模式切換、無人機解鎖，並由高度控制器將無人機送至目標飛行高度
def take_off(target_alt=1.0, color=''):
    """起飛指令：輸入目標高度、追蹤色塊(程序佔用)"""
    global TARGET_ALT, MODE
    # 切換至無衛星導航模式
    uav.mode = VehicleMode("GUIDED_NOGPS")
    while not uav.mode.name == 'GUIDED_NOGPS':
        time.sleep(0.5)
    print('<啟動程序> 無人機已切換至無衛星導航模式')

    # 解鎖無人機
    uav.arm(wait=True)
    print('<啟動程序> 無人機已解鎖')

    # 設定飛行模式
    mode_pos(color)

    # 設定目標高度
    print('<飛行命令> 起飛至目標高度 %.2f m' % target_alt)
    TARGET_ALT = target_alt

    # 等待抵達目標高度
    i = 0
    while ACTUAL_ALT < TARGET_ALT * (TAKEOFF_SETTLE / 100):
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 起飛中，目前高度 %.2f m，距離目標高度尚有 %.2f m' %
                  (ACTUAL_ALT, TARGET_ALT - ACTUAL_ALT))
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報抵達目標高度
    print('<飛行狀態> 已抵達目標高度 %.2f m' % TARGET_ALT)


# ========== 懸停程序 =================================================
def hover(duration=1, color=''):
    """懸停指令：輸入懸停時長(s)、追蹤色塊(程序佔用)"""
    # 設定飛行命令
    duration = int(duration)
    mode_pos(color)

    print('<飛行命令> 懸停於 %.2f m，維持 %d sec' % (TARGET_ALT, duration))

    # 等待時間到達
    i = 0
    for x in range(duration * 10):
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 懸停中，目前高度 %.2f m，剩餘時間 %d sec' % (ACTUAL_ALT, duration - x / 10))
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報懸停結束
    print('<飛行狀態> 懸停狀態結束，目前高度 %.2f m' % ACTUAL_ALT)


# ========== 降落程序 =================================================
def landing(color=''):
    """降落指令：輸入追蹤色塊，等待降落完成(程序佔用)"""
    global TARGET_ALT, MODE
    # 設定目標高度
    print('<飛行命令> 位置已確認，開始降落至地面')
    TARGET_ALT = 0

    # 設定定位顏色
    mode_pos(color)

    # 位置定位
    print('<飛行命令> 移動至降落點上方')
    i = 0
    while 1:
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 降落前定位中')
            flight_info()

        # 判斷是否已抵達中心
        if -LANDING_POS_LIMIT < ACTUAL_POS[0] < LANDING_POS_LIMIT:
            if -LANDING_POS_LIMIT < ACTUAL_POS[1] < LANDING_POS_LIMIT:
                break

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 等待抵達目標高度
    print('<飛行命令> 位置已確認，開始降落至地面')
    while ACTUAL_ALT > LANDING_CUTOFF:
        # 顯示飛行資訊
        if i % 5 == 0:
            print('<飛行狀態> 降落中，目前高度 %.2f m' % ACTUAL_ALT)
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        # 負值以確保不會著地後彈升
        if ACTUAL_ALT < 0.5:
            TARGET_ALT = -1

        time.sleep(0.1)
        i += 1

    # 回報降落成功
    print('<飛行狀態> 已降落地面')


# ========== 設定為定位模式 ============================================
# 切換至定位模式，並輸入目標顏色與追蹤位置來定位目標色塊
def mode_pos(color='', target=(0, 0)):
    """定位指令：輸入追蹤色塊、目標座標(相對)進行色塊定位(切換指令)"""
    global MODE, POS_COLOR, TARGET_POS
    MODE = 'POS'                # 模式旗標切為定位
    POS_COLOR = color           # 定位目標顏色設置
    TARGET_POS = list(target)
    # 回報模式啟用狀態
    print('<飛行狀態> 切換至定位模式，顏色追蹤：%s\n' % POS_COLOR)


# ========== 設定為循跡模式 ===========================================
# 切換至循跡模式，輸入飛行速度做控制
def mode_line(spd_type=0, flight_speed=0.0):
    """循跡指令：輸入飛行速度控制模式與前進速率或前進俯仰角(切換指令)"""
    global MODE, SPD_MODE, TARGET_SPD, SPD_OL_ANGLE
    MODE = 'LINE'       # 模式旗標切為循跡
    SPD_MODE = spd_type

    # 設定目標
    # 設定飛行速度
    if SPD_MODE:
        TARGET_SPD = flight_speed
        print('<飛行狀態> 切換至閉路循跡模式，目標飛行速度 %.2f m/s，開始往前飛行\n' % TARGET_SPD)
    # 設定飛行作用俯仰角
    else:
        SPD_OL_ANGLE = flight_speed
        print('<飛行狀態> 切換至開路循跡模式，目標飛行速度 %d degree，開始往前飛行\n' % TARGET_SPD)


# ========== 模式重設 =================================================
def mode_reset():
    """模式重設指令：清除所有模式旗標與參數(切換指令)"""
    global MODE, SPD_MODE, TARGET_SPD, SPD_OL_ANGLE, POS_COLOR, TARGET_POS
    MODE = ''                   # 模式旗標切為循跡
    SPD_MODE = 0
    TARGET_SPD = 0
    SPD_OL_ANGLE = 0
    POS_COLOR = ''
    TARGET_POS = (0.0, 0.0)


# ========= 顏色抵達偵測 ==============================================
# 條件式，佔用目前程序，直到偵測抵達目標顏色為止，需在循跡模式中使用
# 為避免發生離開後瞬間偵測抵達情形發生，故將面積閥值拉高作預防
def wait_arrive(color):
    """等待抵達指令：於循線模式下等待抵達目標色塊(程序佔用)"""
    # 初始化顏色面積
    img_area = 0

    # 色塊面積偵測
    i = 0
    while img_area < AREA_ARRIVE_LIMIT and MODE == 'LINE':
        # 更新面積數值
        x_raw, y_raw, img_area = image_moment(color)

        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報偵測到顏色
    print('<飛行狀態> 抵達%s色塊上方\n' % color)


# ========== 顏色離開偵測 =============================================
# 條件式，佔用目前程序，直到偵測到離開目標色塊為止，需在循跡模式中使用
def wait_leave(color):
    """等待離開指令：於循線模式下等待離開目標色塊(程序佔用)"""
    # 初始化顏色面積
    img_area = 100000

    # 色塊面積偵測
    i = 0
    while img_area > AREA_LEAVE_LIMIT and MODE == 'LINE':
        # 更新面積數值
        x_raw, y_raw, img_area = image_moment(color)

        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報離開的顏色
    print('<飛行狀態> 已離開%s色塊上方\n' % color)


# ========== T字路口偵測 =============================================
def wait_t_intersection():
    """等待抵達指令：於循跡模式下等待抵達T字路口(程序佔用)"""
    global ANG_T_TURN
    # 一般循線段
    i = 0
    while t_intersection_detect() == 'NaN':
        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 判斷T字路口方向
    right = 0
    left = 0
    while 1:
        # 更新判斷物件
        t_go = t_intersection_detect()
        if t_go == 'right':
            right += 1
        elif t_go == 'left':
            left += 1

        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 判斷結果
        if right == T_JUDGE_TIMES:
            ANG_T_TURN = 'right'
            print('<飛行狀態> 偵測到右側T轉角')
            break
        elif left == T_JUDGE_TIMES:
            ANG_T_TURN = 'left'
            print('<飛行狀態> 偵測到左側T轉角')
            break

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # T字路口轉彎中，等待復歸
    while t_intersection_detect() != 'straight' or 'NaN':
        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 復歸至正常模式
    ANG_T_TURN = ''
    print('<飛行狀態> 已離開T字路口區')


# ========== 號誌轉向控制 =============================================
# 抵達號誌區後，進行燈號等待，並於偵測到顏色變化後做紀錄，並作相對應轉向
def turn():
    """號誌等待與轉彎指令：等待燈號變色、記錄任務顏色後轉至對應方向(程序佔用)"""
    global TARGET_YAW_RATE, MISSION_COLOR, POS_COLOR
    # 轉彎等待與執行迴圈
    i = 0
    while 1:        # 藍右，綠左
        # 讀取顏色面積值做判斷
        block_x, block_y, image_area_blue = image_moment('blue')
        block_x, block_y, image_area_green = image_moment('green')

        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 讀取偏航角
        yaw_start = math.degrees(uav.attitude.yaw)
        yaw_current = math.degrees(uav.attitude.yaw)

        # 若偵測到藍色
        if i > 50:  # image_area_blue >= TURN_AREA_LIMIT:
            # 記錄任務顏色並更改定位顏色
            MISSION_COLOR = 'red'
            POS_COLOR = MISSION_COLOR
            print('<飛行狀態> 偵測到藍色為任務顏色')

            # 右轉
            print('<飛行命令> 往右旋轉90度\n')
            TARGET_YAW_RATE = TURN_YAW_RATE
            while math.fabs(yaw_current - yaw_start) < 90 * (TURN_SETTLE / 100):
                # 更新數值
                yaw_current = math.degrees(uav.attitude.yaw)

                # 顯示飛行資訊
                if i % 5 == 0:
                    print('<飛行狀態> 旋轉中，已轉 %d 度' % math.fabs(yaw_current - yaw_start))
                    flight_info()

                time.sleep(0.1)
                i += 1

                # 急停跳出
                if not STATE_THREAD:
                    break

            # 結束旋轉
            TARGET_YAW_RATE = 0
            break

        # 若偵測到綠色
        elif image_area_green >= TURN_AREA_LIMIT:
            # 記錄任務顏色並更改定位顏色
            MISSION_COLOR = 'green'
            POS_COLOR = MISSION_COLOR
            print('<飛行狀態> 偵測到綠色為任務顏色')

            # 左轉
            print('<飛行命令> 往左旋轉90度\n')
            TARGET_YAW_RATE = -TURN_YAW_RATE
            while math.fabs(yaw_current - yaw_start) < 90 * (TURN_SETTLE / 100):
                # 更新數值
                yaw_current = math.degrees(uav.attitude.yaw)

                # 顯示飛行資訊
                if i % 5 == 0:
                    print('<飛行狀態> 旋轉中，已轉 %d 度' % math.fabs(yaw_current - yaw_start))
                    flight_info()

                time.sleep(0.1)
                i += 1

                # 急停跳出
                if not STATE_THREAD:
                    break

            # 結束旋轉
            TARGET_YAW_RATE = 0
            break

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 回報轉向完成
    print('<飛行狀態> 轉向完成\n')


# ========== 空投程序 ================================================
# 於投擲區上方做定位穩定，並執行空投程序
def drop():
    global ANG_AFTER_DROP
    # 切換定位模式
    mode_pos('red')

    # 定位迴圈，等待移至空投點上方
    print('<飛行命令> 移動至空投點上方')
    i = 0
    while 1:
        # 取得目前型心位置
        block_x, block_y, image_area = image_moment(MISSION_COLOR)

        # 顯示飛行資訊
        if i % 5 == 0:
            flight_info()

        # 判斷是否已抵達中心
        if -DROP_POS_LIMIT_X < block_x < DROP_POS_LIMIT_X:
            if -DROP_POS_LIMIT_Y < block_y < DROP_POS_LIMIT_Y:
                break

        # 急停跳出
        if not STATE_THREAD:
            break

        time.sleep(0.1)
        i += 1

    # 打開沙包盒空投
    print('<飛行命令> 位置已確認，執行空投\n')
    box_open()
    time.sleep(2)

    # 回報空投成功
    print('<飛行狀態> 沙包已成功空投\n')
    ANG_AFTER_DROP = True

    # 關閉沙包盒
    box_close()


# ========== 飛行資料 ==================================================
def flight_info():
    """飛行資料呈現"""
    print('<高度狀態> 總高：%.2f m，超音波：%.2f m，IMU：%.2f m，輸出THR：%.1f%%' %
          (ACTUAL_ALT, ACTUAL_ALT_SONAR, ACTUAL_ALT_IMU, TARGET_THRUST*100))

    if MODE == 'POS':
        print('<飛行狀態> 定位中，相對位置(%.2f , %.2f)px，色塊面積：%d px' %
              (ACTUAL_POS[0], ACTUAL_POS[1], ACTUAL_AREA))
        print('<飛行狀態> 定位中，相對速度(%.2f , %.2f)px/s' %
              (ACTUAL_VEL[0], ACTUAL_VEL[1]))
    if MODE == 'LINE':
        print('<飛行狀態> 循跡中，差角 %.2f 度，角速度 %.2f 度/秒' % (ACTUAL_HDG, TARGET_YAW_RATE))
        print('<飛行狀態> 循跡中，線位置 %.2f px，線速度 %.2f ps/s，線面積 %d px' %
              (ACTUAL_LINE_POS, ACTUAL_LINE_VEL, ACTUAL_LINE_AREA))
        print('<飛行狀態> 循跡中，飛行速度 %.2f m/s，俯仰角 %.2f 度' %
              (ACTUAL_SPD, math.degrees(uav.attitude.pitch)))

    print('<目標姿態> Roll=%.2f, Pitch=%.2f, Yaw=%.2f, Yaw Rate=%.2f' %
          (TARGET_ROLL, TARGET_PITCH, TARGET_YAW, TARGET_YAW_RATE))
    print('<即時姿態> Roll=%.2f, Pitch=%.2f, Yaw=%.2f\n' %
          (math.degrees(uav.attitude.roll), math.degrees(uav.attitude.pitch),
           math.degrees(uav.attitude.yaw)))


# ====================================================================
# ========== 前景主函式區塊 ============================================
# ====================================================================
# 任務主函式
def mission():
    global STATE_THREAD, STATE_UAV, MISSION_COLOR
    print('<啟動程序> 任務程序：啟動')
    # 啟動多線程
    thread_activate()

    # 選擇任務模式
    mission_mode = int(input('請輸入任務模式1/2/3=>'))

    # 程式標準式(勿呼叫該任務)
    if mission == 255:
        take_off(target_alt=1, color='red')
        hover(duration=10, color='blue')
        landing(color='green')
        wait_leave(color='red')
        wait_arrive(color='blue')
        wait_t_intersection()
        turn()
        drop()
        mode_pos(color='red', target=(0, 0))
        mode_line(0, 0.2)
        mode_reset()

    # 飛行任務1-定高測試(通過)
    if mission_mode == 1:
        pre_check()         # 執行起飛前檢查
        take_off(1)         # 起飛至任務高度
        hover(10)           # 懸停
        landing()           # 執行降落與關閉程序

    # 飛行任務2-定高&定位測試(旋轉定位與投擲再測)
    if mission_mode == 2:
        pre_check()         # 執行起飛前檢查
        take_off(1, 'red')  # 起飛至任務高度
        turn()
        hover(5, 'red')     # 懸停
        drop()
        landing('red')      # 執行降落與關閉程序

    # 飛行任務3-定高循跡(分採HSV與GRAY取圖法做比較)
    if mission_mode == 3:
        pre_check()         # 執行起飛前檢查
        take_off(1, 'red')  # 起飛至任務高度
        mode_line(0, 3)     # 循跡模式
        wait_leave('red')
        wait_arrive('blue')
        mode_pos('blue')
        hover(5, 'blue')
        landing('blue')     # 執行降落與關閉程序

    # 飛行任務3-起飛區完整測試
    if mission_mode == 4:
        pre_check()             # 執行起飛前檢查
        take_off(1, 'black')    # 起飛至任務高度
        mode_line(0, 1.5)       # 循跡模式
        wait_arrive('red')      # 等待抵達號誌區
        mode_pos('red')         # 紅燈定位保持
        turn()
        hover(5, 'red')
        landing('red')          # 執行降落與關閉程序

    # 飛行任務4-循線A區測試
    if mission_mode == 5:
        pre_check()
        take_off(1, 'black')
        mode_line(0, 1.5)       # 經過Ｌ轉角
        wait_arrive('blue')     # 色塊抓取與懸停
        mode_pos('blue')
        hover(5, 'blue')
        # wait_arrive('green')    # 色塊抓取與懸停
        # mode_pos('green')
        # hover(5, 'green')
        landing('red')

    # 飛行任務5-循線中色塊空投測試
    if mission_mode == 6:
        pre_check()
        take_off(1, 'black')
        mode_line(0, 2)
        wait_arrive('green')
        mode_pos('green')
        drop()
        mode_line(0, 3)
        wait_leave('green')
        wait_arrive('red')
        hover(5)
        landing()

    # 完整版(未整理)
    if mission_mode == 10:
        # 起飛與紅燈區
        pre_check()                 # 執行起飛前檢查
        take_off(1, 'black')        # 起飛至任務高度
        mode_line(0, 2)             # 循跡模式
        wait_arrive('red')          # 等待抵達號誌區
        mode_pos('red')             # 紅燈定位保持

        # 循線A區
        turn()                      # 轉彎等待模式
        mode_line(1, 0.2)           # 往前循跡飛行
        wait_leave(MISSION_COLOR)   # 離開號誌區

        # 投擲與循線B區
        wait_arrive(MISSION_COLOR)  # 偵測抵達投擲色塊上方
        drop()                      # 空投程序
        mode_line(1, 0.2)           # 往前循跡飛行
        wait_leave(MISSION_COLOR)   # 偵測已離開投擲區

        # T字路口與降落區
        wait_t_intersection()       # 等待與執行T字路口轉彎
        wait_arrive(MISSION_COLOR)  # 偵測抵達降落區
        landing()                   # 執行降落

    # 關閉多線程
    STATE_THREAD = False
    print('<關閉程序> 主程序：關閉')

    # 清空GPIO設定
    GPIO.cleanup()
    print('<關閉程序> GPIO已關閉')

    # 結束與無人機連線
    uav.close()
    STATE_UAV = False
    print('<關閉程序> 無人機連線：關閉')

    # 輸出圖表
    data_sheet()
    pos_ctrl_data()
    vel_ctrl_data()
    # hdg_ctrl_data()
    # alt_ctrl_data()
    # alt_sen_data()
    # spd_ctrl_cl_data()
    # spd_ctrl_ol_data()


# ========== GUI介面區塊 ================================================
# 顏色參數
color_bg = '#eff8fd'

# 圖形化介面建構
flight_monitor = tk.Tk()
flight_monitor.title('TDK 24th 飛行組 C08 台灣大艦隊 飛行控制介面')
flight_monitor.geometry('1280x700')
flight_monitor.configure(background=color_bg)

# ===========================================================================
# ========== Row 1 標題區 ====================================================
# ===========================================================================
# 80*1920
row_1_frame = tk.Frame(flight_monitor, height=80, width=1280, bg=color_bg)
row_1_frame.pack_propagate(0)
row_1_frame.pack(side=tk.TOP)
head_label = tk.Label(row_1_frame, font=('', 30), fg='#004085', bg=color_bg,
                      text='\n----- TDK 24th 飛行組 C08 台灣大艦隊 飛行控制介面 -----\n')
head_label.pack(side=tk.TOP)


# ===========================================================================
# ========== Row 2 系統狀態列 =================================================
# ===========================================================================
# 參數
GUI_ROW_2_H = 130
GUI_ROW_2_W = 115
# 方塊佈局
row_2_frame = tk.Frame(flight_monitor, height=150, width=1280, bg=color_bg)
row_2_frame.pack_propagate(0)
row_2_frame.pack(side=tk.TOP)

# ========== Row 2 標題 ==========
# 系統與控制器狀態
state_frame = tk.Frame(row_2_frame, bg=color_bg)
state_frame.pack(side=tk.TOP)
state_label = tk.Label(state_frame, font=('', 25), bg=color_bg,
                       fg='#007dd1', text='----- 系統與控制器狀態 -----')
state_label.pack(side=tk.TOP)

# 系統狀態
state_sys_frame = tk.Frame(state_frame, height=GUI_ROW_2_H, width=GUI_ROW_2_W, bg=color_bg)
state_sys_frame.pack_propagate(0)
state_sys_frame.pack(side=tk.LEFT)
state_sys_label = tk.Label(state_sys_frame, font=('', 25), text='SYS\n系統', bg=color_bg)
state_sys_label.pack(side=tk.TOP)
state_sys = tk.Label(state_sys_frame, fg='red', font=('', 35), text='[OFF]', bg=color_bg)
state_sys.pack(side=tk.TOP)



# 執行程式
flight_monitor.mainloop()
